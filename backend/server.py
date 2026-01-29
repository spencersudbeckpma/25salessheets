from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta, date as date_class
import bcrypt
import jwt
import string
import random
import base64
from pytz import timezone as pytz_timezone

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-super-secret-jwt-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_DAYS = 30

# Security
security = HTTPBearer()

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Models
class Team(BaseModel):
    """Team model for multi-tenancy"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    settings: Optional[Dict[str, Any]] = {}
    # Branding fields
    branding: Optional[Dict[str, Any]] = Field(default_factory=lambda: {
        "logo_url": None,
        "primary_color": "#1e40af",  # Default blue
        "accent_color": "#3b82f6",   # Default lighter blue
        "display_name": None,
        "tagline": None
    })
    # Feature flags - controls which tabs/features are visible to team members
    features: Optional[Dict[str, bool]] = Field(default_factory=lambda: {
        "activity": True,
        "stats": True,
        "team_view": True,
        "suitability": True,
        "pma_bonuses": True,
        "docusphere": True,
        "leaderboard": True,
        "analytics": True,
        "reports": True,
        "team_mgmt": True,
        "recruiting": True,
        "interviews": True
    })

# Default features for new teams
DEFAULT_TEAM_FEATURES = {
    "activity": True,
    "stats": True,
    "team_view": True,
    "suitability": True,
    "pma_bonuses": True,
    "docusphere": True,
    "leaderboard": True,
    "analytics": True,
    "reports": True,
    "team_mgmt": True,
    "recruiting": False,  # Disabled by default - super_admin enables per team
    "interviews": True,
    "fact_finder": True,
    "sna": True,  # SNA tracker sub-tab
    "npa": True,  # NPA tracker sub-tab
    "new_faces": True  # New Faces sub-tab
}

# Default role-based tab overrides (empty = no restrictions)
DEFAULT_ROLE_TAB_OVERRIDES = {
    "agent": {"hidden_tabs": []},
    "district_manager": {"hidden_tabs": []},
    "regional_manager": {"hidden_tabs": []}
}

# Default team UI settings
DEFAULT_TEAM_UI_SETTINGS = {
    "default_landing_tab": "activity",
    "default_leaderboard_period": "weekly"
}

# Default team view settings (Phase 2)
DEFAULT_TEAM_VIEW_SETTINGS = {
    # KPI cards configuration - order determines display order
    # Labels must match live dashboard (StatsView.jsx, ActivityInput.jsx)
    "kpi_cards": [
        {"id": "contacts", "label": "Contacts", "enabled": True},
        {"id": "appointments", "label": "Appointments", "enabled": True},
        {"id": "presentations", "label": "Presentations", "enabled": True},
        {"id": "referrals", "label": "Referrals", "enabled": True},
        {"id": "testimonials", "label": "Testimonials", "enabled": True},
        {"id": "sales", "label": "Sales", "enabled": True},
        {"id": "new_face_sold", "label": "New Face Sold", "enabled": True},
        {"id": "premium", "label": "Total Premium", "enabled": True}
    ],
    # Sub-tab visibility (enforced server-side)
    "subtabs": {
        "new_faces": True,
        "sna": True,
        "npa": True
    }
}

async def get_team_view_settings(team: dict) -> dict:
    """Get team view settings with defaults merged"""
    if not team:
        return DEFAULT_TEAM_VIEW_SETTINGS.copy()
    
    team_settings = team.get('team_settings', {})
    views = team_settings.get('views', {})
    
    # Merge KPI cards (preserve order from team config if exists)
    kpi_cards = views.get('kpi_cards', DEFAULT_TEAM_VIEW_SETTINGS['kpi_cards'])
    
    # Merge subtabs
    subtabs = {**DEFAULT_TEAM_VIEW_SETTINGS['subtabs'], **views.get('subtabs', {})}
    
    return {
        "kpi_cards": kpi_cards,
        "subtabs": subtabs
    }

async def check_subtab_access(current_user: dict, subtab_name: str):
    """
    Check if user's team has access to a specific sub-tab.
    Enforced server-side - returns 403 if disabled.
    Super admins always have access.
    """
    if current_user.get('role') == 'super_admin':
        return True
    
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail=f"Access denied: {subtab_name} is not available")
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=403, detail=f"Access denied: team not found")
    
    view_settings = await get_team_view_settings(team)
    subtabs = view_settings.get('subtabs', {})
    
    if not subtabs.get(subtab_name, True):
        raise HTTPException(status_code=403, detail=f"Access denied: {subtab_name} is not enabled for your team")
    
    return True

async def get_effective_features(user: dict, team: dict = None) -> dict:
    """
    Compute effective features for a user based on:
    1. Team feature flags
    2. Role-based tab overrides
    3. User's role
    
    Returns the feature set this specific user should see.
    State managers and super_admins are NOT subject to role overrides.
    """
    role = user.get('role', 'agent')
    
    # Super admins see ALL features
    if role == 'super_admin':
        return {k: True for k in DEFAULT_TEAM_FEATURES.keys()}
    
    # State managers see full team configuration (no role overrides)
    # Start with team features merged with defaults
    if team:
        team_features = team.get('features', {})
        effective = {**DEFAULT_TEAM_FEATURES, **team_features}
    else:
        effective = DEFAULT_TEAM_FEATURES.copy()
    
    # State managers are not subject to role overrides
    if role == 'state_manager':
        return effective
    
    # Apply role-based overrides for agents, DMs, and RMs
    if team:
        role_overrides = team.get('role_tab_overrides', DEFAULT_ROLE_TAB_OVERRIDES)
        role_config = role_overrides.get(role, {})
        hidden_tabs = role_config.get('hidden_tabs', [])
        
        # Disable hidden tabs for this role
        for tab in hidden_tabs:
            if tab in effective:
                effective[tab] = False
    
    return effective

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    role: str  # super_admin, state_manager, regional_manager, district_manager, agent
    team_id: Optional[str] = None  # Required for all non-admin users
    manager_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str
    team_id: Optional[str] = None
    manager_id: Optional[str] = None
    invite_code: Optional[str] = None

class UserLogin(BaseModel):
    email: str  # Changed from EmailStr to accept username or email
    password: str

class PasswordChangeRequest(BaseModel):
    current_password: str
    new_password: str

class PasswordResetRequest(BaseModel):
    user_id: str
    new_password: str

class ForgotPasswordRequest(BaseModel):
    email: str

class Activity(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    team_id: Optional[str] = None  # Multi-tenancy support
    date: str  # YYYY-MM-DD format
    contacts: float = 0.0
    appointments: float = 0.0
    presentations: float = 0.0
    referrals: int = 0
    testimonials: int = 0
    apps: int = 0
    sales: int = 0
    new_face_sold: float = 0.0
    bankers_premium: float = 0.0
    premium: float = 0.0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    edited_by: Optional[str] = None
    edited_at: Optional[datetime] = None

class ActivityCreate(BaseModel):
    date: str
    contacts: float = 0.0
    appointments: float = 0.0
    presentations: float = 0.0
    referrals: int = 0
    testimonials: int = 0
    apps: int = 0
    sales: int = 0
    new_face_sold: float = 0.0
    bankers_premium: float = 0.0
    premium: float = 0.0

class NewFaceCustomer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    team_id: Optional[str] = None  # Multi-tenancy support
    date: str  # YYYY-MM-DD format
    customer_name: str
    county: str
    policy_amount: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NewFaceCustomerCreate(BaseModel):
    date: str
    customer_name: str
    county: str
    policy_amount: float

class Invite(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invite_code: str
    name: str
    email: EmailStr
    role: str
    manager_id: str
    team_id: Optional[str] = None  # Multi-tenancy support
    status: str = "pending"  # pending, accepted
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InviteCreate(BaseModel):
    name: str
    email: EmailStr
    role: str

# Helper Functions
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_jwt_token(user_id: str, email: str) -> str:
    payload = {
        'user_id': user_id,
        'email': email,
        'exp': datetime.now(timezone.utc) + timedelta(days=JWT_EXPIRATION_DAYS)
    }
    return jwt.encode(payload, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)

def decode_jwt_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def generate_invite_code() -> str:
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    token = credentials.credentials
    payload = decode_jwt_token(token)
    user = await db.users.find_one({"id": payload['user_id']}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    
    # Super admins can access without team assignment
    if user.get('role') == 'super_admin':
        return user
    
    # State managers have admin privileges - allow access even without team for bootstrap
    if user.get('role') == 'state_manager':
        return user
    
    # Regular users MUST have a team assigned
    if not user.get('team_id'):
        raise HTTPException(
            status_code=403, 
            detail="Access denied. Your account has not been assigned to a team yet. Please contact your administrator."
        )
    
    return user

async def get_team_filter(user: dict) -> dict:
    """Get the team filter for queries based on user's team.
    
    IMPORTANT: ALL users (including super_admin) are scoped to their assigned team
    on product pages. Cross-team visibility is ONLY available in Admin endpoints
    where team_id is explicitly passed as a parameter.
    """
    return {"team_id": user.get('team_id')}

def get_team_filter_with_legacy(team_id: str) -> dict:
    """Get STRICT team filter - NO cross-team visibility allowed.
    
    CRITICAL: This filter ONLY matches records that explicitly belong to the given team.
    Records with NULL/missing team_id are EXCLUDED to prevent cross-team leakage.
    
    Use the migration endpoint to backfill team_id on legacy records BEFORE they 
    will appear in team views.
    """
    if team_id:
        # STRICT: Only exact team_id match. No NULL/missing allowed.
        return {"team_id": team_id}
    return {"team_id": {"$exists": False}}  # If no team_id provided, match nothing meaningful

async def get_all_subordinates(user_id: str, team_id: str = None) -> List[str]:
    """Get all subordinates recursively (exclude archived), scoped to team"""
    subordinates = [user_id]
    query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
    if team_id:
        query["team_id"] = team_id
    direct_reports = await db.users.find(query, {"_id": 0}).to_list(1000)
    for report in direct_reports:
        sub_list = await get_all_subordinates(report['id'], team_id)
        subordinates.extend(sub_list)
    return list(set(subordinates))

def require_super_admin(current_user: dict):
    """Check if user is super_admin ONLY - for team management operations"""
    if current_user.get('role') != 'super_admin':
        raise HTTPException(status_code=403, detail="Super admin access required")

async def check_feature_access(current_user: dict, feature_name: str):
    """
    Check if user has access to a specific feature based on:
    1. Team feature flags
    2. Role-based tab overrides
    
    Super admins always have access.
    State managers see full team config.
    Other roles may have restrictions applied.
    
    Returns True if allowed, raises 403 if not.
    """
    # Super admins always have access
    if current_user.get('role') == 'super_admin':
        return True
    
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail=f"Access denied: {feature_name} is not available")
    
    # Get team config
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if team is None:
        raise HTTPException(status_code=403, detail=f"Access denied: team not found")
    
    # Get effective features for this user (includes role-based overrides)
    effective_features = await get_effective_features(current_user, team)
    
    if not effective_features.get(feature_name, False):
        raise HTTPException(status_code=403, detail=f"Access denied: {feature_name} is not enabled for your role")
    
    return True

# ==================== ADMIN TEAM MANAGEMENT ====================

class TeamCreate(BaseModel):
    name: str
    settings: Optional[Dict[str, Any]] = {}

class TeamBrandingUpdate(BaseModel):
    logo_url: Optional[str] = None
    primary_color: Optional[str] = None
    accent_color: Optional[str] = None
    display_name: Optional[str] = None
    tagline: Optional[str] = None

class UserTeamAssignment(BaseModel):
    user_id: str
    team_id: str
    role: Optional[str] = None
    manager_id: Optional[str] = None

@api_router.get("/admin/activities-team-diagnostic")
async def activities_team_diagnostic(current_user: dict = Depends(get_current_user)):
    """Diagnose activities team_id distribution (super_admin only)"""
    require_super_admin(current_user)
    
    # Get all teams
    teams = await db.teams.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    team_names = {t["id"]: t["name"] for t in teams}
    
    # Count activities by team_id
    pipeline = [
        {"$group": {"_id": "$team_id", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    team_counts = await db.activities.aggregate(pipeline).to_list(100)
    
    # Count activities for January 2026 by team
    jan_pipeline = [
        {"$match": {"date": {"$gte": "2026-01-01", "$lte": "2026-01-31"}}},
        {"$group": {"_id": "$team_id", "count": {"$sum": 1}, "total_presentations": {"$sum": "$presentations"}}},
        {"$sort": {"count": -1}}
    ]
    jan_counts = await db.activities.aggregate(jan_pipeline).to_list(100)
    
    # Check Steve Ahlers specifically
    steve = await db.users.find_one({"name": {"$regex": "Steve Ahlers", "$options": "i"}}, {"_id": 0})
    steve_activities = []
    if steve:
        steve_acts = await db.activities.find(
            {"user_id": steve["id"], "date": {"$gte": "2026-01-01"}}, 
            {"_id": 0, "date": 1, "team_id": 1, "presentations": 1}
        ).sort("date", 1).to_list(100)
        steve_activities = steve_acts
    
    return {
        "total_activities": await db.activities.count_documents({}),
        "activities_by_team": [
            {
                "team_id": tc["_id"],
                "team_name": team_names.get(tc["_id"], "NO TEAM" if tc["_id"] is None else "UNKNOWN"),
                "count": tc["count"]
            }
            for tc in team_counts
        ],
        "january_2026_by_team": [
            {
                "team_id": tc["_id"],
                "team_name": team_names.get(tc["_id"], "NO TEAM" if tc["_id"] is None else "UNKNOWN"),
                "count": tc["count"],
                "presentations": tc["total_presentations"]
            }
            for tc in jan_counts
        ],
        "steve_ahlers": {
            "user_id": steve["id"] if steve else None,
            "team_id": steve.get("team_id") if steve else None,
            "january_activities": steve_activities
        }
    }

@api_router.get("/admin/recruiting-diagnostic")
async def recruiting_diagnostic(current_user: dict = Depends(get_current_user)):
    """Diagnose recruiting/pipeline data - shows STRICT team filtering status.
    
    CRITICAL: No cross-team visibility is allowed. This diagnostic confirms
    that ONLY records with exact team_id match are included.
    """
    require_super_admin(current_user)
    
    user_team_id = current_user.get('team_id')
    
    # Get all teams
    teams = await db.teams.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    team_names = {t["id"]: t["name"] for t in teams}
    
    # Get all users in the current user's team
    users_in_team = await db.users.find(
        {"team_id": user_team_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
        {"_id": 0, "id": 1, "name": 1, "role": 1}
    ).to_list(10000)
    users_in_scope_ids = [u["id"] for u in users_in_team]
    users_by_role = {}
    for u in users_in_team:
        role = u.get("role", "unknown")
        if role not in users_by_role:
            users_by_role[role] = 0
        users_by_role[role] += 1
    
    # Get ALL recruits
    all_recruits = await db.recruits.find({}, {"_id": 0}).to_list(10000)
    
    # Build user lookup
    all_users = await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "role": 1, "team_id": 1}).to_list(10000)
    user_lookup = {u["id"]: u for u in all_users}
    
    # Categorize recruits with STRICT filtering
    recruits_exact_match = []  # team_id == user_team_id (INCLUDED)
    recruits_null_missing = []  # team_id is NULL/missing (EXCLUDED)
    recruits_other_team = []  # team_id belongs to another team (EXCLUDED)
    
    for r in all_recruits:
        recruit_team_id = r.get("team_id")
        created_by = r.get("created_by")
        creator = user_lookup.get(created_by, {})
        
        recruit_info = {
            "name": r.get("name"),
            "recruit_team_id": recruit_team_id if recruit_team_id else "NULL/MISSING",
            "recruit_team_name": team_names.get(recruit_team_id, "UNKNOWN") if recruit_team_id else "N/A",
            "creator_name": creator.get("name", "Unknown"),
            "creator_role": creator.get("role", "unknown"),
            "creator_team_id": creator.get("team_id", "NULL/MISSING")
        }
        
        if recruit_team_id == user_team_id:
            recruits_exact_match.append(recruit_info)
        elif recruit_team_id is None or "team_id" not in r:
            recruits_null_missing.append(recruit_info)
        else:
            recruits_other_team.append(recruit_info)
    
    # The STRICT query that will be used
    strict_query = {"team_id": user_team_id} if user_team_id else {}
    matching_count = await db.recruits.count_documents(strict_query)
    
    return {
        "CRITICAL_NOTICE": "STRICT team filtering enforced. NO cross-team visibility.",
        "current_user": {
            "team_id": user_team_id,
            "team_name": team_names.get(user_team_id, "UNKNOWN")
        },
        "strict_query_applied": str(strict_query),
        "users_in_my_team": {
            "total": len(users_in_scope_ids),
            "by_role": users_by_role
        },
        "recruits_analysis": {
            "total_in_database": len(all_recruits),
            "INCLUDED_exact_team_match": len(recruits_exact_match),
            "EXCLUDED_null_missing_team_id": len(recruits_null_missing),
            "EXCLUDED_other_team": len(recruits_other_team),
            "query_result_count": matching_count
        },
        "cross_team_check": {
            "other_teams_excluded": len(recruits_other_team),
            "NO_CROSS_TEAM_LEAKAGE": len(recruits_other_team) > 0 and matching_count == len(recruits_exact_match)
        },
        "sample_INCLUDED_recruits": recruits_exact_match[:5],
        "sample_EXCLUDED_null_missing": recruits_null_missing[:5],
        "sample_EXCLUDED_other_team": recruits_other_team[:5],
        "action_required": {
            "null_missing_recruits": len(recruits_null_missing),
            "fix": "Run POST /api/admin/migrate-recruits-team-id to backfill team_id on legacy records"
        }
    }

@api_router.post("/admin/migrate-recruits-team-id")
async def migrate_recruits_team_id(current_user: dict = Depends(get_current_user)):
    """
    Backfill missing team_id on recruits based on created_by user's team_id.
    GUARDRAILS:
    - Only updates recruits where team_id is NULL or missing
    - Sets recruit.team_id = creator_user.team_id
    - Does NOT modify any other fields
    - If created_by user not found, leaves recruit unchanged
    (super_admin only)
    """
    require_super_admin(current_user)
    
    # Build user_id -> team_id mapping
    users = await db.users.find({}, {"_id": 0, "id": 1, "team_id": 1}).to_list(10000)
    user_team_map = {u["id"]: u.get("team_id") for u in users}
    
    # Find recruits with missing team_id
    recruits_to_update = await db.recruits.find(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"_id": 1, "created_by": 1, "name": 1}
    ).to_list(100000)
    
    total_scanned = len(recruits_to_update)
    total_updated = 0
    total_skipped_no_user = 0
    total_skipped_user_no_team = 0
    
    for recruit in recruits_to_update:
        created_by = recruit.get("created_by")
        
        if not created_by or created_by not in user_team_map:
            total_skipped_no_user += 1
            continue
        
        user_team_id = user_team_map[created_by]
        
        if not user_team_id:
            total_skipped_user_no_team += 1
            continue
        
        await db.recruits.update_one(
            {"_id": recruit["_id"]},
            {"$set": {"team_id": user_team_id}}
        )
        total_updated += 1
    
    return {
        "migration_report": {
            "total_recruits_in_db": await db.recruits.count_documents({}),
            "total_scanned_missing_team_id": total_scanned,
            "total_updated": total_updated,
            "total_skipped_creator_not_found": total_skipped_no_user,
            "total_skipped_creator_has_no_team": total_skipped_user_no_team
        }
    }

@api_router.post("/admin/migrate-interviews-team-id")
async def migrate_interviews_team_id(current_user: dict = Depends(get_current_user)):
    """
    Backfill missing team_id on interviews based on interviewer's team_id.
    GUARDRAILS:
    - Only updates interviews where team_id is NULL or missing
    - Sets interview.team_id = interviewer_user.team_id
    - Does NOT modify any other fields
    - If interviewer_id not found, leaves interview unchanged
    (super_admin only)
    """
    require_super_admin(current_user)
    
    # Build user_id -> team_id mapping
    users = await db.users.find({}, {"_id": 0, "id": 1, "team_id": 1}).to_list(10000)
    user_team_map = {u["id"]: u.get("team_id") for u in users}
    
    # Find interviews with missing team_id
    interviews_to_update = await db.interviews.find(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"_id": 1, "interviewer_id": 1, "candidate_name": 1}
    ).to_list(100000)
    
    total_scanned = len(interviews_to_update)
    total_updated = 0
    total_skipped_no_user = 0
    total_skipped_user_no_team = 0
    
    for interview in interviews_to_update:
        interviewer_id = interview.get("interviewer_id")
        
        if not interviewer_id or interviewer_id not in user_team_map:
            total_skipped_no_user += 1
            continue
        
        user_team_id = user_team_map[interviewer_id]
        
        if not user_team_id:
            total_skipped_user_no_team += 1
            continue
        
        await db.interviews.update_one(
            {"_id": interview["_id"]},
            {"$set": {"team_id": user_team_id}}
        )
        total_updated += 1
    
    return {
        "migration_report": {
            "collection": "interviews",
            "total_in_db": await db.interviews.count_documents({}),
            "total_scanned_missing_team_id": total_scanned,
            "total_updated": total_updated,
            "total_skipped_interviewer_not_found": total_skipped_no_user,
            "total_skipped_interviewer_has_no_team": total_skipped_user_no_team
        }
    }

@api_router.post("/admin/migrate-new-face-customers-team-id")
async def migrate_new_face_customers_team_id(current_user: dict = Depends(get_current_user)):
    """
    Backfill missing team_id on new_face_customers based on user_id's team_id.
    GUARDRAILS:
    - Only updates records where team_id is NULL or missing
    - Sets new_face_customer.team_id = user.team_id
    - Does NOT modify any other fields
    - If user_id not found, leaves record unchanged
    (super_admin only)
    """
    require_super_admin(current_user)
    
    # Build user_id -> team_id mapping
    users = await db.users.find({}, {"_id": 0, "id": 1, "team_id": 1}).to_list(10000)
    user_team_map = {u["id"]: u.get("team_id") for u in users}
    
    # Find new_face_customers with missing team_id
    records_to_update = await db.new_face_customers.find(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"_id": 1, "user_id": 1, "name": 1}
    ).to_list(100000)
    
    total_scanned = len(records_to_update)
    total_updated = 0
    total_skipped_no_user = 0
    total_skipped_user_no_team = 0
    
    for record in records_to_update:
        user_id = record.get("user_id")
        
        if not user_id or user_id not in user_team_map:
            total_skipped_no_user += 1
            continue
        
        user_team_id = user_team_map[user_id]
        
        if not user_team_id:
            total_skipped_user_no_team += 1
            continue
        
        await db.new_face_customers.update_one(
            {"_id": record["_id"]},
            {"$set": {"team_id": user_team_id}}
        )
        total_updated += 1
    
    return {
        "migration_report": {
            "collection": "new_face_customers",
            "total_in_db": await db.new_face_customers.count_documents({}),
            "total_scanned_missing_team_id": total_scanned,
            "total_updated": total_updated,
            "total_skipped_user_not_found": total_skipped_no_user,
            "total_skipped_user_has_no_team": total_skipped_user_no_team
        }
    }

@api_router.post("/admin/migrate-activities-team-id")
async def migrate_activities_team_id(current_user: dict = Depends(get_current_user)):
    """
    Backfill missing team_id on activities based on user's team_id.
    GUARDRAILS:
    - Only updates activities where team_id is NULL or missing
    - Sets activity.team_id = user.team_id
    - Does NOT modify dates, premium, counts, or any other fields
    - If user_id not found, leaves activity unchanged
    (super_admin only)
    """
    require_super_admin(current_user)
    
    # Build user_id -> team_id mapping
    users = await db.users.find({}, {"_id": 0, "id": 1, "team_id": 1}).to_list(10000)
    user_team_map = {u["id"]: u.get("team_id") for u in users}
    
    # Find activities with missing team_id
    activities_to_update = await db.activities.find(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"_id": 1, "user_id": 1}
    ).to_list(100000)
    
    total_scanned = len(activities_to_update)
    total_updated = 0
    total_skipped_no_user = 0
    total_skipped_user_no_team = 0
    unresolved_user_ids = set()
    
    for activity in activities_to_update:
        user_id = activity.get("user_id")
        
        if not user_id or user_id not in user_team_map:
            # User not found - do not guess, leave unchanged
            total_skipped_no_user += 1
            if user_id:
                unresolved_user_ids.add(user_id)
            continue
        
        user_team_id = user_team_map[user_id]
        
        if not user_team_id:
            # User exists but has no team_id - leave unchanged
            total_skipped_user_no_team += 1
            continue
        
        # Update ONLY the team_id field
        await db.activities.update_one(
            {"_id": activity["_id"]},
            {"$set": {"team_id": user_team_id}}
        )
        total_updated += 1
    
    # Count activities that already had team_id (for completeness)
    total_already_had_team = await db.activities.count_documents(
        {"team_id": {"$ne": None, "$exists": True}}
    ) - total_updated
    
    return {
        "migration_report": {
            "total_activities_in_db": await db.activities.count_documents({}),
            "total_scanned_missing_team_id": total_scanned,
            "total_updated": total_updated,
            "total_skipped_already_had_team_id": total_already_had_team,
            "total_skipped_user_not_found": total_skipped_no_user,
            "total_skipped_user_has_no_team": total_skipped_user_no_team,
            "unresolved_user_ids": list(unresolved_user_ids)[:20]  # Limit to first 20
        }
    }

@api_router.post("/admin/migrate-docusphere-team-id")
async def migrate_docusphere_team_id(current_user: dict = Depends(get_current_user)):
    """
    Backfill team_id on DocuSphere folders and documents.
    Assigns all existing content to Team Sudbeck.
    (super_admin only)
    """
    require_super_admin(current_user)
    
    # Get Team Sudbeck ID
    team_sudbeck = await db.teams.find_one({"name": "Team Sudbeck"})
    if not team_sudbeck:
        raise HTTPException(status_code=404, detail="Team Sudbeck not found")
    
    team_sudbeck_id = team_sudbeck["id"]
    
    # Migrate folders
    folders_result = await db.docusphere_folders.update_many(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"$set": {"team_id": team_sudbeck_id}}
    )
    
    # Migrate documents
    docs_result = await db.docusphere_documents.update_many(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"$set": {"team_id": team_sudbeck_id}}
    )
    
    return {
        "message": "DocuSphere migration complete",
        "team_sudbeck_id": team_sudbeck_id,
        "folders_updated": folders_result.modified_count,
        "documents_updated": docs_result.modified_count
    }

@api_router.get("/admin/user-activities-diagnostic/{user_id}")
async def user_activities_diagnostic(user_id: str, current_user: dict = Depends(get_current_user)):
    """Check a specific user's activities (super_admin only)"""
    require_super_admin(current_user)
    
    # Get user info
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get user's activities
    activities = await db.activities.find(
        {"user_id": user_id},
        {"_id": 0, "date": 1, "team_id": 1, "presentations": 1, "premium": 1}
    ).sort("date", -1).to_list(100)
    
    # Count by team_id
    team_counts = {}
    for a in activities:
        tid = a.get("team_id") or "NULL"
        team_counts[tid] = team_counts.get(tid, 0) + 1
    
    return {
        "user": {
            "id": user.get("id"),
            "name": user.get("name"),
            "email": user.get("email"),
            "team_id": user.get("team_id"),
            "role": user.get("role")
        },
        "total_activities": len(activities),
        "activities_by_team_id": team_counts,
        "recent_activities": activities[:10]
    }

@api_router.post("/admin/fix-user-activities-team-id/{user_id}")
async def fix_user_activities_team_id(user_id: str, current_user: dict = Depends(get_current_user)):
    """Fix team_id on a specific user's activities (super_admin only)"""
    require_super_admin(current_user)
    
    # Get user info
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_team_id = user.get("team_id")
    if not user_team_id:
        raise HTTPException(status_code=400, detail="User has no team_id assigned")
    
    # Update activities that have NULL or missing team_id
    result = await db.activities.update_many(
        {"user_id": user_id, "$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"$set": {"team_id": user_team_id}}
    )
    
    return {
        "message": f"Updated activities for user {user.get('name')}",
        "user_id": user_id,
        "user_team_id": user_team_id,
        "activities_updated": result.modified_count
    }

@api_router.get("/admin/diagnose-orphaned-activities")
async def diagnose_orphaned_activities(current_user: dict = Depends(get_current_user)):
    """
    Find activities with NULL team_id and identify which users they belong to.
    Shows whether the user has a team (fixable) or not (needs user assignment first).
    (super_admin only, read-only)
    """
    require_super_admin(current_user)
    
    # Find all activities with NULL/missing team_id
    orphaned = await db.activities.find(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"_id": 0, "id": 1, "user_id": 1, "date": 1, "presentations": 1, "premium": 1}
    ).to_list(1000)
    
    # Group by user_id
    user_activity_counts = {}
    for act in orphaned:
        uid = act.get("user_id", "unknown")
        if uid not in user_activity_counts:
            user_activity_counts[uid] = {"count": 0, "sample_dates": []}
        user_activity_counts[uid]["count"] += 1
        if len(user_activity_counts[uid]["sample_dates"]) < 3:
            user_activity_counts[uid]["sample_dates"].append(act.get("date"))
    
    # Look up user info for each
    results = []
    fixable_count = 0
    needs_team_assignment_count = 0
    
    for user_id, info in user_activity_counts.items():
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
        if user:
            has_team = bool(user.get("team_id"))
            if has_team:
                fixable_count += info["count"]
                status = "FIXABLE - user has team_id"
            else:
                needs_team_assignment_count += info["count"]
                status = "NEEDS USER TEAM ASSIGNMENT"
            results.append({
                "user_id": user_id,
                "user_name": user.get("name"),
                "user_email": user.get("email"),
                "user_team_id": user.get("team_id"),
                "user_status": user.get("status", "active"),
                "activity_count": info["count"],
                "sample_dates": info["sample_dates"],
                "fix_status": status
            })
        else:
            results.append({
                "user_id": user_id,
                "user_name": "USER NOT FOUND",
                "activity_count": info["count"],
                "sample_dates": info["sample_dates"],
                "fix_status": "ORPHANED - user deleted"
            })
    
    return {
        "total_orphaned_activities": len(orphaned),
        "fixable_activities": fixable_count,
        "needs_team_assignment_activities": needs_team_assignment_count,
        "users_with_orphaned_activities": results,
        "instructions": {
            "fixable": "Run POST /admin/migrate-activities-team-id to auto-fix",
            "needs_team": "First assign user to team via /admin/fix-unassigned-users, then run migration"
        }
    }

@api_router.get("/admin/diagnose-subtabs")
async def diagnose_subtabs(current_user: dict = Depends(get_current_user)):
    """
    Comprehensive diagnostics for New Faces, SNA, and NPA sub-tabs.
    Shows data integrity issues affecting rollups.
    (super_admin only, read-only)
    """
    require_super_admin(current_user)
    
    results = {
        "new_face_customers": {},
        "npa_agents": {},
        "activities_with_premium": {},
        "summary": {}
    }
    
    # ===== NEW FACE CUSTOMERS =====
    nfc_total = await db.new_face_customers.count_documents({})
    nfc_missing_team = await db.new_face_customers.count_documents(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]}
    )
    nfc_missing_user = await db.new_face_customers.count_documents(
        {"$or": [{"user_id": None}, {"user_id": {"$exists": False}}, {"user_id": ""}]}
    )
    
    # Group by year-month
    nfc_pipeline = [
        {"$addFields": {"month": {"$substr": ["$date", 0, 7]}}},
        {"$group": {"_id": "$month", "count": {"$sum": 1}}},
        {"$sort": {"_id": -1}}
    ]
    nfc_by_month = await db.new_face_customers.aggregate(nfc_pipeline).to_list(100)
    
    # Check for orphaned user_ids
    nfc_user_ids = await db.new_face_customers.distinct("user_id")
    existing_user_ids = await db.users.distinct("id")
    nfc_orphaned_users = [uid for uid in nfc_user_ids if uid and uid not in existing_user_ids]
    
    results["new_face_customers"] = {
        "total_records": nfc_total,
        "missing_team_id": nfc_missing_team,
        "missing_user_id": nfc_missing_user,
        "orphaned_user_ids": len(nfc_orphaned_users),
        "orphaned_user_ids_list": nfc_orphaned_users[:10],  # First 10
        "by_month": {m["_id"]: m["count"] for m in nfc_by_month if m["_id"]},
        "filters_used": "Query by user_id + team_id for team scoping"
    }
    
    # ===== NPA AGENTS =====
    npa_total = await db.npa_agents.count_documents({})
    npa_missing_team = await db.npa_agents.count_documents(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]}
    )
    npa_missing_user = await db.npa_agents.count_documents(
        {"$or": [{"user_id": None}, {"user_id": {"$exists": False}}, {"user_id": ""}]}
    )
    
    # Group by year-month from created_at or start_date
    npa_pipeline = [
        {"$addFields": {"month": {"$substr": [{"$ifNull": ["$start_date", "$created_at"]}, 0, 7]}}},
        {"$group": {"_id": "$month", "count": {"$sum": 1}}},
        {"$sort": {"_id": -1}}
    ]
    npa_by_month = await db.npa_agents.aggregate(npa_pipeline).to_list(100)
    
    results["npa_agents"] = {
        "total_records": npa_total,
        "missing_team_id": npa_missing_team,
        "missing_user_id_or_empty": npa_missing_user,
        "by_month": {m["_id"]: m["count"] for m in npa_by_month if m["_id"]},
        "filters_used": "NPA premium calculated from activities.premium WHERE team_id matches",
        "note": "NPA with linked user_id pulls premium from activities collection"
    }
    
    # ===== ACTIVITIES WITH PREMIUM (affects SNA/NPA rollups) =====
    act_total = await db.activities.count_documents({"premium": {"$gt": 0}})
    act_missing_team = await db.activities.count_documents(
        {"premium": {"$gt": 0}, "$or": [{"team_id": None}, {"team_id": {"$exists": False}}]}
    )
    act_missing_user = await db.activities.count_documents(
        {"premium": {"$gt": 0}, "$or": [{"user_id": None}, {"user_id": {"$exists": False}}, {"user_id": ""}]}
    )
    
    # Group premium activities by year-month
    act_pipeline = [
        {"$match": {"premium": {"$gt": 0}}},
        {"$addFields": {"month": {"$substr": ["$date", 0, 7]}}},
        {"$group": {"_id": "$month", "count": {"$sum": 1}, "total_premium": {"$sum": "$premium"}}},
        {"$sort": {"_id": -1}}
    ]
    act_by_month = await db.activities.aggregate(act_pipeline).to_list(100)
    
    # Activities missing team_id by team (via user lookup)
    missing_team_activities = await db.activities.find(
        {"premium": {"$gt": 0}, "$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"_id": 0, "user_id": 1, "date": 1, "premium": 1}
    ).to_list(1000)
    
    # Group missing by user_id
    missing_by_user = {}
    for act in missing_team_activities:
        uid = act.get("user_id", "unknown")
        if uid not in missing_by_user:
            missing_by_user[uid] = {"count": 0, "total_premium": 0}
        missing_by_user[uid]["count"] += 1
        missing_by_user[uid]["total_premium"] += act.get("premium", 0)
    
    # Lookup user info for missing activities
    missing_users_detail = []
    for uid, info in list(missing_by_user.items())[:20]:
        user = await db.users.find_one({"id": uid}, {"_id": 0, "name": 1, "email": 1, "team_id": 1})
        missing_users_detail.append({
            "user_id": uid,
            "user_name": user.get("name") if user else "NOT FOUND",
            "user_email": user.get("email") if user else None,
            "user_team_id": user.get("team_id") if user else None,
            "activities_count": info["count"],
            "total_premium": info["total_premium"],
            "fixable": bool(user and user.get("team_id"))
        })
    
    results["activities_with_premium"] = {
        "total_records": act_total,
        "missing_team_id": act_missing_team,
        "missing_user_id": act_missing_user,
        "by_month": {m["_id"]: {"count": m["count"], "premium": m["total_premium"]} for m in act_by_month if m["_id"]},
        "missing_team_id_by_user": missing_users_detail,
        "filters_used": "SNA/NPA queries filter by team_id - missing team_id = invisible to rollups",
        "impact": "Activities without team_id do NOT appear in SNA tracker or NPA calculations"
    }
    
    # ===== SUMMARY =====
    fixable_premium_activities = sum(1 for u in missing_users_detail if u.get("fixable"))
    unfixable_premium_activities = len(missing_users_detail) - fixable_premium_activities
    
    results["summary"] = {
        "total_data_issues": nfc_missing_team + npa_missing_team + act_missing_team,
        "new_face_customers_missing_team_id": nfc_missing_team,
        "npa_agents_missing_team_id": npa_missing_team,
        "premium_activities_missing_team_id": act_missing_team,
        "fixable_users": fixable_premium_activities,
        "unfixable_users": unfixable_premium_activities,
        "recommendation": "Run POST /admin/migrate-all-team-ids to fix all collections"
    }
    
    return results

@api_router.post("/admin/migrate-all-team-ids")
async def migrate_all_team_ids(current_user: dict = Depends(get_current_user)):
    """
    Migrate team_id for ALL collections: activities, new_face_customers, npa_agents.
    Sets team_id based on the user's current team assignment.
    (super_admin only)
    """
    require_super_admin(current_user)
    
    report = {
        "activities": {"updated": 0, "skipped_no_user_team": 0, "skipped_user_not_found": 0, "skipped_details": []},
        "new_face_customers": {"updated": 0, "skipped_no_user_team": 0, "skipped_user_not_found": 0, "skipped_details": []},
        "npa_agents": {"updated": 0, "skipped_no_user_team": 0, "skipped_user_not_found": 0, "skipped_details": []}
    }
    
    # Build user lookup with full info for reporting
    all_users = await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1, "team_id": 1}).to_list(10000)
    user_team_map = {u["id"]: u.get("team_id") for u in all_users}
    user_info_map = {u["id"]: {"name": u.get("name"), "email": u.get("email"), "team_id": u.get("team_id")} for u in all_users}
    
    # Track skipped users for detailed reporting
    skipped_users = {"activities": {}, "new_face_customers": {}, "npa_agents": {}}
    
    # ===== ACTIVITIES =====
    orphaned_activities = await db.activities.find(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"_id": 0, "id": 1, "user_id": 1, "date": 1, "premium": 1}
    ).to_list(10000)
    
    for act in orphaned_activities:
        user_id = act.get("user_id")
        if not user_id:
            report["activities"]["skipped_user_not_found"] += 1
            continue
        
        team_id = user_team_map.get(user_id)
        if not team_id:
            report["activities"]["skipped_no_user_team"] += 1
            # Track for detailed report
            if user_id not in skipped_users["activities"]:
                user_info = user_info_map.get(user_id, {})
                skipped_users["activities"][user_id] = {
                    "user_id": user_id,
                    "user_name": user_info.get("name", "NOT FOUND"),
                    "user_email": user_info.get("email"),
                    "reason": "user_not_found" if user_id not in user_info_map else "no_team_assigned",
                    "record_count": 0,
                    "total_premium": 0
                }
            skipped_users["activities"][user_id]["record_count"] += 1
            skipped_users["activities"][user_id]["total_premium"] += act.get("premium", 0)
            continue
        
        await db.activities.update_one({"id": act["id"]}, {"$set": {"team_id": team_id}})
        report["activities"]["updated"] += 1
    
    report["activities"]["skipped_details"] = list(skipped_users["activities"].values())
    
    # ===== NEW FACE CUSTOMERS =====
    orphaned_nfc = await db.new_face_customers.find(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"_id": 0, "id": 1, "user_id": 1, "policy_amount": 1}
    ).to_list(10000)
    
    for nfc in orphaned_nfc:
        user_id = nfc.get("user_id")
        if not user_id:
            report["new_face_customers"]["skipped_user_not_found"] += 1
            continue
        
        team_id = user_team_map.get(user_id)
        if not team_id:
            report["new_face_customers"]["skipped_no_user_team"] += 1
            if user_id not in skipped_users["new_face_customers"]:
                user_info = user_info_map.get(user_id, {})
                skipped_users["new_face_customers"][user_id] = {
                    "user_id": user_id,
                    "user_name": user_info.get("name", "NOT FOUND"),
                    "user_email": user_info.get("email"),
                    "reason": "user_not_found" if user_id not in user_info_map else "no_team_assigned",
                    "record_count": 0
                }
            skipped_users["new_face_customers"][user_id]["record_count"] += 1
            continue
        
        await db.new_face_customers.update_one({"id": nfc["id"]}, {"$set": {"team_id": team_id}})
        report["new_face_customers"]["updated"] += 1
    
    report["new_face_customers"]["skipped_details"] = list(skipped_users["new_face_customers"].values())
    
    # ===== NPA AGENTS =====
    orphaned_npa = await db.npa_agents.find(
        {"$or": [{"team_id": None}, {"team_id": {"$exists": False}}]},
        {"_id": 0, "id": 1, "user_id": 1, "added_by": 1, "name": 1}
    ).to_list(10000)
    
    for npa in orphaned_npa:
        # Try user_id first, then added_by
        user_id = npa.get("user_id") or npa.get("added_by")
        if not user_id:
            report["npa_agents"]["skipped_user_not_found"] += 1
            skipped_users["npa_agents"][npa.get("id", "unknown")] = {
                "npa_name": npa.get("name"),
                "reason": "no_user_id_or_added_by"
            }
            continue
        
        team_id = user_team_map.get(user_id)
        if not team_id:
            report["npa_agents"]["skipped_no_user_team"] += 1
            user_info = user_info_map.get(user_id, {})
            skipped_users["npa_agents"][user_id] = {
                "user_id": user_id,
                "user_name": user_info.get("name", "NOT FOUND"),
                "user_email": user_info.get("email"),
                "npa_name": npa.get("name"),
                "reason": "user_not_found" if user_id not in user_info_map else "no_team_assigned"
            }
            continue
        
        await db.npa_agents.update_one({"id": npa["id"]}, {"$set": {"team_id": team_id}})
        report["npa_agents"]["updated"] += 1
    
    report["npa_agents"]["skipped_details"] = list(skipped_users["npa_agents"].values())
    
    total_updated = sum(r["updated"] for r in report.values())
    total_skipped = sum(r["skipped_no_user_team"] + r["skipped_user_not_found"] for r in report.values())
    
    return {
        "message": f"Migration complete. Updated {total_updated} records, skipped {total_skipped}.",
        "total_updated": total_updated,
        "total_skipped": total_skipped,
        "report": report
    }

@api_router.get("/admin/teams")
async def get_all_teams(current_user: dict = Depends(get_current_user)):
    """Get all teams (super_admin only)"""
    require_super_admin(current_user)
    teams = await db.teams.find({}, {"_id": 0}).to_list(1000)
    
    # Add user count for each team
    for team in teams:
        user_count = await db.users.count_documents({"team_id": team['id']})
        team['user_count'] = user_count
    
    return teams

@api_router.post("/admin/create-missing-team-record")
async def create_missing_team_record(current_user: dict = Depends(get_current_user)):
    """
    Create Team Sudbeck if it doesn't exist. This is a one-time bootstrap for legacy data.
    (super_admin only)
    """
    require_super_admin(current_user)
    
    # Check if Team Sudbeck already exists
    existing = await db.teams.find_one({"name": "Team Sudbeck"})
    if existing:
        return {"message": "Team Sudbeck already exists", "team": {"id": existing.get('id'), "name": existing.get('name')}, "created": False}
    
    # Create Team Sudbeck
    new_team = {
        "id": str(uuid.uuid4()),
        "name": "Team Sudbeck",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "settings": {"is_default": True}
    }
    
    await db.teams.insert_one(new_team)
    new_team.pop('_id', None)
    
    return {
        "message": "Team Sudbeck created successfully",
        "team": new_team,
        "created": True
    }

@api_router.post("/admin/setup-all-branding")
async def setup_all_branding(current_user: dict = Depends(get_current_user)):
    """
    One-time setup: Apply branding (logos, colors) to all teams.
    (super_admin only)
    """
    require_super_admin(current_user)
    
    # Define branding for each team
    branding_config = {
        "sudbeck": {
            "logo_url": "https://customer-assets.emergentagent.com/job_0086b560-8cff-4294-89b3-b94a427d032c/artifacts/rxkifhzg_IMG_2351.jpeg",
            "primary_color": "#1e3a5f",
            "accent_color": "#2563eb",
            "display_name": "Team Sudbeck",
            "tagline": "We Build Leaders"
        },
        "gaines": {
            "logo_url": "https://customer-assets.emergentagent.com/job_0086b560-8cff-4294-89b3-b94a427d032c/artifacts/5zlvaqi5_image.png",
            "primary_color": "#1e3a5f",
            "accent_color": "#3b82f6",
            "display_name": "Team Gaines",
            "tagline": "Arizona • Texas"
        },
        "koch": {
            "logo_url": "https://customer-assets.emergentagent.com/job_0086b560-8cff-4294-89b3-b94a427d032c/artifacts/grsbn8m3_image.png",
            "primary_color": "#1e3a5f",
            "accent_color": "#3b82f6",
            "display_name": "Team Koch",
            "tagline": "South Dakota"
        },
        "quick": {
            "logo_url": "https://customer-assets.emergentagent.com/job_0086b560-8cff-4294-89b3-b94a427d032c/artifacts/9odjpoz2_image.png",
            "primary_color": "#1e3a5f",
            "accent_color": "#3b82f6",
            "display_name": "Team Quick",
            "tagline": "Joining Forces Changing Lives"
        },
        "graham": {
            "logo_url": "https://customer-assets.emergentagent.com/job_0086b560-8cff-4294-89b3-b94a427d032c/artifacts/hq27bw63_image.png",
            "primary_color": "#1e3a5f",
            "accent_color": "#3b82f6",
            "display_name": "Team Graham",
            "tagline": "Kansas • Colorado • Nebraska"
        }
    }
    
    # Get all teams
    teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    
    results = []
    for team in teams:
        team_name = team.get('name', '').lower()
        
        # Find matching branding config
        branding = None
        matched_key = None
        for key, config in branding_config.items():
            if key in team_name:
                branding = config
                matched_key = key
                break
        
        if branding:
            await db.teams.update_one(
                {"id": team.get('id')},
                {"$set": {"branding": branding}}
            )
            results.append({
                "team": team.get('name'),
                "status": "updated",
                "branding": branding['display_name']
            })
        else:
            results.append({
                "team": team.get('name'),
                "status": "skipped",
                "reason": "no matching branding config"
            })
    
    updated_count = len([r for r in results if r['status'] == 'updated'])
    
    return {
        "message": f"Branding applied to {updated_count} of {len(teams)} teams",
        "results": results
    }

# Team Feature Flags Endpoints
class TeamFeaturesUpdate(BaseModel):
    features: Dict[str, bool]

@api_router.get("/admin/teams/{team_id}/features")
async def get_team_features(team_id: str, current_user: dict = Depends(get_current_user)):
    """Get team feature flags (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Return features with defaults for any missing flags
    features = team.get('features', {})
    merged_features = {**DEFAULT_TEAM_FEATURES, **features}
    
    return {
        "team_id": team_id,
        "team_name": team.get('name'),
        "features": merged_features
    }

@api_router.put("/admin/teams/{team_id}/features")
async def update_team_features(team_id: str, data: TeamFeaturesUpdate, current_user: dict = Depends(get_current_user)):
    """Update team feature flags (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    await db.teams.update_one(
        {"id": team_id},
        {"$set": {"features": data.features}}
    )
    
    return {
        "message": f"Features updated for {team.get('name')}",
        "features": data.features
    }

@api_router.post("/admin/teams/{team_id}/features/reset")
async def reset_team_features(team_id: str, current_user: dict = Depends(get_current_user)):
    """Reset team features to defaults (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    await db.teams.update_one(
        {"id": team_id},
        {"$set": {"features": DEFAULT_TEAM_FEATURES}}
    )
    
    return {
        "message": f"Features reset to defaults for {team.get('name')}",
        "features": DEFAULT_TEAM_FEATURES
    }

@api_router.post("/admin/teams/{team_id}/features/copy-from/{source_team_id}")
async def copy_team_features(team_id: str, source_team_id: str, current_user: dict = Depends(get_current_user)):
    """Copy features from another team (super_admin only)"""
    require_super_admin(current_user)
    
    target_team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not target_team:
        raise HTTPException(status_code=404, detail="Target team not found")
    
    source_team = await db.teams.find_one({"id": source_team_id}, {"_id": 0})
    if not source_team:
        raise HTTPException(status_code=404, detail="Source team not found")
    
    source_features = source_team.get('features', DEFAULT_TEAM_FEATURES)
    
    await db.teams.update_one(
        {"id": team_id},
        {"$set": {"features": source_features}}
    )
    
    return {
        "message": f"Features copied from {source_team.get('name')} to {target_team.get('name')}",
        "features": source_features
    }

# ==================== TEAM ROLE TAB OVERRIDES ====================

class RoleTabOverridesUpdate(BaseModel):
    role_tab_overrides: Dict[str, Dict[str, list]]

@api_router.get("/admin/teams/{team_id}/role-overrides")
async def get_team_role_overrides(team_id: str, current_user: dict = Depends(get_current_user)):
    """Get team role-based tab overrides (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    role_overrides = team.get('role_tab_overrides', DEFAULT_ROLE_TAB_OVERRIDES)
    
    return {
        "team_id": team_id,
        "team_name": team.get('name'),
        "role_tab_overrides": role_overrides,
        "available_tabs": list(DEFAULT_TEAM_FEATURES.keys())
    }

@api_router.put("/admin/teams/{team_id}/role-overrides")
async def update_team_role_overrides(team_id: str, data: RoleTabOverridesUpdate, current_user: dict = Depends(get_current_user)):
    """Update team role-based tab overrides (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Validate role names
    valid_roles = ['agent', 'district_manager', 'regional_manager']
    for role in data.role_tab_overrides.keys():
        if role not in valid_roles:
            raise HTTPException(status_code=400, detail=f"Invalid role: {role}. Must be one of {valid_roles}")
    
    # Validate tab names
    valid_tabs = list(DEFAULT_TEAM_FEATURES.keys())
    for role, config in data.role_tab_overrides.items():
        hidden_tabs = config.get('hidden_tabs', [])
        for tab in hidden_tabs:
            if tab not in valid_tabs:
                raise HTTPException(status_code=400, detail=f"Invalid tab: {tab}. Must be one of {valid_tabs}")
    
    await db.teams.update_one(
        {"id": team_id},
        {"$set": {"role_tab_overrides": data.role_tab_overrides}}
    )
    
    return {
        "message": "Role tab overrides updated",
        "role_tab_overrides": data.role_tab_overrides
    }

# ==================== TEAM UI SETTINGS ====================

class TeamUISettingsUpdate(BaseModel):
    default_landing_tab: Optional[str] = None
    default_leaderboard_period: Optional[str] = None

@api_router.get("/admin/teams/{team_id}/ui-settings")
async def get_team_ui_settings(team_id: str, current_user: dict = Depends(get_current_user)):
    """Get team UI settings (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    ui_settings = team.get('ui_settings', DEFAULT_TEAM_UI_SETTINGS)
    merged_settings = {**DEFAULT_TEAM_UI_SETTINGS, **ui_settings}
    
    return {
        "team_id": team_id,
        "team_name": team.get('name'),
        "ui_settings": merged_settings,
        "available_tabs": list(DEFAULT_TEAM_FEATURES.keys()),
        "available_periods": ["weekly", "monthly", "quarterly", "yearly"]
    }

@api_router.put("/admin/teams/{team_id}/ui-settings")
async def update_team_ui_settings(team_id: str, data: TeamUISettingsUpdate, current_user: dict = Depends(get_current_user)):
    """Update team UI settings (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Validate landing tab
    valid_tabs = list(DEFAULT_TEAM_FEATURES.keys())
    if data.default_landing_tab and data.default_landing_tab not in valid_tabs:
        raise HTTPException(status_code=400, detail=f"Invalid landing tab. Must be one of {valid_tabs}")
    
    # Validate leaderboard period
    valid_periods = ["weekly", "monthly", "quarterly", "yearly"]
    if data.default_leaderboard_period and data.default_leaderboard_period not in valid_periods:
        raise HTTPException(status_code=400, detail=f"Invalid period. Must be one of {valid_periods}")
    
    # Build update object
    update_data = {}
    if data.default_landing_tab:
        update_data["ui_settings.default_landing_tab"] = data.default_landing_tab
    if data.default_leaderboard_period:
        update_data["ui_settings.default_leaderboard_period"] = data.default_leaderboard_period
    
    if update_data:
        await db.teams.update_one(
            {"id": team_id},
            {"$set": update_data}
        )
    
    # Get updated settings
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    ui_settings = team.get('ui_settings', DEFAULT_TEAM_UI_SETTINGS)
    
    return {
        "message": "UI settings updated",
        "ui_settings": {**DEFAULT_TEAM_UI_SETTINGS, **ui_settings}
    }

# ==================== FULL TEAM CONFIG (Combined View) ====================

@api_router.get("/admin/teams/{team_id}/full-config")
async def get_team_full_config(team_id: str, current_user: dict = Depends(get_current_user)):
    """Get complete team configuration including features, role overrides, UI settings, view settings, and branding (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Merge with defaults
    features = {**DEFAULT_TEAM_FEATURES, **team.get('features', {})}
    role_overrides = {**DEFAULT_ROLE_TAB_OVERRIDES, **team.get('role_tab_overrides', {})}
    ui_settings = {**DEFAULT_TEAM_UI_SETTINGS, **team.get('ui_settings', {})}
    view_settings = await get_team_view_settings(team)
    branding = team.get('branding', {
        "logo_url": None,
        "primary_color": "#1e40af",
        "accent_color": "#3b82f6",
        "display_name": None,
        "tagline": None
    })
    
    return {
        "team_id": team_id,
        "team_name": team.get('name'),
        "features": features,
        "role_tab_overrides": role_overrides,
        "ui_settings": ui_settings,
        "view_settings": view_settings,
        "branding": branding,
        "available_tabs": list(DEFAULT_TEAM_FEATURES.keys()),
        "available_roles": ["agent", "district_manager", "regional_manager"],
        "available_periods": ["weekly", "monthly", "quarterly", "yearly"],
        "available_kpi_cards": [card["id"] for card in DEFAULT_TEAM_VIEW_SETTINGS["kpi_cards"]]
    }

@api_router.put("/admin/teams/{team_id}/full-config")
async def update_team_full_config(team_id: str, config: dict, current_user: dict = Depends(get_current_user)):
    """Update complete team configuration (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    update_data = {}
    
    # Update features if provided
    if 'features' in config:
        update_data['features'] = config['features']
    
    # Update role overrides if provided
    if 'role_tab_overrides' in config:
        # Validate roles
        valid_roles = ['agent', 'district_manager', 'regional_manager']
        for role in config['role_tab_overrides'].keys():
            if role not in valid_roles:
                raise HTTPException(status_code=400, detail=f"Invalid role: {role}")
        update_data['role_tab_overrides'] = config['role_tab_overrides']
    
    # Update UI settings if provided
    if 'ui_settings' in config:
        update_data['ui_settings'] = config['ui_settings']
    
    # Update view settings if provided (Phase 2)
    if 'view_settings' in config:
        update_data['team_settings.views'] = config['view_settings']
    
    # Update branding if provided
    if 'branding' in config:
        update_data['branding'] = config['branding']
    
    if update_data:
        await db.teams.update_one(
            {"id": team_id},
            {"$set": update_data}
        )
    
    return {"message": "Team configuration updated", "updated_fields": list(update_data.keys())}

@api_router.get("/teams/my-features")
async def get_my_team_features(current_user: dict = Depends(get_current_user)):
    """
    Get current user's EFFECTIVE features for frontend tab visibility.
    This includes team feature flags + role-based tab overrides.
    Also returns UI settings (default landing tab, leaderboard period).
    """
    team_id = current_user.get('team_id')
    role = current_user.get('role', 'agent')
    
    # Get team config
    team = None
    if team_id:
        team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    
    # Get effective features (handles super_admin, state_manager, and role overrides)
    effective_features = await get_effective_features(current_user, team)
    
    # Get UI settings from team config (or defaults)
    ui_settings = DEFAULT_TEAM_UI_SETTINGS.copy()
    if team:
        team_ui = team.get('ui_settings', {})
        ui_settings = {**ui_settings, **team_ui}
    
    # Get view settings (Phase 2) - KPI cards and subtabs
    view_settings = await get_team_view_settings(team)
    
    return {
        "features": effective_features,
        "ui_settings": ui_settings,
        "view_settings": view_settings,
        "role": role
    }

@api_router.get("/admin/debug-teams")
async def debug_teams(current_user: dict = Depends(get_current_user)):
    """Raw database dump of all teams for debugging (super_admin only)"""
    require_super_admin(current_user)
    
    # Get ALL teams from database
    all_teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    
    # Get all distinct team_ids from users
    user_team_ids = await db.users.distinct("team_id")
    
    # Count users per team
    team_stats = []
    for team in all_teams:
        count = await db.users.count_documents({"team_id": team.get('id')})
        team_stats.append({
            "id": team.get('id'),
            "name": team.get('name'),
            "settings": team.get('settings'),
            "user_count": count
        })
    
    return {
        "total_teams_in_db": len(all_teams),
        "teams": team_stats,
        "user_team_ids": [t for t in user_team_ids if t],
        "raw_teams": all_teams
    }

@api_router.get("/admin/default-team")
async def get_default_team(current_user: dict = Depends(get_current_user)):
    """Get the default team (Team Sudbeck) directly from database (super_admin only)"""
    require_super_admin(current_user)
    
    # Get ALL teams
    all_teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    
    # Find Team Sudbeck by name or default flag
    default_team = None
    for team in all_teams:
        name = team.get('name', '').lower()
        if 'sudbeck' in name or team.get('settings', {}).get('is_default') == True:
            default_team = team
            break
    
    # Get ALL distinct team_ids that users have - this reveals ALL teams in use
    all_team_ids_in_use = await db.users.distinct("team_id")
    
    # For each team_id in use, get count and check if it matches a known team
    team_usage = []
    for tid in all_team_ids_in_use:
        if not tid:
            continue
        count = await db.users.count_documents({"team_id": tid})
        # Find if this team_id has a team record
        team_record = await db.teams.find_one({"id": tid}, {"_id": 0})
        team_usage.append({
            "team_id": tid,
            "user_count": count,
            "team_name": team_record.get('name') if team_record else "NO TEAM RECORD - ORPHANED",
            "has_team_record": team_record is not None
        })
    
    # Sort by user count descending - the biggest team is likely Team Sudbeck
    team_usage.sort(key=lambda x: x['user_count'], reverse=True)
    
    # The team with most users that's NOT in all_teams is likely Team Sudbeck
    orphaned_teams = [t for t in team_usage if not t['has_team_record']]
    
    if default_team:
        return {
            "found": True,
            "team": default_team,
            "all_teams": all_teams,
            "team_usage": team_usage,
            "message": f"Found: {default_team.get('name')}"
        }
    
    return {
        "found": False,
        "all_teams": all_teams,
        "team_usage": team_usage,
        "orphaned_teams": orphaned_teams,
        "message": "Team Sudbeck not in teams collection. Check team_usage for orphaned team_ids - largest one is likely Team Sudbeck."
    }

@api_router.post("/admin/teams")
async def create_team(team_data: TeamCreate, current_user: dict = Depends(get_current_user)):
    """Create a new team (super_admin only)"""
    require_super_admin(current_user)
    
    # Check if team name exists
    existing = await db.teams.find_one({"name": team_data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Team name already exists")
    
    team = Team(name=team_data.name, settings=team_data.settings)
    team_dict = team.model_dump()
    team_dict['created_at'] = team_dict['created_at'].isoformat()
    
    await db.teams.insert_one(team_dict)
    team_dict.pop('_id', None)
    
    return {"message": f"Team '{team_data.name}' created successfully", "team": team_dict}

@api_router.put("/admin/teams/{team_id}")
async def update_team(team_id: str, team_data: TeamCreate, current_user: dict = Depends(get_current_user)):
    """Update a team (super_admin only)"""
    require_super_admin(current_user)
    
    existing = await db.teams.find_one({"id": team_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Team not found")
    
    await db.teams.update_one(
        {"id": team_id}, 
        {"$set": {"name": team_data.name, "settings": team_data.settings}}
    )
    
    return {"message": "Team updated successfully"}

@api_router.get("/admin/teams/{team_id}/branding")
async def get_team_branding(team_id: str, current_user: dict = Depends(get_current_user)):
    """Get team branding (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    return {
        "team_id": team_id,
        "team_name": team.get('name'),
        "branding": team.get('branding', {
            "logo_url": None,
            "primary_color": "#1e40af",
            "accent_color": "#3b82f6",
            "display_name": None,
            "tagline": None
        })
    }

@api_router.put("/admin/teams/{team_id}/branding")
async def update_team_branding(team_id: str, branding: TeamBrandingUpdate, current_user: dict = Depends(get_current_user)):
    """Update team branding (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Get existing branding and merge with updates
    existing_branding = team.get('branding', {})
    updated_branding = {
        "logo_url": branding.logo_url if branding.logo_url is not None else existing_branding.get('logo_url'),
        "primary_color": branding.primary_color if branding.primary_color is not None else existing_branding.get('primary_color', '#1e40af'),
        "accent_color": branding.accent_color if branding.accent_color is not None else existing_branding.get('accent_color', '#3b82f6'),
        "display_name": branding.display_name if branding.display_name is not None else existing_branding.get('display_name'),
        "tagline": branding.tagline if branding.tagline is not None else existing_branding.get('tagline')
    }
    
    await db.teams.update_one(
        {"id": team_id},
        {"$set": {"branding": updated_branding}}
    )
    
    return {
        "message": "Branding updated successfully",
        "branding": updated_branding
    }

@api_router.delete("/admin/teams/{team_id}")
async def delete_team(team_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a team (super_admin only) - only if no users assigned"""
    require_super_admin(current_user)
    
    user_count = await db.users.count_documents({"team_id": team_id})
    if user_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete team with {user_count} users assigned. Reassign users first.")
    
    await db.teams.delete_one({"id": team_id})
    return {"message": "Team deleted successfully"}

# ============================================
# Document Generation Endpoints (Admin Only)
# ============================================

@api_router.get("/admin/teams/{team_id}/roster/csv")
async def download_team_roster_csv(team_id: str, current_user: dict = Depends(get_current_user)):
    """Download team roster as CSV (super_admin only)"""
    require_super_admin(current_user)
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Get all users for this team
    users = await db.users.find(
        {"team_id": team_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
        {"_id": 0, "password_hash": 0}
    ).to_list(1000)
    
    # Build user lookup for manager names
    all_users = await db.users.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(10000)
    user_dict = {u['id']: u['name'] for u in all_users}
    
    # Role order for sorting
    role_order = {'state_manager': 0, 'regional_manager': 1, 'district_manager': 2, 'agent': 3}
    users.sort(key=lambda u: (role_order.get(u.get('role', 'agent'), 4), u.get('name', '')))
    
    # Build CSV
    import io
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name', 'Email', 'Role', 'Reports To'])
    
    for user in users:
        manager_name = user_dict.get(user.get('manager_id'), '—') if user.get('manager_id') else '—'
        role_display = user.get('role', '').replace('_', ' ').title()
        writer.writerow([user.get('name', ''), user.get('email', ''), role_display, manager_name])
    
    csv_content = output.getvalue()
    safe_name = team['name'].replace(' ', '_').replace('/', '_')
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={safe_name}_roster.csv"}
    )

@api_router.get("/admin/teams/{team_id}/roster/pdf")
async def download_team_roster_pdf(team_id: str, current_user: dict = Depends(get_current_user)):
    """Download team roster as PDF (super_admin only)"""
    require_super_admin(current_user)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Get all users for this team
    users = await db.users.find(
        {"team_id": team_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
        {"_id": 0, "password_hash": 0}
    ).to_list(1000)
    
    # Build user lookup for manager names
    all_users = await db.users.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(10000)
    user_dict = {u['id']: u['name'] for u in all_users}
    
    # Role order for sorting
    role_order = {'state_manager': 0, 'regional_manager': 1, 'district_manager': 2, 'agent': 3}
    users.sort(key=lambda u: (role_order.get(u.get('role', 'agent'), 4), u.get('name', '')))
    
    # Build PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=50, bottomMargin=50)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=20)
    elements.append(Paragraph(f"{team['name']} - Team Roster", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    table_data = [['Name', 'Email', 'Role', 'Reports To']]
    for user in users:
        manager_name = user_dict.get(user.get('manager_id'), '—') if user.get('manager_id') else '—'
        role_display = user.get('role', '').replace('_', ' ').title()
        table_data.append([user.get('name', ''), user.get('email', ''), role_display, manager_name])
    
    # Create table
    table = Table(table_data, colWidths=[140, 180, 100, 140])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(f"Total: {len(users)} team members", styles['Normal']))
    
    doc.build(elements)
    pdf_content = buffer.getvalue()
    safe_name = team['name'].replace(' ', '_').replace('/', '_')
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={safe_name}_roster.pdf"}
    )

@api_router.get("/admin/guides/state-manager")
async def download_state_manager_guide(current_user: dict = Depends(get_current_user)):
    """Download State Manager Quick Start Guide as PDF (super_admin only)"""
    require_super_admin(current_user)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, ListFlowable, ListItem
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=40, bottomMargin=40, leftMargin=50, rightMargin=50)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, spaceAfter=10, textColor=colors.HexColor('#1e40af'))
    h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=14, spaceBefore=15, spaceAfter=8, textColor=colors.HexColor('#1e40af'))
    body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, spaceAfter=6, leading=14)
    
    # Title
    elements.append(Paragraph("PMA Agent - State Manager Quick Start Guide", title_style))
    elements.append(Paragraph("Welcome! This guide covers everything you need to know to get started.", body_style))
    elements.append(Spacer(1, 15))
    
    # Logging In
    elements.append(Paragraph("Logging In", h2_style))
    elements.append(Paragraph("1. Go to your team's app URL", body_style))
    elements.append(Paragraph("2. Enter your email (username) and password", body_style))
    elements.append(Paragraph("3. Click Login - Your dashboard will load with your team's branding", body_style))
    elements.append(Spacer(1, 10))
    
    # Tabs
    elements.append(Paragraph("What You'll See", h2_style))
    tab_data = [
        ['Tab', 'What It Does'],
        ['Activity', 'Log your daily activity (contacts, appointments, presentations)'],
        ['Stats', 'View your personal statistics and trends'],
        ['Team View', 'See your entire team hierarchy with rollup totals'],
        ['Leaderboard', 'Team rankings for presentations, referrals, and more'],
        ['Reports', 'Download team reports by period'],
        ['Suitability', 'SNA/NPA forms - view and export'],
        ['DocuSphere', 'Team document library (you can upload/manage)'],
        ['Team Mgmt', 'Manage your team members and hierarchy'],
    ]
    tab_table = Table(tab_data, colWidths=[80, 400])
    tab_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(tab_table)
    elements.append(Spacer(1, 15))
    
    # Understanding Rollups
    elements.append(Paragraph("Understanding Rollups", h2_style))
    elements.append(Paragraph("<b>What is a rollup?</b> A rollup is the sum of your activity PLUS all your subordinates' activity.", body_style))
    elements.append(Paragraph("<b>Example:</b> You logged 5 presentations + Your 3 Regional Managers logged 10 + Their teams logged 85 = Your rollup is <b>100 presentations</b>", body_style))
    elements.append(Paragraph("<b>Where to see rollups:</b> Team View (click your name) or Reports (download hierarchy report)", body_style))
    elements.append(Spacer(1, 10))
    
    # What you CAN do
    elements.append(Paragraph("What You CAN Do", h2_style))
    can_do = ["Log your own daily activity", "View all team members' activity and stats", "Download reports for any time period",
              "Upload documents to DocuSphere", "Create/delete folders in DocuSphere", "View and reassign team members",
              "See all suitability forms from your team", "Export suitability data to CSV/Excel"]
    for item in can_do:
        elements.append(Paragraph(f"✓ {item}", body_style))
    elements.append(Spacer(1, 10))
    
    # What you CANNOT do
    elements.append(Paragraph("What You CANNOT Do", h2_style))
    cannot_do = ["Create new users (Admin only)", "Delete users (Admin only)", "Change feature flags (Admin only)", "See other teams' data"]
    for item in cannot_do:
        elements.append(Paragraph(f"✗ {item}", body_style))
    elements.append(Spacer(1, 10))
    
    # DocuSphere
    elements.append(Paragraph("DocuSphere (Document Library)", h2_style))
    elements.append(Paragraph("You are the <b>only role</b> that can manage DocuSphere for your team. Your team members can VIEW documents but cannot upload or delete.", body_style))
    elements.append(Paragraph("<b>To upload:</b> Go to DocuSphere → Navigate to folder → Click Upload Document", body_style))
    elements.append(Paragraph("<b>To create folder:</b> Click New Folder → Enter name → Click Create", body_style))
    elements.append(Spacer(1, 10))
    
    # Tips
    elements.append(Paragraph("Tips", h2_style))
    elements.append(Paragraph("• Check rollups weekly - Make sure numbers match expectations", body_style))
    elements.append(Paragraph("• Upload key docs to DocuSphere - Your team can access them anytime", body_style))
    elements.append(Paragraph("• Use the Leaderboard - Motivate top performers", body_style))
    elements.append(Paragraph("• Archive departed users promptly - Keeps your team list clean", body_style))
    elements.append(Spacer(1, 15))
    
    # Footer
    elements.append(Paragraph("Need Help? Contact your Super Admin for password resets, adding new team members, or technical problems.", body_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<i>Generated: {datetime.now().strftime('%B %d, %Y')}</i>", body_style))
    
    doc.build(elements)
    pdf_content = buffer.getvalue()
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=State_Manager_Quick_Start_Guide.pdf"}
    )

@api_router.get("/admin/guides/admin-playbook")
async def download_admin_playbook(current_user: dict = Depends(get_current_user)):
    """Download Admin Playbook as PDF (super_admin only)"""
    require_super_admin(current_user)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=40, bottomMargin=40, leftMargin=50, rightMargin=50)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=22, spaceAfter=10, textColor=colors.HexColor('#1e40af'))
    h1_style = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=16, spaceBefore=20, spaceAfter=10, textColor=colors.HexColor('#1e40af'))
    h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=12, spaceBefore=12, spaceAfter=6, textColor=colors.HexColor('#334155'))
    body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, spaceAfter=6, leading=14)
    step_style = ParagraphStyle('Step', parent=styles['Normal'], fontSize=10, spaceAfter=4, leftIndent=15, leading=13)
    
    # Title Page
    elements.append(Paragraph("PMA Agent", title_style))
    elements.append(Paragraph("Super Admin Playbook", ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=16, textColor=colors.HexColor('#64748b'))))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("A step-by-step guide for managing teams, users, and system administration.", body_style))
    elements.append(Spacer(1, 30))
    
    # Table of Contents
    elements.append(Paragraph("Table of Contents", h2_style))
    toc = ["1. Logging In as Super Admin", "2. Creating a New Team", "3. Configuring Feature Flags",
           "4. Creating Users", "5. Setting Up Hierarchy", "6. Team Branding", "7. DocuSphere Rules",
           "8. Diagnostics & Recovery", "9. Common Issues & Solutions"]
    for item in toc:
        elements.append(Paragraph(item, body_style))
    elements.append(PageBreak())
    
    # Section 1: Logging In
    elements.append(Paragraph("1. Logging In as Super Admin", h1_style))
    elements.append(Paragraph("1. Go to your app URL", step_style))
    elements.append(Paragraph("2. Enter your super admin email and password", step_style))
    elements.append(Paragraph("3. Click Login - You will see the full navigation with an Admin tab", step_style))
    elements.append(Paragraph("<b>Note:</b> Super admins can see all teams and bypass feature restrictions.", body_style))
    
    # Section 2: Creating Team
    elements.append(Paragraph("2. Creating a New Team", h1_style))
    elements.append(Paragraph("1. Click <b>Admin</b> in the navigation", step_style))
    elements.append(Paragraph("2. You're on the <b>Teams</b> tab by default", step_style))
    elements.append(Paragraph("3. Click <b>+ New Team</b>", step_style))
    elements.append(Paragraph("4. Enter the team name (e.g., 'Team Nebraska')", step_style))
    elements.append(Paragraph("5. Click <b>Create Team</b>", step_style))
    elements.append(Paragraph("<b>What happens:</b> A new team is created with default feature flags (all ON except Recruiting).", body_style))
    
    # Section 3: Feature Flags
    elements.append(Paragraph("3. Configuring Feature Flags", h1_style))
    elements.append(Paragraph("Feature flags control which tabs/features each team can access.", body_style))
    
    feature_data = [
        ['Feature', 'Description', 'Default'],
        ['Activity', 'Daily activity logging', 'ON'],
        ['Stats', 'Personal statistics view', 'ON'],
        ['Team View', 'Hierarchy tree with rollups', 'ON'],
        ['Suitability', 'SNA/NPA forms', 'ON'],
        ['PMA Bonuses', 'Bonus PDF uploads', 'ON'],
        ['DocuSphere', 'Document library', 'ON'],
        ['Leaderboard', 'Team rankings', 'ON'],
        ['Analytics', 'Charts and trends', 'ON'],
        ['Reports', 'Manager reports & exports', 'ON'],
        ['Team Mgmt', 'Team management', 'ON'],
        ['Interviews', 'Interview tracking', 'ON'],
        ['Recruiting', 'Recruiting pipeline', 'OFF'],
    ]
    feature_table = Table(feature_data, colWidths=[80, 280, 50])
    feature_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(feature_table)
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("<b>To configure:</b> Admin → Teams → Click gear icon (⚙️) → Toggle features ON/OFF", body_style))
    
    # Section 4: Creating Users
    elements.append(Paragraph("4. Creating Users", h1_style))
    elements.append(Paragraph("1. Go to <b>Admin → Users</b>", step_style))
    elements.append(Paragraph("2. Click <b>+ Create User</b>", step_style))
    elements.append(Paragraph("3. Fill in: Team, Full Name, Email (login), Password, Role", step_style))
    elements.append(Paragraph("4. Optionally set Manager (can set later in Team Mgmt)", step_style))
    elements.append(Paragraph("5. Click <b>Create User</b>", step_style))
    
    role_data = [
        ['Role', 'Access Level'],
        ['State Manager', 'Full team access, can manage DocuSphere, see all reports'],
        ['Regional Manager', 'Sees direct + indirect reports, limited admin'],
        ['District Manager', 'Sees direct reports only'],
        ['Agent', 'Personal data only'],
    ]
    role_table = Table(role_data, colWidths=[120, 350])
    role_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(role_table)
    
    # Section 5: Hierarchy
    elements.append(Paragraph("5. Setting Up Hierarchy", h1_style))
    elements.append(Paragraph("Hierarchy determines who reports to whom and affects data visibility and rollups.", body_style))
    elements.append(Paragraph("<b>Option A:</b> Set during user creation - select manager from dropdown", body_style))
    elements.append(Paragraph("<b>Option B:</b> Edit existing user - Admin → Users → pencil icon → change Manager", body_style))
    elements.append(Paragraph("<b>Option C:</b> Use Team Management (as State Manager) → Reorganize tab", body_style))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("<b>Hierarchy Rules:</b>", body_style))
    elements.append(Paragraph("• State Manager at top (no manager)", step_style))
    elements.append(Paragraph("• Regional Managers → State Manager", step_style))
    elements.append(Paragraph("• District Managers → Regional Managers", step_style))
    elements.append(Paragraph("• Agents → District Managers", step_style))
    elements.append(Paragraph("• Users can only have ONE manager from the SAME team", step_style))
    
    # Section 6: Branding
    elements.append(Paragraph("6. Team Branding", h1_style))
    elements.append(Paragraph("Each team can have custom colors, logo, and display name.", body_style))
    elements.append(Paragraph("1. Go to <b>Admin → Teams</b>", step_style))
    elements.append(Paragraph("2. Click the <b>paint palette icon</b> for the team", step_style))
    elements.append(Paragraph("3. Set: Display Name, Tagline, Primary Color, Accent Color, Logo URL", step_style))
    elements.append(Paragraph("4. Click <b>Save Branding</b>", step_style))
    
    # Section 7: DocuSphere
    elements.append(Paragraph("7. DocuSphere Rules", h1_style))
    docu_data = [
        ['Role', 'View', 'Upload', 'Create Folders', 'Delete'],
        ['State Manager', '✓', '✓', '✓', '✓'],
        ['Regional Manager', '✓', '✗', '✗', '✗'],
        ['District Manager', '✓', '✗', '✗', '✗'],
        ['Agent', '✓', '✗', '✗', '✗'],
    ]
    docu_table = Table(docu_data, colWidths=[120, 60, 60, 90, 60])
    docu_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(docu_table)
    elements.append(Paragraph("<b>Key:</b> Documents are team-scoped. Team A cannot see Team B's documents.", body_style))
    
    # Section 8: Diagnostics
    elements.append(PageBreak())
    elements.append(Paragraph("8. Diagnostics & Recovery", h1_style))
    elements.append(Paragraph("Access: <b>Admin → Diagnostics</b>", body_style))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("<b>A. Diagnose Interviews</b> - Finds interviews with deleted owners. Click Fix to reassign.", body_style))
    elements.append(Paragraph("<b>B. Find Unassigned Users</b> - Users without a team cannot log in. Assign them to fix.", body_style))
    elements.append(Paragraph("<b>C. Diagnose Orphaned Activities</b> - Finds activities missing team_id. Click Fix to repair.", body_style))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("<b>When to run:</b> After bulk imports, if rollups seem wrong, if users report missing data.", body_style))
    
    # Section 9: Common Issues
    elements.append(Paragraph("9. Common Issues & Solutions", h1_style))
    
    elements.append(Paragraph("<b>User can't log in</b>", h2_style))
    elements.append(Paragraph("Check: Is user assigned to a team? Is password correct? Is status 'active'?", body_style))
    elements.append(Paragraph("Fix: Assign user to team or reset password.", body_style))
    
    elements.append(Paragraph("<b>Manager sees wrong rollup data</b>", h2_style))
    elements.append(Paragraph("Check: Is hierarchy correct? Are subordinates in same team? Run Diagnostics.", body_style))
    elements.append(Paragraph("Fix: Correct hierarchy or run activity migration.", body_style))
    
    elements.append(Paragraph("<b>Feature tab missing</b>", h2_style))
    elements.append(Paragraph("Check: Is feature enabled for their team? Does role have access?", body_style))
    elements.append(Paragraph("Fix: Enable feature flag for team.", body_style))
    
    elements.append(Paragraph("<b>DocuSphere empty</b>", h2_style))
    elements.append(Paragraph("Check: Is DocuSphere enabled? Is user in correct team? Have docs been uploaded?", body_style))
    elements.append(Paragraph("Fix: Enable feature and/or upload documents.", body_style))
    
    # Footer
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(f"<i>PMA Agent Admin Playbook - Generated: {datetime.now().strftime('%B %d, %Y')}</i>", body_style))
    
    doc.build(elements)
    pdf_content = buffer.getvalue()
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=Admin_Playbook.pdf"}
    )

@api_router.get("/admin/users")
async def get_all_users_admin(current_user: dict = Depends(get_current_user)):
    """Get all users with team info and subordinate counts (super_admin only)"""
    require_super_admin(current_user)
    
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(10000)
    teams = await db.teams.find({}, {"_id": 0}).to_list(1000)
    team_map = {t['id']: t['name'] for t in teams}
    
    # Build a map of manager_id -> count of direct reports
    subordinate_counts = {}
    for user in users:
        manager_id = user.get('manager_id')
        if manager_id:
            subordinate_counts[manager_id] = subordinate_counts.get(manager_id, 0) + 1
    
    for user in users:
        user['team_name'] = team_map.get(user.get('team_id'), 'Unassigned')
        user['subordinate_count'] = subordinate_counts.get(user['id'], 0)
        # Get manager name
        if user.get('manager_id'):
            manager = next((u for u in users if u['id'] == user.get('manager_id')), None)
            user['manager_name'] = manager['name'] if manager else 'Unknown'
        else:
            user['manager_name'] = None
    
    return users

@api_router.post("/admin/users/assign-team")
async def assign_user_to_team(assignment: UserTeamAssignment, current_user: dict = Depends(get_current_user)):
    """Assign a user to a team (super_admin only)"""
    require_super_admin(current_user)
    
    # Verify user exists
    user = await db.users.find_one({"id": assignment.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify team exists
    team = await db.teams.find_one({"id": assignment.team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    update_data = {"team_id": assignment.team_id}
    if assignment.role:
        update_data["role"] = assignment.role
    if assignment.manager_id:
        update_data["manager_id"] = assignment.manager_id
    
    await db.users.update_one({"id": assignment.user_id}, {"$set": update_data})
    
    return {"message": f"User assigned to team '{team['name']}' successfully"}

@api_router.post("/admin/users/remove-team")
async def remove_user_from_team(user_id: str, current_user: dict = Depends(get_current_user)):
    """Remove a user from their team (super_admin only)"""
    require_super_admin(current_user)
    
    await db.users.update_one({"id": user_id}, {"$set": {"team_id": None}})
    return {"message": "User removed from team"}

@api_router.post("/admin/migrate-to-teams")
async def migrate_existing_data_to_teams(current_user: dict = Depends(get_current_user)):
    """One-time migration: Create default team and assign all existing data (super_admin only)"""
    require_super_admin(current_user)
    
    # Check if default team exists
    default_team = await db.teams.find_one({"name": "Team Sudbeck"})
    
    if not default_team:
        # Create Team Sudbeck
        default_team = {
            "id": str(uuid.uuid4()),
            "name": "Team Sudbeck",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "settings": {"is_default": True}
        }
        await db.teams.insert_one(default_team)
    
    team_id = default_team['id']
    
    # Collections to migrate
    collections = [
        'users', 'activities', 'interviews', 'suitability_forms', 
        'npa_agents', 'new_face_customers', 'recruits', 'team_goals',
        'goals', 'pma_bonuses', 'docusphere_documents', 'docusphere_folders', 'invites'
    ]
    
    results = {}
    for collection_name in collections:
        collection = db[collection_name]
        # Update all documents that don't have team_id
        result = await collection.update_many(
            {"team_id": {"$exists": False}},
            {"$set": {"team_id": team_id}}
        )
        results[collection_name] = result.modified_count
    
    return {
        "message": "Migration completed successfully",
        "team_id": team_id,
        "team_name": "Team Sudbeck",
        "migrated_records": results
    }

@api_router.post("/admin/create-super-admin")
async def create_super_admin_user(user_data: UserCreate, admin_secret: str):
    """Create a super admin user (requires secret key)"""
    # Simple security - in production, use a more secure method
    if admin_secret != os.environ.get('JWT_SECRET_KEY', 'your-super-secret-jwt-key-change-in-production'):
        raise HTTPException(status_code=403, detail="Invalid admin secret")
    
    # Case-insensitive email check
    existing = await db.users.find_one({"email": {"$regex": f"^{user_data.email}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    hashed_pw = hash_password(user_data.password)
    
    admin_user = {
        "id": str(uuid.uuid4()),
        "email": user_data.email,
        "name": user_data.name,
        "role": "super_admin",
        "team_id": None,  # Super admins don't belong to a team
        "manager_id": None,
        "password_hash": hashed_pw,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(admin_user)
    admin_user.pop('_id', None)
    admin_user.pop('password_hash', None)
    
    return {"message": "Super admin created successfully", "user": admin_user}

class AdminUserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: str  # state_manager, regional_manager, district_manager, agent
    team_id: str
    manager_id: Optional[str] = None

@api_router.post("/admin/users")
async def admin_create_user(user_data: AdminUserCreate, current_user: dict = Depends(get_current_user)):
    """Create a new user directly into a specific team (super_admin only)"""
    require_super_admin(current_user)
    
    # Validate team exists
    team = await db.teams.find_one({"id": user_data.team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Check if email already exists (case-insensitive)
    existing = await db.users.find_one({"email": {"$regex": f"^{user_data.email}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Validate role
    valid_roles = ['state_manager', 'regional_manager', 'district_manager', 'agent']
    if user_data.role not in valid_roles:
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {valid_roles}")
    
    # If manager_id provided, validate it exists and is in the same team
    if user_data.manager_id:
        manager = await db.users.find_one({"id": user_data.manager_id, "team_id": user_data.team_id})
        if not manager:
            raise HTTPException(status_code=404, detail="Manager not found in this team")
    
    hashed_pw = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    new_user = {
        "id": str(uuid.uuid4()),
        "email": user_data.email,
        "name": user_data.name,
        "role": user_data.role,
        "team_id": user_data.team_id,
        "manager_id": user_data.manager_id,
        "password_hash": hashed_pw,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(new_user)
    new_user.pop('_id', None)
    new_user.pop('password_hash', None)
    
    return {"message": "User created successfully", "user": new_user}

@api_router.put("/admin/users/{user_id}/manager")
async def update_user_manager(user_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Update a user's manager_id (super_admin only)"""
    require_super_admin(current_user)
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    manager_id = data.get('manager_id')
    if manager_id:
        manager = await db.users.find_one({"id": manager_id})
        if not manager:
            raise HTTPException(status_code=404, detail="Manager not found")
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"manager_id": manager_id}}
    )
    
    return {"message": "Manager updated successfully"}

@api_router.get("/admin/teams/{team_id}/users")
async def get_team_users(team_id: str, current_user: dict = Depends(get_current_user)):
    """Get all users in a specific team (super_admin only)"""
    require_super_admin(current_user)
    
    users = await db.users.find(
        {"team_id": team_id},
        {"_id": 0, "password_hash": 0}
    ).to_list(1000)
    
    return users

@api_router.get("/admin/teams/{team_id}/hierarchy")
async def get_admin_team_hierarchy(team_id: str, current_user: dict = Depends(get_current_user)):
    """Get the full hierarchy tree for a specific team (super_admin only)"""
    require_super_admin(current_user)
    
    # Get all users in this team
    users = await db.users.find(
        {"team_id": team_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
        {"_id": 0, "password_hash": 0}
    ).to_list(1000)
    
    # Build hierarchy tree
    user_map = {u['id']: u for u in users}
    
    # Find state managers (top of hierarchy - no manager_id or manager not in team)
    roots = []
    for u in users:
        manager_id = u.get('manager_id')
        if not manager_id or manager_id not in user_map:
            roots.append(u)
    
    # Build children map
    children_map = {}
    for u in users:
        manager_id = u.get('manager_id')
        if manager_id and manager_id in user_map:
            if manager_id not in children_map:
                children_map[manager_id] = []
            children_map[manager_id].append(u)
    
    def build_tree(user):
        node = {
            "id": user['id'],
            "name": user.get('name', 'Unknown'),
            "email": user.get('email', ''),
            "role": user.get('role', 'unknown'),
            "manager_id": user.get('manager_id'),
            "children": []
        }
        if user['id'] in children_map:
            for child in children_map[user['id']]:
                node['children'].append(build_tree(child))
        return node
    
    hierarchy = [build_tree(root) for root in roots]
    
    return {
        "team_id": team_id,
        "total_users": len(users),
        "roots_count": len(roots),
        "hierarchy": hierarchy
    }

class ManagerIdRepairRequest(BaseModel):
    user_id: str
    manager_id: str

@api_router.post("/admin/repair-manager-ids")
async def repair_manager_ids_batch(repairs: List[ManagerIdRepairRequest], current_user: dict = Depends(get_current_user)):
    """
    Batch repair manager_id for multiple users (super_admin only).
    This is a ONE-TIME fix for users created before the manager_id bug was fixed.
    """
    require_super_admin(current_user)
    
    results = []
    for repair in repairs:
        # Verify user exists
        user = await db.users.find_one({"id": repair.user_id})
        if not user:
            results.append({"user_id": repair.user_id, "status": "error", "message": "User not found"})
            continue
        
        # Verify manager exists (if provided)
        if repair.manager_id:
            manager = await db.users.find_one({"id": repair.manager_id})
            if not manager:
                results.append({"user_id": repair.user_id, "status": "error", "message": "Manager not found"})
                continue
        
        # Update the manager_id
        await db.users.update_one(
            {"id": repair.user_id},
            {"$set": {"manager_id": repair.manager_id if repair.manager_id else None}}
        )
        results.append({
            "user_id": repair.user_id,
            "user_name": user.get('name'),
            "status": "success",
            "new_manager_id": repair.manager_id
        })
    
    return {
        "message": f"Processed {len(repairs)} manager_id repairs",
        "results": results
    }

@api_router.get("/admin/teams/{team_id}/broken-hierarchy")
async def get_broken_hierarchy_users(team_id: str, current_user: dict = Depends(get_current_user)):
    """
    Find users in a team with broken hierarchy (missing or invalid manager_id).
    This helps identify users that need manager_id repair.
    (super_admin only)
    """
    require_super_admin(current_user)
    
    # Get team info
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Get all users in this team
    users = await db.users.find(
        {"team_id": team_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
        {"_id": 0, "password_hash": 0}
    ).to_list(1000)
    
    user_ids = {u['id'] for u in users}
    
    broken_users = []
    for u in users:
        role = u.get('role', '')
        manager_id = u.get('manager_id')
        
        # State managers don't need a manager_id (they're at the top)
        if role == 'state_manager':
            continue
        
        # All other roles should have a valid manager_id pointing to someone in the same team
        if not manager_id:
            broken_users.append({
                "id": u['id'],
                "name": u.get('name'),
                "email": u.get('email'),
                "role": role,
                "issue": "Missing manager_id"
            })
        elif manager_id not in user_ids:
            broken_users.append({
                "id": u['id'],
                "name": u.get('name'),
                "email": u.get('email'),
                "role": role,
                "manager_id": manager_id,
                "issue": "manager_id points to user not in this team"
            })
    
    # Get potential managers for fixing
    potential_managers = []
    for u in users:
        if u.get('role') in ['state_manager', 'regional_manager', 'district_manager']:
            potential_managers.append({
                "id": u['id'],
                "name": u.get('name'),
                "role": u.get('role')
            })
    
    return {
        "team_name": team.get('name'),
        "team_id": team_id,
        "total_users": len(users),
        "broken_count": len(broken_users),
        "broken_users": broken_users,
        "potential_managers": potential_managers
    }

@api_router.get("/admin/diagnose-interviews")
async def diagnose_interviews(current_user: dict = Depends(get_current_user)):
    """
    Diagnostic endpoint to check interview data integrity (super_admin only).
    Shows total interviews, interviews by team, and orphaned interviews.
    
    SAFETY:
    - Does NOT modify any data
    - Does NOT affect Team Sudbeck
    - Read-only diagnostic
    """
    require_super_admin(current_user)
    
    # Get all interviews
    all_interviews = await db.interviews.find({}, {"_id": 0}).to_list(10000)
    
    # Get all users
    all_users = await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "team_id": 1}).to_list(10000)
    user_map = {u['id']: u for u in all_users}
    
    # Get all teams
    all_teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    team_map = {t['id']: t['name'] for t in all_teams}
    
    # Find Team Sudbeck ID to exclude
    team_sudbeck_id = None
    for t in all_teams:
        if t.get('name') == 'Team Sudbeck':
            team_sudbeck_id = t['id']
            break
    
    # Analyze interviews
    by_team = {}
    orphaned_by_team = {}
    orphaned = []
    
    for interview in all_interviews:
        team_id = interview.get('team_id')
        interviewer_id = interview.get('interviewer_id')
        team_name = team_map.get(team_id, 'No Team')
        
        # Count total by team
        by_team[team_name] = by_team.get(team_name, 0) + 1
        
        # Check if interviewer still exists
        if interviewer_id and interviewer_id not in user_map:
            # Skip Team Sudbeck for orphaned analysis
            if team_id == team_sudbeck_id:
                continue
                
            orphaned_by_team[team_name] = orphaned_by_team.get(team_name, 0) + 1
            orphaned.append({
                "interview_id": interview.get('id'),
                "candidate_name": interview.get('candidate_name'),
                "interviewer_id": interviewer_id,
                "team_id": team_id,
                "team_name": team_name,
                "interview_date": interview.get('interview_date'),
                "issue": "Interviewer user no longer exists"
            })
    
    return {
        "total_interviews": len(all_interviews),
        "interviews_by_team": by_team,
        "orphaned_total": len(orphaned),
        "orphaned_by_team": orphaned_by_team,
        "orphaned_interviews": orphaned[:50],  # Limit to first 50 for display
        "team_sudbeck_excluded": True,
        "message": "Orphaned interviews have interviewer_id pointing to deleted users. Team Sudbeck is excluded from orphaned analysis."
    }

@api_router.post("/admin/fix-orphaned-interviews")
async def fix_orphaned_interviews(current_user: dict = Depends(get_current_user)):
    """
    Fix orphaned interviews by reassigning them to the team's State Manager (super_admin only).
    
    SAFETY:
    - Super admin only
    - Does NOT modify team_id
    - Does NOT affect Team Sudbeck
    - Only updates interviews where interviewer_id points to non-existent user
    - Preserves original_interviewer_id for audit trail
    """
    require_super_admin(current_user)
    
    # Get all teams
    all_teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    team_map = {t['id']: t['name'] for t in all_teams}
    
    # Find Team Sudbeck ID to exclude
    team_sudbeck_id = None
    for t in all_teams:
        if t.get('name') == 'Team Sudbeck':
            team_sudbeck_id = t['id']
            break
    
    # Get all interviews
    all_interviews = await db.interviews.find({}, {"_id": 0}).to_list(10000)
    
    # Get all users
    all_users = await db.users.find({}, {"_id": 0}).to_list(10000)
    user_ids = {u['id'] for u in all_users}
    
    # Group state managers by team
    state_managers_by_team = {}
    for u in all_users:
        if u.get('role') == 'state_manager' and u.get('team_id'):
            state_managers_by_team[u['team_id']] = u
    
    fixed_by_team = {}
    fixed = []
    skipped_sudbeck = 0
    skipped_no_sm = 0
    
    for interview in all_interviews:
        interviewer_id = interview.get('interviewer_id')
        team_id = interview.get('team_id')
        
        # Check if interviewer no longer exists
        if interviewer_id and interviewer_id not in user_ids:
            # Skip Team Sudbeck - do not modify
            if team_id == team_sudbeck_id:
                skipped_sudbeck += 1
                continue
            
            # Find state manager for this team
            sm = state_managers_by_team.get(team_id)
            if sm:
                team_name = team_map.get(team_id, 'Unknown')
                
                # Preserve original interviewer_id for audit trail
                await db.interviews.update_one(
                    {"id": interview.get('id')},
                    {"$set": {
                        "interviewer_id": sm['id'],
                        "original_interviewer_id": interviewer_id,
                        "reassigned_at": datetime.now(timezone.utc).isoformat(),
                        "reassigned_by": current_user['id']
                    }}
                )
                
                fixed_by_team[team_name] = fixed_by_team.get(team_name, 0) + 1
                fixed.append({
                    "interview_id": interview.get('id'),
                    "candidate_name": interview.get('candidate_name'),
                    "team_name": team_name,
                    "original_interviewer_id": interviewer_id,
                    "new_interviewer": sm['name'],
                    "new_interviewer_id": sm['id']
                })
            else:
                skipped_no_sm += 1
    
    return {
        "message": f"Fixed {len(fixed)} orphaned interviews",
        "fixed_total": len(fixed),
        "fixed_by_team": fixed_by_team,
        "skipped_team_sudbeck": skipped_sudbeck,
        "skipped_no_state_manager": skipped_no_sm,
        "team_sudbeck_protected": True,
        "audit_trail": "original_interviewer_id preserved on each fixed interview",
        "details": fixed[:30]  # Limit details for response size
    }

@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """
    Delete a user (super_admin only). 
    Use this to remove duplicate users or clean up bad data.
    """
    require_super_admin(current_user)
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent deleting yourself
    if user_id == current_user['id']:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    # Delete user and their activities
    await db.users.delete_one({"id": user_id})
    deleted_activities = await db.activities.delete_many({"user_id": user_id})
    
    return {
        "message": f"User '{user.get('name')}' deleted successfully",
        "activities_deleted": deleted_activities.deleted_count
    }

class AdminUserUpdate(BaseModel):
    name: str = None
    email: str = None
    role: str = None
    team_id: str = None
    manager_id: str = None

@api_router.put("/admin/users/{user_id}")
async def admin_update_user(user_id: str, update_data: AdminUserUpdate, current_user: dict = Depends(get_current_user)):
    """
    Update user details (super_admin only).
    Can update name, email, role, team_id, or manager_id.
    """
    require_super_admin(current_user)
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Build update dict with only provided fields
    update_dict = {}
    if update_data.name:
        update_dict["name"] = update_data.name
    if update_data.email:
        # Check if email is already taken by another user (case-insensitive)
        existing = await db.users.find_one({
            "email": {"$regex": f"^{update_data.email}$", "$options": "i"},
            "id": {"$ne": user_id}
        })
        if existing:
            raise HTTPException(status_code=400, detail="Email already in use by another user")
        update_dict["email"] = update_data.email
    if update_data.role:
        update_dict["role"] = update_data.role
    if update_data.team_id:
        update_dict["team_id"] = update_data.team_id
    if update_data.manager_id is not None:
        update_dict["manager_id"] = update_data.manager_id if update_data.manager_id else None
    
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    await db.users.update_one({"id": user_id}, {"$set": update_dict})
    
    return {
        "message": f"User '{user.get('name')}' updated successfully",
        "updated_fields": list(update_dict.keys())
    }

@api_router.get("/admin/diagnose-unassigned-users")
async def diagnose_unassigned_users(current_user: dict = Depends(get_current_user)):
    """
    Find all users who don't have a team_id assigned (super_admin only).
    These users will get the "Access denied - not assigned to team" error.
    
    SAFETY:
    - Does NOT modify any data
    - Read-only diagnostic
    """
    require_super_admin(current_user)
    
    # Find users without team_id or with null/empty team_id
    unassigned_users = await db.users.find(
        {
            "$or": [
                {"team_id": {"$exists": False}},
                {"team_id": None},
                {"team_id": ""},
                {"team_id": "None"}
            ],
            "role": {"$nin": ["super_admin"]}  # super_admin doesn't need a team
        },
        {"_id": 0, "password_hash": 0}
    ).to_list(1000)
    
    # Get all teams for reference - include all fields to debug visibility issues
    all_teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    
    # Sort teams alphabetically by name for easier finding
    all_teams_sorted = sorted(all_teams, key=lambda t: t.get('name', '').lower())
    
    return {
        "unassigned_count": len(unassigned_users),
        "unassigned_users": [
            {
                "id": u.get('id'),
                "email": u.get('email'),
                "name": u.get('name'),
                "role": u.get('role'),
                "status": u.get('status', 'active'),
                "current_team_id": u.get('team_id')
            }
            for u in unassigned_users
        ],
        "available_teams": [{"id": t.get('id'), "name": t.get('name')} for t in all_teams_sorted],
        "all_teams_debug": all_teams_sorted,  # Full team data for debugging
        "message": "These users cannot access the app until assigned to a team. Use /admin/fix-unassigned-users to assign them."
    }

class BulkTeamAssignment(BaseModel):
    user_ids: List[str]
    team_id: str
    set_manager_id: Optional[str] = None  # Optional: also set their manager

@api_router.post("/admin/fix-unassigned-users")
async def fix_unassigned_users(assignment: BulkTeamAssignment, current_user: dict = Depends(get_current_user)):
    """
    Bulk assign team_id to users who don't have one (super_admin only).
    Optionally also sets their manager_id.
    
    SAFETY:
    - Only updates users in the provided list
    - Validates team exists before assignment
    """
    require_super_admin(current_user)
    
    # Validate team exists
    team = await db.teams.find_one({"id": assignment.team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Validate manager if provided
    if assignment.set_manager_id:
        manager = await db.users.find_one({"id": assignment.set_manager_id}, {"_id": 0})
        if not manager:
            raise HTTPException(status_code=404, detail="Manager not found")
    
    results = []
    for user_id in assignment.user_ids:
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            results.append({"user_id": user_id, "status": "error", "message": "User not found"})
            continue
        
        update_data = {"team_id": assignment.team_id}
        if assignment.set_manager_id:
            update_data["manager_id"] = assignment.set_manager_id
        
        await db.users.update_one({"id": user_id}, {"$set": update_data})
        results.append({
            "user_id": user_id,
            "email": user.get('email'),
            "status": "success",
            "assigned_to_team": team.get('name')
        })
    
    success_count = len([r for r in results if r['status'] == 'success'])
    
    return {
        "message": f"Assigned {success_count} of {len(assignment.user_ids)} users to team '{team.get('name')}'",
        "results": results
    }

@api_router.post("/admin/auto-repair-all-teams")
async def auto_repair_all_teams(current_user: dict = Depends(get_current_user)):
    """
    Automatically repair manager_id for ALL teams (except Team Sudbeck).
    This is a ONE-CLICK fix that:
    - Finds all users with missing/invalid manager_id
    - Assigns them to the appropriate manager based on role hierarchy
    - Does NOT touch team_id, does NOT reset users
    (super_admin only)
    """
    require_super_admin(current_user)
    
    # Get all teams except Team Sudbeck
    teams = await db.teams.find({"name": {"$ne": "Team Sudbeck"}}, {"_id": 0}).to_list(100)
    
    total_repaired = 0
    team_results = []
    
    for team in teams:
        team_id = team['id']
        team_name = team.get('name', 'Unknown')
        
        # Get all users in this team
        users = await db.users.find(
            {"team_id": team_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
            {"_id": 0, "password_hash": 0}
        ).to_list(1000)
        
        if not users:
            team_results.append({"team": team_name, "repaired": 0, "message": "No users found"})
            continue
        
        user_ids = {u['id'] for u in users}
        
        # Find managers by role
        state_managers = [u for u in users if u.get('role') == 'state_manager']
        regional_managers = [u for u in users if u.get('role') == 'regional_manager']
        district_managers = [u for u in users if u.get('role') == 'district_manager']
        
        if not state_managers:
            team_results.append({"team": team_name, "repaired": 0, "message": "No state manager found - skipped"})
            continue
        
        # Default state manager for this team
        default_sm = state_managers[0]
        
        repairs_made = 0
        
        # Find and fix broken users
        for u in users:
            role = u.get('role', '')
            manager_id = u.get('manager_id')
            
            # State managers don't need a manager_id
            if role == 'state_manager':
                continue
            
            # Check if manager_id is missing or invalid
            needs_repair = False
            new_manager_id = None
            
            if not manager_id or manager_id not in user_ids:
                needs_repair = True
                
                # Smart assignment based on role
                if role == 'regional_manager':
                    # Regional managers report to state manager
                    new_manager_id = default_sm['id']
                elif role == 'district_manager':
                    # District managers report to regional manager (or state if none)
                    if regional_managers:
                        new_manager_id = regional_managers[0]['id']
                    else:
                        new_manager_id = default_sm['id']
                elif role == 'agent':
                    # Agents report to district manager (or regional, or state)
                    if district_managers:
                        new_manager_id = district_managers[0]['id']
                    elif regional_managers:
                        new_manager_id = regional_managers[0]['id']
                    else:
                        new_manager_id = default_sm['id']
                else:
                    # Default to state manager
                    new_manager_id = default_sm['id']
            
            if needs_repair and new_manager_id:
                await db.users.update_one(
                    {"id": u['id']},
                    {"$set": {"manager_id": new_manager_id}}
                )
                repairs_made += 1
        
        total_repaired += repairs_made
        team_results.append({
            "team": team_name,
            "repaired": repairs_made,
            "total_users": len(users),
            "message": "Repaired successfully" if repairs_made > 0 else "No repairs needed"
        })
    
    return {
        "message": f"Auto-repair complete. Fixed {total_repaired} users across {len(teams)} teams.",
        "total_repaired": total_repaired,
        "teams_processed": len(teams),
        "details": team_results
    }

@api_router.post("/admin/teams/{team_id}/force-rebuild-hierarchy")
async def force_rebuild_team_hierarchy(team_id: str, current_user: dict = Depends(get_current_user)):
    """
    Force rebuild the entire hierarchy for a team (super_admin only).
    This will:
    - Find the State Manager
    - Assign ALL Regional Managers to report to the State Manager
    - Assign ALL District Managers to report to a Regional Manager
    - Assign ALL Agents to report to a District Manager
    
    Use this when auto-repair fails due to deleted users or corrupted manager_ids.
    """
    require_super_admin(current_user)
    
    # Get team info
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Get all active users in this team
    users = await db.users.find(
        {"team_id": team_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
        {"_id": 0, "password_hash": 0}
    ).to_list(1000)
    
    if not users:
        return {"message": "No users found in team", "repaired": 0}
    
    # Group users by role
    state_managers = [u for u in users if u.get('role') == 'state_manager']
    regional_managers = [u for u in users if u.get('role') == 'regional_manager']
    district_managers = [u for u in users if u.get('role') == 'district_manager']
    agents = [u for u in users if u.get('role') == 'agent']
    
    if not state_managers:
        raise HTTPException(status_code=400, detail="No State Manager found in team. Cannot rebuild hierarchy.")
    
    # Use the first state manager as the top of hierarchy
    state_manager = state_managers[0]
    repairs = []
    
    # State Manager should have no manager_id (top of hierarchy)
    if state_manager.get('manager_id'):
        await db.users.update_one({"id": state_manager['id']}, {"$set": {"manager_id": None}})
        repairs.append(f"State Manager {state_manager['name']}: cleared manager_id (top level)")
    
    # All Regional Managers report to State Manager
    for rm in regional_managers:
        if rm.get('manager_id') != state_manager['id']:
            await db.users.update_one({"id": rm['id']}, {"$set": {"manager_id": state_manager['id']}})
            repairs.append(f"Regional Manager {rm['name']}: now reports to {state_manager['name']}")
    
    # All District Managers report to a Regional Manager (first one if multiple)
    target_rm = regional_managers[0] if regional_managers else state_manager
    for dm in district_managers:
        if dm.get('manager_id') != target_rm['id']:
            await db.users.update_one({"id": dm['id']}, {"$set": {"manager_id": target_rm['id']}})
            repairs.append(f"District Manager {dm['name']}: now reports to {target_rm['name']}")
    
    # All Agents report to a District Manager (first one if multiple), or Regional, or State
    target_dm = district_managers[0] if district_managers else (regional_managers[0] if regional_managers else state_manager)
    for agent in agents:
        if agent.get('manager_id') != target_dm['id']:
            await db.users.update_one({"id": agent['id']}, {"$set": {"manager_id": target_dm['id']}})
            repairs.append(f"Agent {agent['name']}: now reports to {target_dm['name']}")
    
    return {
        "message": f"Force rebuilt hierarchy for {team['name']}",
        "team_name": team['name'],
        "state_manager": state_manager['name'],
        "total_users": len(users),
        "repairs_made": len(repairs),
        "details": repairs
    }

# ==================== END ADMIN TEAM MANAGEMENT ====================

# Authentication Routes
@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    # Check if email exists (case-insensitive)
    existing = await db.users.find_one({"email": {"$regex": f"^{user_data.email}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # If invite code provided, validate it
    if user_data.invite_code:
        invite = await db.invites.find_one({"invite_code": user_data.invite_code, "status": "pending"}, {"_id": 0})
        if not invite:
            raise HTTPException(status_code=400, detail="Invalid invite code")
        
        # Use invite details
        user_data.name = invite['name']
        user_data.role = invite['role']
        user_data.manager_id = invite['manager_id']
        # Inherit team_id from invite
        user_data.team_id = invite.get('team_id')
        
        # Mark invite as accepted
        await db.invites.update_one({"invite_code": user_data.invite_code}, {"$set": {"status": "accepted"}})
    
    # Hash password
    hashed_pw = hash_password(user_data.password)
    
    # Create user
    user = User(
        email=user_data.email,
        name=user_data.name,
        role=user_data.role,
        team_id=user_data.team_id,
        manager_id=user_data.manager_id
    )
    
    doc = user.model_dump()
    doc['password_hash'] = hashed_pw
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    # Create token
    token = create_jwt_token(user.id, user.email)
    
    return {
        "token": token,
        "user": {
            "id": user.id,
            "email": user.email,
            "name": user.name,
            "role": user.role,
            "manager_id": user.manager_id
        }
    }

@api_router.post("/new-face-customers")
async def create_new_face_customer(customer: NewFaceCustomerCreate, current_user: dict = Depends(get_current_user)):
    """Add a new face customer record"""
    # Check how many records exist for this user on this date
    count = await db.new_face_customers.count_documents({
        "user_id": current_user['id'],
        "date": customer.date
    })
    
    if count >= 3:
        raise HTTPException(status_code=400, detail="Maximum 3 new face customers per day")
    
    new_customer = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        "user_name": current_user['name'],
        "team_id": current_user.get('team_id'),  # Multi-tenancy
        "date": customer.date,
        "customer_name": customer.customer_name,
        "county": customer.county,
        "policy_amount": customer.policy_amount,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.new_face_customers.insert_one(new_customer)
    return {"message": "New face customer added", "id": new_customer['id']}

@api_router.get("/new-face-customers/my")
async def get_my_new_face_customers(current_user: dict = Depends(get_current_user)):
    """Get my new face customer records"""
    customers = await db.new_face_customers.find(
        {"user_id": current_user['id']},
        {"_id": 0}
    ).sort("date", -1).to_list(1000)
    return customers

@api_router.get("/new-face-customers/date/{date}")
async def get_new_face_customers_by_date(date: str, current_user: dict = Depends(get_current_user)):
    """Get new face customers for a specific date"""
    customers = await db.new_face_customers.find(
        {"user_id": current_user['id'], "date": date},
        {"_id": 0}
    ).to_list(10)
    return customers

@api_router.get("/new-face-customers/all")
async def get_all_new_face_customers(current_user: dict = Depends(get_current_user)):
    """Get new face customers - STRICTLY scoped to current user's team_id.
    
    NO cross-team visibility allowed under any circumstance.
    """
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Check if New Faces sub-tab is enabled for this team
    await check_subtab_access(current_user, 'new_faces')
    
    team_id = current_user.get('team_id')
    
    # CRITICAL: No team_id = no data
    if not team_id:
        return []
    
    # Get all team members recursively (scoped to team)
    async def get_all_team_ids(user_id: str):
        ids = [user_id]
        query = {"manager_id": user_id, "team_id": team_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        subordinates = await db.users.find(query, {"_id": 0, "id": 1}).to_list(1000)
        for sub in subordinates:
            ids.extend(await get_all_team_ids(sub['id']))
        return ids
    
    team_ids = await get_all_team_ids(current_user['id'])
    
    # STRICT: Only exact team_id match
    query = {"$and": [
        {"user_id": {"$in": team_ids}},
        {"team_id": team_id}
    ]}
    
    customers = await db.new_face_customers.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    return customers

@api_router.delete("/new-face-customers/{customer_id}")
async def delete_new_face_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a new face customer record"""
    customer = await db.new_face_customers.find_one({"id": customer_id}, {"_id": 0})
    
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Only owner or managers can delete
    if customer['user_id'] != current_user['id'] and current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.new_face_customers.delete_one({"id": customer_id})
    return {"message": "Customer deleted"}

@api_router.get("/reports/excel/newface/{period}")
async def generate_newface_report(period: str, current_user: dict = Depends(get_current_user)):
    """
    Generate Excel report for new face customers by period.
    Shows customer name, county, policy amount, date, and agent.
    """
    # Check feature access
    await check_feature_access(current_user, "reports")
    
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Access denied")
    
    team_id = current_user.get('team_id')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Use Central Time for date calculations
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    if period == "monthly":
        start_date = today.replace(day=1)
        period_name = start_date.strftime('%B %Y')
    elif period == "quarterly":
        quarter = (today.month - 1) // 3
        start_date = today.replace(month=quarter * 3 + 1, day=1)
        period_name = f"Q{quarter + 1} {today.year}"
    elif period == "yearly":
        start_date = today.replace(month=1, day=1)
        period_name = f"Year {today.year}"
    else:
        raise HTTPException(status_code=400, detail="Invalid period")
    
    # Get all team members recursively (exclude archived) - SCOPED TO TEAM
    async def get_all_team_ids(user_id: str):
        ids = [user_id]
        query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        subordinates = await db.users.find(query, {"_id": 0, "id": 1}).to_list(1000)
        for sub in subordinates:
            ids.extend(await get_all_team_ids(sub['id']))
        return ids
    
    team_ids = await get_all_team_ids(current_user['id'])
    
    # Get new face customers for the period - SCOPED TO TEAM
    query = {"user_id": {"$in": team_ids}, "date": {"$gte": start_date.isoformat()}}
    if team_id:
        query["team_id"] = team_id
    customers = await db.new_face_customers.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "New Face Customers"
    
    # Add title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"New Face Customer Report - {period_name}"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    
    # Add headers
    headers = ["Date", "Agent Name", "Customer Name", "County", "Policy Amount", "Total"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    total_policy = 0
    for row_num, customer in enumerate(customers, 3):
        ws.cell(row=row_num, column=1).value = customer.get('date', '')
        ws.cell(row=row_num, column=2).value = customer.get('user_name', 'Unknown')
        ws.cell(row=row_num, column=3).value = customer.get('customer_name', '')
        ws.cell(row=row_num, column=4).value = customer.get('county', '')
        ws.cell(row=row_num, column=5).value = customer.get('policy_amount', 0)
        total_policy += customer.get('policy_amount', 0)
    
    # Add summary row
    if customers:
        summary_row = len(customers) + 3
        ws.cell(row=summary_row, column=1).value = "TOTAL"
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row, column=2).value = f"{len(customers)} Customers"
        ws.cell(row=summary_row, column=5).value = total_policy
        ws.cell(row=summary_row, column=5).font = Font(bold=True)
        
        for col in range(1, 7):
            ws.cell(row=summary_row, column=col).fill = PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid")
    
    # Auto-size columns
    for col_idx in range(1, 7):
        column_letter = ws.cell(row=2, column=col_idx).column_letter
        max_length = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Format currency column
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=5)
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.00'
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"new_face_report_{period}_{today.isoformat()}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/excel/{period}")
async def generate_excel_report(period: str, current_user: dict = Depends(get_current_user)):
    """
    Generate Excel report for weekly, monthly, quarterly, or yearly data.
    Returns Excel file with one row per team member showing totals.
    """
    # Check feature access
    await check_feature_access(current_user, "reports")
    
    team_id = current_user.get('team_id')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Use Central Time for date calculations
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    if period == "weekly":
        start_date = today - timedelta(days=today.weekday())
        period_name = f"Week of {start_date.strftime('%B %d, %Y')}"
    elif period == "monthly":
        start_date = today.replace(day=1)
        period_name = start_date.strftime('%B %Y')
    elif period == "quarterly":
        quarter = (today.month - 1) // 3
        start_date = today.replace(month=quarter * 3 + 1, day=1)
        period_name = f"Q{quarter + 1} {today.year}"
    elif period == "yearly":
        start_date = today.replace(month=1, day=1)
        period_name = f"Year {today.year}"
    else:
        raise HTTPException(status_code=400, detail="Invalid period")
    
    # Get all subordinates recursively - SCOPED TO TEAM
    async def get_all_team_members(user_id: str):
        members = []
        user_doc = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
        if user_doc:
            # Convert to dict to ensure it's serializable
            user = dict(user_doc)
            members.append(user)
            query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
            if team_id:
                query["team_id"] = team_id
            subordinates = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
            for sub in subordinates:
                sub_members = await get_all_team_members(sub['id'])
                members.extend(sub_members)
        return members
    
    team_members = await get_all_team_members(current_user['id'])
    
    # Get activities for each team member - SCOPED TO TEAM
    report_data = []
    for member in team_members:
        act_query = {"user_id": member['id'], "date": {"$gte": start_date.isoformat()}}
        if team_id:
            act_query["team_id"] = team_id
        activities = await db.activities.find(act_query, {"_id": 0}).to_list(1000)
        
        totals = {
            "name": str(member.get('name', 'Unknown')),
            "email": str(member.get('email', '')),
            "role": str(member.get('role', 'unknown')).replace('_', ' ').title(),
            "contacts": float(sum(a.get('contacts', 0) for a in activities)),
            "appointments": float(sum(a.get('appointments', 0) for a in activities)),
            "presentations": float(sum(a.get('presentations', 0) for a in activities)),
            "referrals": int(sum(a.get('referrals', 0) for a in activities)),
            "testimonials": int(sum(a.get('testimonials', 0) for a in activities)),
            "sales": int(sum(a.get('sales', 0) for a in activities)),
            "new_face_sold": float(sum(a.get('new_face_sold', 0) for a in activities)),
            "premium": float(sum(a.get('premium', 0) for a in activities))
        }
        report_data.append(totals)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{period.capitalize()} Report"
    
    # Add title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"Sales Activity Report - {period_name}"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Add headers
    headers = ["Name", "Email", "Role", "Contacts", "Appointments", "Presentations", 
               "Referrals", "Testimonials", "Sales", "New Face Sold", "Total Premium"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_num, data in enumerate(report_data, 3):
        ws.cell(row=row_num, column=1).value = data['name']
        ws.cell(row=row_num, column=2).value = data['email']
        ws.cell(row=row_num, column=3).value = data['role']
        ws.cell(row=row_num, column=4).value = data['contacts']
        ws.cell(row=row_num, column=5).value = data['appointments']
        ws.cell(row=row_num, column=6).value = data['presentations']
        ws.cell(row=row_num, column=7).value = data['referrals']
        ws.cell(row=row_num, column=8).value = data['testimonials']
        ws.cell(row=row_num, column=9).value = data['sales']
        ws.cell(row=row_num, column=10).value = data['new_face_sold']
        ws.cell(row=row_num, column=11).value = f"${data['premium']:.2f}"
    
    # Auto-size columns
    for col_idx in range(1, 12):  # 11 columns (A to K)
        column_letter = ws.cell(row=2, column=col_idx).column_letter
        max_length = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"sales_report_{period}_{today.isoformat()}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/manager-hierarchy/{manager_id}")
async def get_manager_hierarchy_report(manager_id: str, period: str, current_user: dict = Depends(get_current_user), date: str = None, month: str = None, quarter: str = None, year: str = None):
    """
    Get full hierarchy report for a specific manager.
    Shows the manager + all their direct and indirect reports with totals for the period.
    manager_id: ID of the manager to get hierarchy for
    period: 'daily', 'monthly', 'quarterly', or 'yearly'
    date: ISO format date string for daily reports (YYYY-MM-DD)
    month: Optional - specific month for monthly reports in YYYY-MM format
    quarter: Optional - specific quarter for quarterly reports in YYYY-Q1 format
    year: Optional - specific year for yearly reports in YYYY format
    """
    # Check feature access
    await check_feature_access(current_user, "reports")
    
    team_id = current_user.get('team_id')
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only Managers can access hierarchy reports")
    
    # Verify the requested manager is in current user's hierarchy - SCOPED TO TEAM
    async def get_all_subordinates(user_id: str):
        members = []
        query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        subordinates = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
        for sub in subordinates:
            members.append(sub)
            sub_members = await get_all_subordinates(sub['id'])
            members.extend(sub_members)
        return members
    
    user_hierarchy = await get_all_subordinates(current_user['id'])
    user_hierarchy.insert(0, current_user)  # Include self
    
    target_manager = None
    for member in user_hierarchy:
        if member['id'] == manager_id:
            target_manager = member
            break
    
    if not target_manager:
        raise HTTPException(status_code=403, detail="Manager not found in your hierarchy")
    
    # Get the target manager's hierarchy
    manager_hierarchy = await get_all_subordinates(manager_id)
    manager_hierarchy.insert(0, target_manager)  # Include the manager themselves
    
    # Calculate date range based on period
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    if period == "daily":
        if date:
            start_date = datetime.strptime(date, '%Y-%m-%d').date()
            period_name = f"Daily - {start_date.strftime('%B %d, %Y')}"
            date_filter = start_date.isoformat()
        else:
            start_date = today
            period_name = f"Daily - {start_date.strftime('%B %d, %Y')}"
            date_filter = start_date.isoformat()
    elif period == "monthly":
        if month:
            # Parse the selected month (YYYY-MM format)
            try:
                year_str, month_num = month.split('-')
                start_date = date_class(int(year_str), int(month_num), 1)
                period_name = f"Month of {start_date.strftime('%B %Y')}"
                date_filter = {"$gte": start_date.isoformat()}
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        else:
            # Default to current month
            start_date = today.replace(day=1)
            period_name = f"Month of {start_date.strftime('%B %Y')}"
            date_filter = {"$gte": start_date.isoformat()}
    elif period == "quarterly":
        if quarter:
            # Parse the selected quarter (YYYY-Q1 format)
            try:
                year_str, quarter_str = quarter.split('-Q')
                quarter_num = int(quarter_str)
                if quarter_num < 1 or quarter_num > 4:
                    raise ValueError("Quarter must be 1-4")
                start_date = date_class(int(year_str), (quarter_num - 1) * 3 + 1, 1)
                period_name = f"Q{quarter_num} {year_str}"
                date_filter = {"$gte": start_date.isoformat()}
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid quarter format. Use YYYY-Q1")
        else:
            # Default to current quarter
            quarter_num = (today.month - 1) // 3
            start_date = today.replace(month=quarter_num * 3 + 1, day=1)
            period_name = f"Q{quarter_num + 1} {today.year}"
            date_filter = {"$gte": start_date.isoformat()}
    elif period == "yearly":
        if year:
            # Parse the selected year
            try:
                year_num = int(year)
                start_date = date_class(year_num, 1, 1)
                period_name = f"Year {year_num}"
                date_filter = {"$gte": start_date.isoformat()}
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid year format. Use YYYY")
        else:
            # Default to current year
            start_date = today.replace(month=1, day=1)
            period_name = f"Year {today.year}"
            date_filter = {"$gte": start_date.isoformat()}
    else:
        raise HTTPException(status_code=400, detail="Invalid period. Use 'daily', 'monthly', 'quarterly', or 'yearly'")
    
    # Build hierarchy structure with totals
    hierarchy_data = []
    
    for member in manager_hierarchy:
        # Get activities for the period - SCOPED TO TEAM
        act_query = {"user_id": member['id'], "date": date_filter}
        if team_id:
            act_query["team_id"] = team_id
        
        if period == "daily":
            activities = await db.activities.find(act_query, {"_id": 0}).to_list(1000)
        else:
            activities = await db.activities.find(act_query, {"_id": 0}).to_list(10000)
        
        # Calculate totals
        if period == "daily":
            # For daily, show the single day's data
            activity = activities[0] if activities else {}
            totals = {
                "contacts": activity.get('contacts', 0),
                "appointments": activity.get('appointments', 0),
                "presentations": activity.get('presentations', 0),
                "referrals": activity.get('referrals', 0),
                "testimonials": activity.get('testimonials', 0),
                "sales": activity.get('sales', 0),
                "new_face_sold": activity.get('new_face_sold', 0),
                "premium": activity.get('premium', 0)
            }
        else:
            # For periods, sum all activities
            totals = {
                "contacts": sum(a.get('contacts', 0) for a in activities),
                "appointments": sum(a.get('appointments', 0) for a in activities),
                "presentations": sum(a.get('presentations', 0) for a in activities),
                "referrals": sum(a.get('referrals', 0) for a in activities),
                "testimonials": sum(a.get('testimonials', 0) for a in activities),
                "sales": sum(a.get('sales', 0) for a in activities),
                "new_face_sold": sum(a.get('new_face_sold', 0) for a in activities),
                "premium": sum(a.get('premium', 0) for a in activities)
            }
        
        # Determine relationship to target manager
        if member['id'] == manager_id:
            relationship = "Manager"
        elif member.get('manager_id') == manager_id:
            relationship = "Direct Report"
        else:
            relationship = "Indirect Report"
        
        hierarchy_data.append({
            "id": member['id'],
            "name": member.get('name', 'Unknown'),
            "email": member.get('email', ''),
            "role": member.get('role', 'unknown').replace('_', ' ').title(),
            "relationship": relationship,
            "manager_id": member.get('manager_id'),
            **totals
        })
    
    return {
        "manager_name": target_manager.get('name', 'Unknown'),
        "manager_role": target_manager.get('role', 'unknown').replace('_', ' ').title(),
        "period": period,
        "period_name": period_name,
        "hierarchy_data": hierarchy_data,
        "total_members": len(hierarchy_data)
    }

@api_router.get("/reports/managers")
async def get_available_managers(current_user: dict = Depends(get_current_user)):
    """
    Get list of managers available for individual reporting.
    Returns users under the current user's hierarchy, scoped by team.
    
    - super_admin: Can see all users across all teams
    - state_manager: Can see ALL users in their team (they're at the top)
    - regional_manager/district_manager: ONLY their team's downline
    """
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only Managers can access manager list")
    
    team_id = current_user.get('team_id')
    
    # Helper function to get all subordinates recursively (exclude archived), scoped by team
    async def get_subordinates_with_info(user_id: str, filter_team_id: str = None):
        members = []
        query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if filter_team_id:
            query["team_id"] = filter_team_id
        
        subordinates = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
        for sub in subordinates:
            members.append(sub)
            sub_members = await get_subordinates_with_info(sub['id'], filter_team_id)
            members.extend(sub_members)
        return members
    
    # super_admin uses same team scoping as state_manager on product pages
    # Cross-team visibility is ONLY available in Admin endpoints
    if current_user['role'] in ['super_admin', 'state_manager']:
        # State manager and super_admin see ALL users in their team (they're at the top)
        if not team_id:
            return {"managers": []}
        team_members = await db.users.find(
            {"team_id": team_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
            {"_id": 0, "password_hash": 0}
        ).to_list(1000)
        # Remove self from list (will be added back)
        team_members = [m for m in team_members if m['id'] != current_user['id']]
    else:
        # For regional/district managers, enforce team + hierarchy filtering
        if not team_id:
            return {"managers": []}
        team_members = await get_subordinates_with_info(current_user['id'], team_id)
    
    team_members.insert(0, current_user)  # Include self at the top
    
    # Format for dropdown display
    manager_list = []
    for member in team_members:
        manager_list.append({
            "id": member.get('id'),
            "name": member.get('name', 'Unknown'),
            "email": member.get('email', ''),
            "role": member.get('role', 'unknown').replace('_', ' ').title()
        })
    
    return {"managers": manager_list}

@api_router.get("/reports/daily/{report_type}")
async def get_daily_report(report_type: str, date: str, current_user: dict = Depends(get_current_user), user_id: str = None):
    """
    Get daily report for a specific date.
    report_type: 'individual', 'team', or 'organization'
    date: ISO format date string (YYYY-MM-DD)
    user_id: Optional - specific user ID for individual reports (defaults to all team members)
    Returns JSON data for on-screen viewing
    """
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only Managers (State, Regional, District) can access daily reports")
    
    team_id = current_user.get('team_id')
    
    # Validate date format - keep it simple like other endpoints
    try:
        # Simple validation - just ensure format is correct
        datetime.strptime(date, '%Y-%m-%d')
        report_date = date  # Use the date string directly for querying
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Helper function to get all subordinates recursively (exclude archived) - SCOPED TO TEAM
    async def get_all_subordinates(user_id: str):
        members = []
        query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        subordinates = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
        for sub in subordinates:
            members.append(sub)
            sub_members = await get_all_subordinates(sub['id'])
            members.extend(sub_members)
        return members
    
    if report_type == "individual":
        # Get all team members under current user
        team_members = await get_all_subordinates(current_user['id'])
        team_members.insert(0, current_user)  # Include self
        
        # If user_id specified, filter to just that user
        if user_id:
            # Verify the requested user is in the hierarchy
            target_user = None
            for member in team_members:
                if member['id'] == user_id:
                    target_user = member
                    break
            
            if not target_user:
                raise HTTPException(status_code=403, detail="User not found in your hierarchy")
            
            team_members = [target_user]  # Show only the selected user
        
        report_data = []
        for member in team_members:
            # SCOPED TO TEAM
            act_query = {"user_id": member['id'], "date": report_date}
            if team_id:
                act_query["team_id"] = team_id
            activity = await db.activities.find_one(act_query, {"_id": 0})
            
            report_data.append({
                "name": member.get('name', 'Unknown'),
                "email": member.get('email', ''),
                "role": member.get('role', 'unknown').replace('_', ' ').title(),
                "contacts": activity.get('contacts', 0) if activity else 0,
                "appointments": activity.get('appointments', 0) if activity else 0,
                "presentations": activity.get('presentations', 0) if activity else 0,
                "referrals": activity.get('referrals', 0) if activity else 0,
                "testimonials": activity.get('testimonials', 0) if activity else 0,
                "sales": activity.get('sales', 0) if activity else 0,
                "new_face_sold": activity.get('new_face_sold', 0) if activity else 0,
                "premium": activity.get('premium', 0) if activity else 0
            })
        
        return {
            "report_type": "individual",
            "date": report_date,
            "data": report_data,
            "selected_user": user_id
        }
    
    elif report_type == "team":
        # Get all team members - show each individual person, not aggregated teams
        target_manager = None
        if user_id:
            # Get all subordinates to verify the user is in the hierarchy
            all_subordinates = await get_all_subordinates(current_user['id'])
            all_subordinates.insert(0, current_user)  # Include self
            
            # Verify the requested user is in the hierarchy
            for member in all_subordinates:
                if member['id'] == user_id:
                    target_manager = member
                    break
            
            if not target_manager:
                raise HTTPException(status_code=403, detail="Manager not found in your hierarchy")
        
        # Get all team members (individuals) under the viewing user or selected manager
        base_user = target_manager if target_manager else current_user
        all_team_members = await get_all_subordinates(base_user['id'])
        all_team_members.insert(0, base_user)  # Include the manager themselves first
        
        report_data = []
        
        # Add each individual team member's data - SCOPED TO TEAM
        for member in all_team_members:
            act_query = {"user_id": member['id'], "date": report_date}
            if team_id:
                act_query["team_id"] = team_id
            activity = await db.activities.find_one(act_query, {"_id": 0})
            
            member_totals = {
                "contacts": activity.get('contacts', 0) if activity else 0,
                "appointments": activity.get('appointments', 0) if activity else 0,
                "presentations": activity.get('presentations', 0) if activity else 0,
                "referrals": activity.get('referrals', 0) if activity else 0,
                "testimonials": activity.get('testimonials', 0) if activity else 0,
                "sales": activity.get('sales', 0) if activity else 0,
                "new_face_sold": activity.get('new_face_sold', 0) if activity else 0,
                "premium": activity.get('premium', 0) if activity else 0
            }
            
            report_data.append({
                "team_name": member.get('name', 'Unknown'),
                "manager": member.get('name', 'Unknown'),
                "role": member.get('role', 'unknown').replace('_', ' ').title(),
                **member_totals
            })
        
        return {
            "report_type": "team",
            "date": report_date,
            "data": report_data,
            "selected_user": user_id
        }
    
    elif report_type == "organization":
        # Get all organization members
        all_members = await get_all_subordinates(current_user['id'])
        all_members.insert(0, current_user)  # Include self
        
        # Aggregate organization totals - SCOPED TO TEAM
        org_totals = {
            "contacts": 0, "appointments": 0, "presentations": 0,
            "referrals": 0, "testimonials": 0, "sales": 0,
            "new_face_sold": 0, "premium": 0
        }
        
        for member in all_members:
            act_query = {"user_id": member['id'], "date": report_date}
            if team_id:
                act_query["team_id"] = team_id
            activity = await db.activities.find_one(act_query, {"_id": 0})
            
            if activity:
                org_totals["contacts"] += activity.get('contacts', 0)
                org_totals["appointments"] += activity.get('appointments', 0)
                org_totals["presentations"] += activity.get('presentations', 0)
                org_totals["referrals"] += activity.get('referrals', 0)
                org_totals["testimonials"] += activity.get('testimonials', 0)
                org_totals["sales"] += activity.get('sales', 0)
                org_totals["new_face_sold"] += activity.get('new_face_sold', 0)
                org_totals["premium"] += activity.get('premium', 0)
        
        return {
            "report_type": "organization",
            "date": report_date,
            "total_members": len(all_members),
            "data": org_totals
        }
    
    else:
        raise HTTPException(status_code=400, detail="Invalid report type. Use 'individual', 'team', or 'organization'")

@api_router.get("/reports/period/{report_type}")
async def get_period_report(report_type: str, period: str, current_user: dict = Depends(get_current_user), user_id: str = None, month: str = None, quarter: str = None, year: str = None, week_start: str = None, week_end: str = None):
    """
    Get period report (weekly, monthly, quarterly, yearly) for a specific period.
    report_type: 'individual', 'team', or 'organization'
    period: 'weekly', 'monthly', 'quarterly', or 'yearly'
    user_id: Optional - specific user ID for individual reports (defaults to all team members)
    month: Optional - specific month for monthly reports in YYYY-MM format (defaults to current month)
    quarter: Optional - specific quarter for quarterly reports in YYYY-Q1 format (defaults to current quarter)
    year: Optional - specific year for yearly reports in YYYY format (defaults to current year)
    week_start, week_end: Optional - specific week range in YYYY-MM-DD format (defaults to current week)
    Returns JSON data for on-screen viewing
    """
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only Managers (State, Regional, District) can access period reports")
    
    team_id = current_user.get('team_id')
    
    # Use Central Time for date calculations
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    # Calculate date range based on period
    if period == "weekly":
        if week_start and week_end:
            # Use provided week range
            try:
                start_date = datetime.strptime(week_start, '%Y-%m-%d').date()
                end_date = datetime.strptime(week_end, '%Y-%m-%d').date()
                period_name = f"Week of {start_date.strftime('%b %d')} - {end_date.strftime('%b %d, %Y')}"
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid week format. Use YYYY-MM-DD")
        else:
            # Current week (Monday to Sunday)
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            period_name = f"Week of {start_date.strftime('%b %d')} - {end_date.strftime('%b %d, %Y')}"
    elif period == "monthly":
        if month:
            # Parse the selected month (YYYY-MM format)
            try:
                year, month_num = month.split('-')
                start_date = date_class(int(year), int(month_num), 1)
                period_name = f"Month of {start_date.strftime('%B %Y')}"
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        else:
            # Default to current month
            start_date = today.replace(day=1)
            period_name = f"Month of {start_date.strftime('%B %Y')}"
    elif period == "quarterly":
        if quarter:
            # Parse the selected quarter (YYYY-Q1 format)
            try:
                year_str, quarter_str = quarter.split('-Q')
                quarter_num = int(quarter_str)
                if quarter_num < 1 or quarter_num > 4:
                    raise ValueError("Quarter must be 1-4")
                start_date = date_class(int(year_str), (quarter_num - 1) * 3 + 1, 1)
                period_name = f"Q{quarter_num} {year_str}"
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid quarter format. Use YYYY-Q1")
        else:
            # Default to current quarter
            quarter_num = (today.month - 1) // 3
            start_date = today.replace(month=quarter_num * 3 + 1, day=1)
            period_name = f"Q{quarter_num + 1} {today.year}"
    elif period == "yearly":
        if year:
            # Parse the selected year
            try:
                year_num = int(year)
                start_date = date_class(year_num, 1, 1)
                period_name = f"Year {year_num}"
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid year format. Use YYYY")
        else:
            # Default to current year
            start_date = today.replace(month=1, day=1)
            period_name = f"Year {today.year}"
    else:
        raise HTTPException(status_code=400, detail="Invalid period. Use 'weekly', 'monthly', 'quarterly', or 'yearly'")
    
    # Helper function to get all subordinates recursively (exclude archived) - SCOPED TO TEAM
    async def get_all_subordinates(user_id: str):
        members = []
        query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        subordinates = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
        for sub in subordinates:
            members.append(sub)
            sub_members = await get_all_subordinates(sub['id'])
            members.extend(sub_members)
        return members
    
    if report_type == "individual":
        # Get all team members under current user
        team_members = await get_all_subordinates(current_user['id'])
        team_members.insert(0, current_user)  # Include self
        
        # If user_id specified, filter to just that user
        if user_id:
            # Verify the requested user is in the hierarchy
            target_user = None
            for member in team_members:
                if member['id'] == user_id:
                    target_user = member
                    break
            
            if not target_user:
                raise HTTPException(status_code=403, detail="User not found in your hierarchy")
            
            team_members = [target_user]  # Show only the selected user
        
        report_data = []
        for member in team_members:
            # SCOPED TO TEAM
            act_query = {"user_id": member['id'], "date": {"$gte": start_date.isoformat()}}
            if team_id:
                act_query["team_id"] = team_id
            activities = await db.activities.find(act_query, {"_id": 0}).to_list(10000)
            
            totals = {
                "name": member.get('name', 'Unknown'),
                "email": member.get('email', ''),
                "role": member.get('role', 'unknown').replace('_', ' ').title(),
                "contacts": sum(a.get('contacts', 0) for a in activities),
                "appointments": sum(a.get('appointments', 0) for a in activities),
                "presentations": sum(a.get('presentations', 0) for a in activities),
                "referrals": sum(a.get('referrals', 0) for a in activities),
                "testimonials": sum(a.get('testimonials', 0) for a in activities),
                "sales": sum(a.get('sales', 0) for a in activities),
                "new_face_sold": sum(a.get('new_face_sold', 0) for a in activities),
                "premium": sum(a.get('premium', 0) for a in activities)
            }
            report_data.append(totals)
        
        return {
            "report_type": "individual",
            "period": period,
            "period_name": period_name,
            "start_date": start_date.isoformat(),
            "data": report_data,
            "selected_user": user_id
        }
    
    elif report_type == "team":
        # Get all team members - show each individual person, not aggregated teams
        # If user_id specified, show that manager's hierarchy instead
        target_manager = None
        if user_id:
            # Get all subordinates to verify the user is in the hierarchy
            all_subordinates = await get_all_subordinates(current_user['id'])
            all_subordinates.insert(0, current_user)  # Include self
            
            # Verify the requested user is in the hierarchy
            for member in all_subordinates:
                if member['id'] == user_id:
                    target_manager = member
                    break
            
            if not target_manager:
                raise HTTPException(status_code=403, detail="Manager not found in your hierarchy")
        
        # Get all team members (individuals) under the viewing user or selected manager
        base_user = target_manager if target_manager else current_user
        all_team_members = await get_all_subordinates(base_user['id'])
        all_team_members.insert(0, base_user)  # Include the manager themselves first
        
        report_data = []
        
        # Add each individual team member's data - SCOPED TO TEAM
        for member in all_team_members:
            act_query = {"user_id": member['id'], "date": {"$gte": start_date.isoformat()}}
            if team_id:
                act_query["team_id"] = team_id
            activities = await db.activities.find(act_query, {"_id": 0}).to_list(10000)
            
            member_totals = {
                "contacts": sum(a.get('contacts', 0) for a in activities),
                "appointments": sum(a.get('appointments', 0) for a in activities),
                "presentations": sum(a.get('presentations', 0) for a in activities),
                "referrals": sum(a.get('referrals', 0) for a in activities),
                "testimonials": sum(a.get('testimonials', 0) for a in activities),
                "sales": sum(a.get('sales', 0) for a in activities),
                "new_face_sold": sum(a.get('new_face_sold', 0) for a in activities),
                "premium": sum(a.get('premium', 0) for a in activities)
            }
            
            report_data.append({
                "team_name": member.get('name', 'Unknown'),
                "manager": member.get('name', 'Unknown'),
                "role": member.get('role', 'unknown').replace('_', ' ').title(),
                **member_totals
            })
        
        return {
            "report_type": "team",
            "period": period,
            "period_name": period_name,
            "start_date": start_date.isoformat(),
            "data": report_data,
            "selected_user": user_id
        }
    
    elif report_type == "organization":
        # Get all organization members
        all_members = await get_all_subordinates(current_user['id'])
        all_members.insert(0, current_user)  # Include self
        
        # Aggregate organization totals
        org_totals = {
            "contacts": 0, "appointments": 0, "presentations": 0,
            "referrals": 0, "testimonials": 0, "sales": 0,
            "new_face_sold": 0, "premium": 0
        }
        
        # SCOPED TO TEAM
        for member in all_members:
            act_query = {"user_id": member['id'], "date": {"$gte": start_date.isoformat()}}
            if team_id:
                act_query["team_id"] = team_id
            activities = await db.activities.find(act_query, {"_id": 0}).to_list(10000)
            
            for activity in activities:
                org_totals["contacts"] += activity.get('contacts', 0)
                org_totals["appointments"] += activity.get('appointments', 0)
                org_totals["presentations"] += activity.get('presentations', 0)
                org_totals["referrals"] += activity.get('referrals', 0)
                org_totals["testimonials"] += activity.get('testimonials', 0)
                org_totals["sales"] += activity.get('sales', 0)
                org_totals["new_face_sold"] += activity.get('new_face_sold', 0)
                org_totals["premium"] += activity.get('premium', 0)
        
        return {
            "report_type": "organization",
            "period": period,
            "period_name": period_name,
            "start_date": start_date.isoformat(),
            "total_members": len(all_members),
            "data": org_totals
        }
    
    else:
        raise HTTPException(status_code=400, detail="Invalid report type. Use 'individual', 'team', or 'organization'")

@api_router.get("/reports/period/excel/{report_type}")
async def download_period_report_excel(report_type: str, period: str, current_user: dict = Depends(get_current_user), user_id: str = None, month: str = None, quarter: str = None, year: str = None):
    """
    Download period report (monthly, quarterly, yearly) as Excel file.
    report_type: 'individual', 'team', or 'organization'
    period: 'monthly', 'quarterly', or 'yearly'
    user_id: Optional - specific user ID for individual/team reports
    month: Optional - specific month for monthly reports in YYYY-MM format
    quarter: Optional - specific quarter for quarterly reports in YYYY-Q1 format
    year: Optional - specific year for yearly reports in YYYY format
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only Managers (State, Regional, District) can download period reports")
    
    # Get the report data with all parameters
    report_json = await get_period_report(report_type, period, current_user, user_id, month, quarter, year)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{period.capitalize()} {report_type.capitalize()}"
    
    # Add title
    period_name = report_json.get('period_name', f'{period.capitalize()} Report')
    
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"{period.capitalize()} {report_type.capitalize()} Report - {period_name}"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    if report_type == "individual":
        # Add headers
        headers = ["Name", "Email", "Role", "Contacts", "Appointments", "Presentations", 
                   "Referrals", "Testimonials", "Sales", "New Face Sold", "Total Premium"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, data in enumerate(report_json['data'], 3):
            ws.cell(row=row_num, column=1).value = data['name']
            ws.cell(row=row_num, column=2).value = data['email']
            ws.cell(row=row_num, column=3).value = data['role']
            ws.cell(row=row_num, column=4).value = data['contacts']
            ws.cell(row=row_num, column=5).value = data['appointments']
            ws.cell(row=row_num, column=6).value = data['presentations']
            ws.cell(row=row_num, column=7).value = data['referrals']
            ws.cell(row=row_num, column=8).value = data['testimonials']
            ws.cell(row=row_num, column=9).value = data['sales']
            ws.cell(row=row_num, column=10).value = data['new_face_sold']
            ws.cell(row=row_num, column=11).value = data['premium']
        
        # Add totals row for individual reports
        if report_json['data']:
            totals_row = len(report_json['data']) + 4  # +3 for headers, +1 for spacing
            ws.cell(row=totals_row, column=1).value = "TOTALS"
            ws.cell(row=totals_row, column=1).font = Font(bold=True)
            ws.cell(row=totals_row, column=1).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # Calculate totals
            total_contacts = sum(data['contacts'] for data in report_json['data'])
            total_appointments = sum(data['appointments'] for data in report_json['data'])
            total_presentations = sum(data['presentations'] for data in report_json['data'])
            total_referrals = sum(data['referrals'] for data in report_json['data'])
            total_testimonials = sum(data['testimonials'] for data in report_json['data'])
            total_sales = sum(data['sales'] for data in report_json['data'])
            total_new_face = sum(data['new_face_sold'] for data in report_json['data'])
            total_premium = sum(data['premium'] for data in report_json['data'])
            
            ws.cell(row=totals_row, column=4).value = total_contacts
            ws.cell(row=totals_row, column=4).font = Font(bold=True)
            ws.cell(row=totals_row, column=5).value = total_appointments
            ws.cell(row=totals_row, column=5).font = Font(bold=True)
            ws.cell(row=totals_row, column=6).value = total_presentations
            ws.cell(row=totals_row, column=6).font = Font(bold=True)
            ws.cell(row=totals_row, column=7).value = total_referrals
            ws.cell(row=totals_row, column=7).font = Font(bold=True)
            ws.cell(row=totals_row, column=8).value = total_testimonials
            ws.cell(row=totals_row, column=8).font = Font(bold=True)
            ws.cell(row=totals_row, column=9).value = total_sales
            ws.cell(row=totals_row, column=9).font = Font(bold=True)
            ws.cell(row=totals_row, column=10).value = total_new_face
            ws.cell(row=totals_row, column=10).font = Font(bold=True)
            ws.cell(row=totals_row, column=11).value = total_premium
            ws.cell(row=totals_row, column=11).font = Font(bold=True)
    
    elif report_type == "team":
        # Add headers
        headers = ["Team Name", "Manager", "Role", "Contacts", "Appointments", "Presentations", 
                   "Referrals", "Testimonials", "Sales", "New Face Sold", "Total Premium"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, data in enumerate(report_json['data'], 3):
            ws.cell(row=row_num, column=1).value = data['team_name']
            ws.cell(row=row_num, column=2).value = data['manager']
            ws.cell(row=row_num, column=3).value = data['role']
            ws.cell(row=row_num, column=4).value = data['contacts']
            ws.cell(row=row_num, column=5).value = data['appointments']
            ws.cell(row=row_num, column=6).value = data['presentations']
            ws.cell(row=row_num, column=7).value = data['referrals']
            ws.cell(row=row_num, column=8).value = data['testimonials']
            ws.cell(row=row_num, column=9).value = data['sales']
            ws.cell(row=row_num, column=10).value = data['new_face_sold']
            ws.cell(row=row_num, column=11).value = data['premium']
        
        # Add totals row for team reports
        if report_json['data']:
            totals_row = len(report_json['data']) + 4  # +3 for headers, +1 for spacing
            ws.cell(row=totals_row, column=1).value = "TOTALS"
            ws.cell(row=totals_row, column=1).font = Font(bold=True)
            ws.cell(row=totals_row, column=1).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # Calculate totals
            total_contacts = sum(data['contacts'] for data in report_json['data'])
            total_appointments = sum(data['appointments'] for data in report_json['data'])
            total_presentations = sum(data['presentations'] for data in report_json['data'])
            total_referrals = sum(data['referrals'] for data in report_json['data'])
            total_testimonials = sum(data['testimonials'] for data in report_json['data'])
            total_sales = sum(data['sales'] for data in report_json['data'])
            total_new_face = sum(data['new_face_sold'] for data in report_json['data'])
            total_premium = sum(data['premium'] for data in report_json['data'])
            
            ws.cell(row=totals_row, column=4).value = total_contacts
            ws.cell(row=totals_row, column=4).font = Font(bold=True)
            ws.cell(row=totals_row, column=5).value = total_appointments
            ws.cell(row=totals_row, column=5).font = Font(bold=True)
            ws.cell(row=totals_row, column=6).value = total_presentations
            ws.cell(row=totals_row, column=6).font = Font(bold=True)
            ws.cell(row=totals_row, column=7).value = total_referrals
            ws.cell(row=totals_row, column=7).font = Font(bold=True)
            ws.cell(row=totals_row, column=8).value = total_testimonials
            ws.cell(row=totals_row, column=8).font = Font(bold=True)
            ws.cell(row=totals_row, column=9).value = total_sales
            ws.cell(row=totals_row, column=9).font = Font(bold=True)
            ws.cell(row=totals_row, column=10).value = total_new_face
            ws.cell(row=totals_row, column=10).font = Font(bold=True)
            ws.cell(row=totals_row, column=11).value = total_premium
            ws.cell(row=totals_row, column=11).font = Font(bold=True)
    
    elif report_type == "organization":
        # Add headers
        headers = ["Metric", "Total"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        metrics = [
            ("Total Members", report_json['total_members']),
            ("Contacts", report_json['data']['contacts']),
            ("Appointments", report_json['data']['appointments']),
            ("Presentations", report_json['data']['presentations']),
            ("Referrals", report_json['data']['referrals']),
            ("Testimonials", report_json['data']['testimonials']),
            ("Sales", report_json['data']['sales']),
            ("New Face Sold", report_json['data']['new_face_sold']),
            ("Total Premium", report_json['data']['premium'])
        ]
        
        for row_num, (metric, value) in enumerate(metrics, 3):
            ws.cell(row=row_num, column=1).value = metric
            ws.cell(row=row_num, column=2).value = value
    
    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = ws.cell(row=2, column=col_idx).column_letter
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"{period}_{report_type}_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/daily/excel/{report_type}")
async def download_daily_report_excel(report_type: str, date: str, current_user: dict = Depends(get_current_user), user_id: str = None):
    """
    Download daily report as Excel file.
    report_type: 'individual', 'team', or 'organization'
    date: ISO format date string (YYYY-MM-DD)
    user_id: Optional - specific user ID for individual/team reports
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only Managers (State, Regional, District) can download daily reports")
    
    # Get the report data with all parameters
    report_json = await get_daily_report(report_type, date, current_user, user_id)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Daily {report_type.capitalize()}"
    
    # Add title
    report_date_obj = datetime.fromisoformat(date).date()
    date_str = report_date_obj.strftime('%B %d, %Y')
    
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"Daily {report_type.capitalize()} Report - {date_str}"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    if report_type == "individual":
        # Add headers
        headers = ["Name", "Email", "Role", "Contacts", "Appointments", "Presentations", 
                   "Referrals", "Testimonials", "Sales", "New Face Sold", "Total Premium"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, data in enumerate(report_json['data'], 3):
            ws.cell(row=row_num, column=1).value = data['name']
            ws.cell(row=row_num, column=2).value = data['email']
            ws.cell(row=row_num, column=3).value = data['role']
            ws.cell(row=row_num, column=4).value = data['contacts']
            ws.cell(row=row_num, column=5).value = data['appointments']
            ws.cell(row=row_num, column=6).value = data['presentations']
            ws.cell(row=row_num, column=7).value = data['referrals']
            ws.cell(row=row_num, column=8).value = data['testimonials']
            ws.cell(row=row_num, column=9).value = data['sales']
            ws.cell(row=row_num, column=10).value = data['new_face_sold']
            ws.cell(row=row_num, column=11).value = data['premium']
        
        # Add totals row for individual reports
        if report_json['data']:
            totals_row = len(report_json['data']) + 4  # +3 for headers, +1 for spacing
            ws.cell(row=totals_row, column=1).value = "TOTALS"
            ws.cell(row=totals_row, column=1).font = Font(bold=True)
            ws.cell(row=totals_row, column=1).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # Calculate totals
            total_contacts = sum(data['contacts'] for data in report_json['data'])
            total_appointments = sum(data['appointments'] for data in report_json['data'])
            total_presentations = sum(data['presentations'] for data in report_json['data'])
            total_referrals = sum(data['referrals'] for data in report_json['data'])
            total_testimonials = sum(data['testimonials'] for data in report_json['data'])
            total_sales = sum(data['sales'] for data in report_json['data'])
            total_new_face = sum(data['new_face_sold'] for data in report_json['data'])
            total_premium = sum(data['premium'] for data in report_json['data'])
            
            ws.cell(row=totals_row, column=4).value = total_contacts
            ws.cell(row=totals_row, column=4).font = Font(bold=True)
            ws.cell(row=totals_row, column=5).value = total_appointments
            ws.cell(row=totals_row, column=5).font = Font(bold=True)
            ws.cell(row=totals_row, column=6).value = total_presentations
            ws.cell(row=totals_row, column=6).font = Font(bold=True)
            ws.cell(row=totals_row, column=7).value = total_referrals
            ws.cell(row=totals_row, column=7).font = Font(bold=True)
            ws.cell(row=totals_row, column=8).value = total_testimonials
            ws.cell(row=totals_row, column=8).font = Font(bold=True)
            ws.cell(row=totals_row, column=9).value = total_sales
            ws.cell(row=totals_row, column=9).font = Font(bold=True)
            ws.cell(row=totals_row, column=10).value = total_new_face
            ws.cell(row=totals_row, column=10).font = Font(bold=True)
            ws.cell(row=totals_row, column=11).value = total_premium
            ws.cell(row=totals_row, column=11).font = Font(bold=True)
    
    elif report_type == "team":
        # Add headers
        headers = ["Team Name", "Manager", "Role", "Contacts", "Appointments", "Presentations", 
                   "Referrals", "Testimonials", "Sales", "New Face Sold", "Total Premium"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, data in enumerate(report_json['data'], 3):
            ws.cell(row=row_num, column=1).value = data['team_name']
            ws.cell(row=row_num, column=2).value = data['manager']
            ws.cell(row=row_num, column=3).value = data['role']
            ws.cell(row=row_num, column=4).value = data['contacts']
            ws.cell(row=row_num, column=5).value = data['appointments']
            ws.cell(row=row_num, column=6).value = data['presentations']
            ws.cell(row=row_num, column=7).value = data['referrals']
            ws.cell(row=row_num, column=8).value = data['testimonials']
            ws.cell(row=row_num, column=9).value = data['sales']
            ws.cell(row=row_num, column=10).value = data['new_face_sold']
            ws.cell(row=row_num, column=11).value = data['premium']
        
        # Add totals row for team reports
        if report_json['data']:
            totals_row = len(report_json['data']) + 4  # +3 for headers, +1 for spacing
            ws.cell(row=totals_row, column=1).value = "TOTALS"
            ws.cell(row=totals_row, column=1).font = Font(bold=True)
            ws.cell(row=totals_row, column=1).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # Calculate totals
            total_contacts = sum(data['contacts'] for data in report_json['data'])
            total_appointments = sum(data['appointments'] for data in report_json['data'])
            total_presentations = sum(data['presentations'] for data in report_json['data'])
            total_referrals = sum(data['referrals'] for data in report_json['data'])
            total_testimonials = sum(data['testimonials'] for data in report_json['data'])
            total_sales = sum(data['sales'] for data in report_json['data'])
            total_new_face = sum(data['new_face_sold'] for data in report_json['data'])
            total_premium = sum(data['premium'] for data in report_json['data'])
            
            ws.cell(row=totals_row, column=4).value = total_contacts
            ws.cell(row=totals_row, column=4).font = Font(bold=True)
            ws.cell(row=totals_row, column=5).value = total_appointments
            ws.cell(row=totals_row, column=5).font = Font(bold=True)
            ws.cell(row=totals_row, column=6).value = total_presentations
            ws.cell(row=totals_row, column=6).font = Font(bold=True)
            ws.cell(row=totals_row, column=7).value = total_referrals
            ws.cell(row=totals_row, column=7).font = Font(bold=True)
            ws.cell(row=totals_row, column=8).value = total_testimonials
            ws.cell(row=totals_row, column=8).font = Font(bold=True)
            ws.cell(row=totals_row, column=9).value = total_sales
            ws.cell(row=totals_row, column=9).font = Font(bold=True)
            ws.cell(row=totals_row, column=10).value = total_new_face
            ws.cell(row=totals_row, column=10).font = Font(bold=True)
            ws.cell(row=totals_row, column=11).value = total_premium
            ws.cell(row=totals_row, column=11).font = Font(bold=True)
    
    elif report_type == "organization":
        # Add headers
        headers = ["Metric", "Total"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        metrics = [
            ("Total Members", report_json['total_members']),
            ("Contacts", report_json['data']['contacts']),
            ("Appointments", report_json['data']['appointments']),
            ("Presentations", report_json['data']['presentations']),
            ("Referrals", report_json['data']['referrals']),
            ("Testimonials", report_json['data']['testimonials']),
            ("Sales", report_json['data']['sales']),
            ("New Face Sold", report_json['data']['new_face_sold']),
            ("Total Premium", report_json['data']['premium'])
        ]
        
        for row_num, (metric, value) in enumerate(metrics, 3):
            ws.cell(row=row_num, column=1).value = metric
            ws.cell(row=row_num, column=2).value = value
    
    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = ws.cell(row=2, column=col_idx).column_letter
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"daily_{report_type}_report_{date}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/admin/diagnostic")
async def diagnostic(current_user: dict = Depends(get_current_user)):
    """Diagnostic endpoint to check database state"""
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from datetime import date
    from pytz import timezone as pytz_timezone
    
    # Get today in Central Time
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date().isoformat()
    
    # Get counts
    user_count = await db.users.count_documents({})
    activity_count = await db.activities.count_documents({})
    today_activity_count = await db.activities.count_documents({"date": today})
    
    # Get current user's activities
    my_activities = await db.activities.find({"user_id": current_user['id']}, {"_id": 0}).to_list(100)
    my_today = [a for a in my_activities if a.get('date') == today]
    
    # Get my subordinates (exclude archived)
    subordinates = await db.users.find(
        {"manager_id": current_user['id'], "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
        {"_id": 0, "name": 1, "id": 1}
    ).to_list(100)
    
    return {
        "database": os.getenv('DB_NAME', 'unknown'),
        "today": today,
        "today_central_time": f"{today} (America/Chicago)",
        "counts": {
            "total_users": user_count,
            "total_activities": activity_count,
            "activities_for_today": today_activity_count
        },
        "current_user": {
            "name": current_user['name'],
            "id": current_user['id'],
            "total_activities": len(my_activities),
            "activities_today": len(my_today),
            "today_data": {
                "contacts": my_today[0].get('contacts', 0),
                "appointments": my_today[0].get('appointments', 0),
                "premium": my_today[0].get('premium', 0)
            } if my_today else None
        },
        "subordinates_count": len(subordinates),
        "subordinates": [s['name'] for s in subordinates]
    }

@api_router.post("/admin/populate-todays-activities")
async def populate_todays_activities(current_user: dict = Depends(get_current_user)):
    """
    Admin endpoint to populate today's activities for all users.
    Run this ONCE after deployment if daily rollup shows zeros.
    Only accessible to state_manager role.
    """
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from datetime import date
    from pytz import timezone as pytz_timezone
    import uuid
    
    # Get today in Central Time
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date().isoformat()
    
    # Get all users
    all_users = await db.users.find({}, {"_id": 0}).to_list(1000)
    
    added_count = 0
    for user in all_users:
        # Check if activity exists for today
        existing = await db.activities.find_one({"user_id": user['id'], "date": today})
        
        if not existing:
            # Add today's activity with sample data
            activity = {
                "id": str(uuid.uuid4()),
                "user_id": user['id'],
                "date": today,
                "contacts": 10.0,
                "appointments": 5.0,
                "presentations": 3.0,
                "referrals": 2,
                "testimonials": 1,
                "sales": 1,
                "new_face_sold": 1,
                "premium": 1000.00,
                "created_at": datetime.now(timezone.utc),
                "edited_by": user['id'],
                "edited_at": datetime.now(timezone.utc)
            }
            
            await db.activities.insert_one(activity)
            added_count += 1
    
    return {
        "message": f"Successfully added {added_count} activities for today ({today})",
        "date": today,
        "activities_added": added_count,
        "total_users": len(all_users)
    }

@api_router.post("/auth/login")
async def login(login_data: UserLogin):
    # Try to find user by email first, then by username
    login_identifier = login_data.email.strip()  # Strip whitespace
    
    # Check if it looks like an email (contains @)
    if '@' in login_identifier:
        # Case-insensitive email lookup
        user = await db.users.find_one(
            {"email": {"$regex": f"^{login_identifier}$", "$options": "i"}},
            {"_id": 0}
        )
    else:
        # Try to find by username (case-insensitive)
        user = await db.users.find_one(
            {"username": {"$regex": f"^{login_identifier}$", "$options": "i"}},
            {"_id": 0}
        )
    
    logging.info(f"Login attempt for: {login_identifier}, user found: {user is not None}")
        
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Check if user is archived
    if user.get('status') == 'archived':
        raise HTTPException(status_code=403, detail="Account is archived. Please contact your administrator.")
    
    password_valid = verify_password(login_data.password, user['password_hash'])
    logging.info(f"Password valid: {password_valid}")
    
    if not password_valid:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_jwt_token(user['id'], user['email'])
    
    # Get team branding, features, and UI settings
    branding = None
    team_name = None
    team = None
    ui_settings = DEFAULT_TEAM_UI_SETTINGS.copy()
    
    if user.get('team_id'):
        team = await db.teams.find_one({"id": user['team_id']}, {"_id": 0})
        if team:
            team_name = team.get('name')
            branding = team.get('branding', {
                "logo_url": None,
                "primary_color": "#1e40af",
                "accent_color": "#3b82f6",
                "display_name": None,
                "tagline": None
            })
            # Get UI settings
            team_ui = team.get('ui_settings', {})
            ui_settings = {**ui_settings, **team_ui}
    
    # Get EFFECTIVE features for this user (includes role-based overrides)
    features = await get_effective_features(user, team)
    
    return {
        "token": token,
        "user": {
            "id": user['id'],
            "email": user['email'],
            "name": user['name'],
            "role": user['role'],
            "team_id": user.get('team_id'),
            "team_name": team_name,
            "manager_id": user.get('manager_id')
        },
        "branding": branding,
        "features": features,
        "ui_settings": ui_settings
    }

@api_router.get("/auth/branding")
async def get_user_branding(current_user: dict = Depends(get_current_user)):
    """Get branding for the current user's team"""
    team_id = current_user.get('team_id')
    
    if not team_id:
        return {
            "branding": None,
            "team_name": None
        }
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        return {
            "branding": None,
            "team_name": None
        }
    
    return {
        "team_name": team.get('name'),
        "branding": team.get('branding', {
            "logo_url": None,
            "primary_color": "#1e40af",
            "accent_color": "#3b82f6",
            "display_name": None,
            "tagline": None
        })
    }

class UserCreate(BaseModel):
    name: str
    email: str
    password: str
    role: str
    manager_id: Optional[str] = None

@api_router.post("/auth/create-user")
async def create_user_directly(user_data: UserCreate, current_user: dict = Depends(get_current_user)):
    """Create a new user directly with password (managers only)"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can create users")
    
    # Check if email already exists (case-insensitive)
    existing_user = await db.users.find_one(
        {"email": {"$regex": f"^{user_data.email}$", "$options": "i"}},
        {"_id": 0}
    )
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Hash the password
    password_hash = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    # Create user
    user_id = str(uuid.uuid4())
    new_user = {
        "id": user_id,
        "name": user_data.name,
        "email": user_data.email,
        "password_hash": password_hash,
        "role": user_data.role,
        "manager_id": user_data.manager_id if user_data.manager_id else None,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(new_user)
    
    return {
        "message": "User created successfully",
        "user": {
            "id": user_id,
            "name": user_data.name,
            "email": user_data.email,
            "role": user_data.role
        }
    }

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return {
        "id": current_user['id'],
        "email": current_user['email'],
        "name": current_user['name'],
        "role": current_user['role'],
        "manager_id": current_user.get('manager_id')
    }

@api_router.post("/auth/change-password")
async def change_password(password_request: PasswordChangeRequest, current_user: dict = Depends(get_current_user)):
    """
    Change user's password.
    Requires current password verification and sets new password.
    """
    try:
        # Get the user's current password hash from database
        user = await db.users.find_one({"id": current_user['id']})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Verify current password
        if not bcrypt.checkpw(password_request.current_password.encode('utf-8'), user['password_hash'].encode('utf-8')):
            raise HTTPException(status_code=400, detail="Current password is incorrect")
        
        # Validate new password (basic validation)
        if len(password_request.new_password) < 6:
            raise HTTPException(status_code=400, detail="New password must be at least 6 characters long")
        
        # Hash the new password
        salt = bcrypt.gensalt()
        new_password_hash = bcrypt.hashpw(password_request.new_password.encode('utf-8'), salt)
        
        # Update password in database
        await db.users.update_one(
            {"id": current_user['id']},
            {"$set": {"password_hash": new_password_hash.decode('utf-8')}}
        )
        
        return {"message": "Password changed successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to change password: {str(e)}")

@api_router.post("/auth/admin-reset-password")
async def admin_reset_password(reset_request: PasswordResetRequest, current_user: dict = Depends(get_current_user)):
    """
    Admin reset password for team members.
    Only State Managers can reset passwords for users in their hierarchy.
    """
    try:
        # Only State Managers can reset passwords
        if current_user['role'] not in ['state_manager', 'super_admin']:
            raise HTTPException(status_code=403, detail="Only State Managers and Super Admins can reset passwords")
        
        team_id = current_user.get('team_id')
        
        # Helper function to get all subordinates recursively (exclude archived) - SCOPED TO TEAM
        async def get_all_subordinates(user_id: str):
            members = []
            query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
            if team_id:
                query["team_id"] = team_id
            subordinates = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
            for sub in subordinates:
                members.append(sub)
                sub_members = await get_all_subordinates(sub['id'])
                members.extend(sub_members)
            return members
        
        # Get all users under current user's hierarchy
        all_subordinates = await get_all_subordinates(current_user['id'])
        all_subordinates.insert(0, current_user)  # Include self
        
        # Find the target user
        target_user = None
        
        # Super admin can reset any user's password
        if current_user['role'] == 'super_admin':
            target_user = await db.users.find_one({"id": reset_request.user_id}, {"_id": 0, "password_hash": 0})
        else:
            for member in all_subordinates:
                if member['id'] == reset_request.user_id:
                    target_user = member
                    break
        
        if not target_user:
            raise HTTPException(status_code=403, detail="User not found in your hierarchy")
        
        # Validate new password (basic validation)
        if len(reset_request.new_password) < 6:
            raise HTTPException(status_code=400, detail="New password must be at least 6 characters long")
        
        # Hash the new password
        salt = bcrypt.gensalt()
        new_password_hash = bcrypt.hashpw(reset_request.new_password.encode('utf-8'), salt)
        
        # Update password in database
        await db.users.update_one(
            {"id": reset_request.user_id},
            {"$set": {"password_hash": new_password_hash.decode('utf-8')}}
        )
        
        return {
            "message": f"Password reset successfully for {target_user.get('name', 'user')}",
            "user_name": target_user.get('name', 'Unknown'),
            "user_email": target_user.get('email', 'Unknown')
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to reset password: {str(e)}")

@api_router.post("/auth/forgot-password")
async def forgot_password(forgot_request: ForgotPasswordRequest):
    """
    Generate temporary password for users who forgot their password.
    Since we don't have email service, generates a temporary password that admins can share.
    """
    try:
        # Check if user exists (case-insensitive)
        user = await db.users.find_one({"email": {"$regex": f"^{forgot_request.email}$", "$options": "i"}})
        if not user:
            # Don't reveal if user exists or not for security
            return {"message": "If the email exists, password reset instructions would be sent"}
        
        # Generate a temporary password
        import secrets
        import string
        temp_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
        
        # Hash the temporary password
        salt = bcrypt.gensalt()
        temp_password_hash = bcrypt.hashpw(temp_password.encode('utf-8'), salt)
        
        # Update user with temporary password
        await db.users.update_one(
            {"email": forgot_request.email},
            {"$set": {"password_hash": temp_password_hash.decode('utf-8')}}
        )
        
        return {
            "message": "Temporary password generated",
            "temporary_password": temp_password,
            "user_name": user.get('name', 'Unknown'),
            "instructions": "Use this temporary password to login, then change it immediately using the Change Password feature"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate temporary password: {str(e)}")
# Activity Routes
@api_router.post("/activities")
async def create_activity(activity_data: ActivityCreate, current_user: dict = Depends(get_current_user)):
    # Check if activity exists for date (scoped to user)
    existing = await db.activities.find_one({"user_id": current_user['id'], "date": activity_data.date})
    if existing:
        raise HTTPException(status_code=400, detail="Activity already exists for this date")
    
    activity = Activity(
        user_id=current_user['id'],
        team_id=current_user.get('team_id'),  # Multi-tenancy
        date=activity_data.date,
        contacts=activity_data.contacts,
        appointments=activity_data.appointments,
        presentations=activity_data.presentations,
        referrals=activity_data.referrals,
        testimonials=activity_data.testimonials,
        sales=activity_data.sales,
        new_face_sold=activity_data.new_face_sold,
        premium=activity_data.premium
    )
    
    doc = activity.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.activities.insert_one(doc)
    
    return {"message": "Activity created", "activity": activity.model_dump()}

@api_router.put("/activities/{date}")
async def update_activity(date: str, activity_data: ActivityCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.activities.find_one({"user_id": current_user['id'], "date": date})
    
    update_data = activity_data.model_dump()
    update_data['edited_by'] = current_user['id']
    update_data['edited_at'] = datetime.now(timezone.utc).isoformat()
    update_data['team_id'] = current_user.get('team_id')  # Multi-tenancy
    
    if existing:
        await db.activities.update_one({"user_id": current_user['id'], "date": date}, {"$set": update_data})
    else:
        # Create new activity
        data_dict = activity_data.model_dump()
        activity = Activity(
            user_id=current_user['id'],
            team_id=current_user.get('team_id'),  # Multi-tenancy
            date=date,
            contacts=data_dict['contacts'],
            appointments=data_dict['appointments'],
            presentations=data_dict['presentations'],
            referrals=data_dict['referrals'],
            testimonials=data_dict['testimonials'],
            apps=data_dict.get('apps', 0),
            sales=data_dict['sales'],
            new_face_sold=data_dict['new_face_sold'],
            bankers_premium=data_dict.get('bankers_premium', 0.0),
            premium=data_dict['premium'],
            edited_by=current_user['id'],
            edited_at=datetime.now(timezone.utc)
        )
        doc = activity.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        if doc.get('edited_at'):
            doc['edited_at'] = doc['edited_at'].isoformat()
        await db.activities.insert_one(doc)
    
    return {"message": "Activity updated"}

@api_router.get("/activities/my")
async def get_my_activities(current_user: dict = Depends(get_current_user)):
    activities = await db.activities.find({"user_id": current_user['id']}, {"_id": 0}).sort("date", -1).to_list(1000)
    return activities

# Manager editing team member activities
@api_router.put("/users/{user_id}/activities/{date}")
async def update_team_activity(user_id: str, date: str, activity_data: ActivityCreate, current_user: dict = Depends(get_current_user)):
    # Get all subordinates - SCOPED TO TEAM
    team_id = current_user.get('team_id')
    subordinates = await get_all_subordinates(current_user['id'], team_id)
    
    if user_id not in subordinates or user_id == current_user['id']:
        raise HTTPException(status_code=403, detail="Not authorized to edit this user's data")
    
    existing = await db.activities.find_one({"user_id": user_id, "date": date})
    
    update_data = activity_data.model_dump()
    update_data['edited_by'] = current_user['id']
    update_data['edited_at'] = datetime.now(timezone.utc).isoformat()
    
    if existing:
        await db.activities.update_one({"user_id": user_id, "date": date}, {"$set": update_data})
    else:
        # Create new activity
        data_dict = activity_data.model_dump()
        activity = Activity(
            user_id=user_id,
            date=date,
            contacts=data_dict['contacts'],
            appointments=data_dict['appointments'],
            presentations=data_dict['presentations'],
            referrals=data_dict['referrals'],
            testimonials=data_dict['testimonials'],
            apps=data_dict.get('apps', 0),
            sales=data_dict['sales'],
            new_face_sold=data_dict['new_face_sold'],
            bankers_premium=data_dict.get('bankers_premium', 0.0),
            premium=data_dict['premium'],
            edited_by=current_user['id'],
            edited_at=datetime.now(timezone.utc)
        )
        doc = activity.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        if doc.get('edited_at'):
            doc['edited_at'] = doc['edited_at'].isoformat()
        await db.activities.insert_one(doc)
    
    return {"message": "Team member activity updated"}

@api_router.get("/users/{user_id}/activities")
async def get_team_member_activities(user_id: str, current_user: dict = Depends(get_current_user)):
    # SCOPED TO TEAM
    team_id = current_user.get('team_id')
    subordinates = await get_all_subordinates(current_user['id'], team_id)
    
    if user_id not in subordinates:
        raise HTTPException(status_code=403, detail="Not authorized to view this user's data")
    
    activities = await db.activities.find({"user_id": user_id}, {"_id": 0}).sort("date", -1).to_list(1000)
    return activities

# Statistics Routes
@api_router.get("/stats/my/{period}")
async def get_my_stats(period: str, current_user: dict = Depends(get_current_user), user_date: str = None):
    from datetime import timedelta
    
    # Use Central Time for date calculations
    if user_date:
        today = datetime.strptime(user_date, '%Y-%m-%d').date()
    else:
        central_tz = pytz_timezone('America/Chicago')
        today = datetime.now(central_tz).date()
    
    if period == "daily":
        start_date = today
    elif period == "weekly":
        start_date = today - timedelta(days=today.weekday())
    elif period == "monthly":
        start_date = today.replace(day=1)
    elif period == "quarterly":
        quarter = (today.month - 1) // 3
        start_date = today.replace(month=quarter * 3 + 1, day=1)
    elif period == "yearly":
        start_date = today.replace(month=1, day=1)
    else:
        raise HTTPException(status_code=400, detail="Invalid period")
    
    activities = await db.activities.find({
        "user_id": current_user['id'],
        "date": {"$gte": start_date.isoformat()}
    }, {"_id": 0}).to_list(1000)
    
    # Aggregate stats
    stats = {
        "contacts": sum(a['contacts'] for a in activities),
        "appointments": sum(a['appointments'] for a in activities),
        "presentations": sum(a['presentations'] for a in activities),
        "referrals": sum(a['referrals'] for a in activities),
        "testimonials": sum(a['testimonials'] for a in activities),
        "sales": sum(a['sales'] for a in activities),
        "new_face_sold": sum(a['new_face_sold'] for a in activities),
        "premium": sum(a['premium'] for a in activities)
    }
    
    return stats

# Team Routes
@api_router.get("/team/members")
async def get_team_members(current_user: dict = Depends(get_current_user)):
    team_id = current_user.get('team_id')
    query = {"manager_id": current_user['id'], "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
    if team_id:
        query["team_id"] = team_id
    members = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    return members

@api_router.get("/team/all-members")
async def get_all_team_members(current_user: dict = Depends(get_current_user)):
    """Get ALL team members in the hierarchy (for NPA/SNA tracker dropdowns)"""
    team_id = current_user.get('team_id')
    
    if current_user['role'] == 'state_manager':
        # State manager sees everyone in their team except other state managers
        query = {"role": {"$ne": "state_manager"}, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        members = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    elif current_user['role'] == 'regional_manager':
        # Regional manager sees their subordinates (scoped to team)
        subordinate_ids = await get_all_subordinates(current_user['id'], team_id)
        members = await db.users.find(
            {"id": {"$in": subordinate_ids}, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
            {"_id": 0, "password_hash": 0}
        ).to_list(1000)
    else:
        # District manager sees their direct reports
        query = {"manager_id": current_user['id'], "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        members = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    
    return members

@api_router.get("/team/week-dates")
async def get_week_dates(current_user: dict = Depends(get_current_user)):
    """
    Get the current week's date range in Central Time for frontend use.
    Returns Monday through Sunday dates for the current week.
    """
    # Use Central Time for date calculations
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    # Calculate Monday of current week
    monday = today - timedelta(days=today.weekday())
    
    # Generate all 7 days of the week
    week_dates = []
    days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    
    for i in range(7):
        current_day = monday + timedelta(days=i)
        week_dates.append({
            "date": current_day.isoformat(),
            "day_name": days_of_week[i],
            "is_today": current_day == today
        })
    
    return {
        "week_dates": week_dates,
        "week_start": monday.isoformat(),
        "today": today.isoformat()
    }

@api_router.get("/team/hierarchy/{period}")
async def get_team_hierarchy(period: str, current_user: dict = Depends(get_current_user), user_date: str = None):
    from datetime import timedelta
    
    team_id = current_user.get('team_id')
    
    # Use Central Time for date calculations
    if user_date:
        today = datetime.strptime(user_date, '%Y-%m-%d').date()
    else:
        central_tz = pytz_timezone('America/Chicago')
        today = datetime.now(central_tz).date()
    
    if period == "daily":
        start_date = today
    elif period == "weekly":
        start_date = today - timedelta(days=today.weekday())
    elif period == "monthly":
        start_date = today.replace(day=1)
    elif period == "yearly":
        start_date = today.replace(month=1, day=1)
    else:
        raise HTTPException(status_code=400, detail="Invalid period")
    
    async def build_hierarchy(user_id: str) -> dict:
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
        if not user:
            return None
        
        # Get user's own stats for the specified period
        # NOTE: We do NOT filter by team_id here because:
        # 1. The user is already confirmed to be in the hierarchy (and thus the team)
        # 2. Some legacy activities may not have team_id set
        # 3. This matches /stats/my/{period} behavior for consistency
        act_query = {"user_id": user_id, "date": {"$gte": start_date.isoformat()}}
        activities = await db.activities.find(act_query, {"_id": 0}).to_list(1000)
        
        own_stats = {
            "contacts": sum(a['contacts'] for a in activities),
            "appointments": sum(a['appointments'] for a in activities),
            "presentations": sum(a['presentations'] for a in activities),
            "referrals": sum(a['referrals'] for a in activities),
            "testimonials": sum(a['testimonials'] for a in activities),
            "sales": sum(a['sales'] for a in activities),
            "new_face_sold": sum(a['new_face_sold'] for a in activities),
            "premium": sum(a['premium'] for a in activities)
        }
        
        # Get subordinates and build their hierarchies (exclude archived users, scoped to team)
        sub_query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            sub_query["team_id"] = team_id
        subordinates = await db.users.find(sub_query, {"_id": 0}).to_list(1000)
        children = []
        
        # Initialize rolled up stats with own stats
        rolled_up_stats = {
            "contacts": own_stats["contacts"],
            "appointments": own_stats["appointments"],
            "presentations": own_stats["presentations"],
            "referrals": own_stats["referrals"],
            "testimonials": own_stats["testimonials"],
            "sales": own_stats["sales"],
            "new_face_sold": own_stats["new_face_sold"],
            "premium": own_stats["premium"]
        }
        
        # Recursively build children and roll up their stats
        for sub in subordinates:
            child_hierarchy = await build_hierarchy(sub['id'])
            if child_hierarchy:
                children.append(child_hierarchy)
                # Add child's rolled up stats to parent's rolled up stats
                rolled_up_stats["contacts"] += child_hierarchy["stats"]["contacts"]
                rolled_up_stats["appointments"] += child_hierarchy["stats"]["appointments"]
                rolled_up_stats["presentations"] += child_hierarchy["stats"]["presentations"]
                rolled_up_stats["referrals"] += child_hierarchy["stats"]["referrals"]
                rolled_up_stats["testimonials"] += child_hierarchy["stats"]["testimonials"]
                rolled_up_stats["sales"] += child_hierarchy["stats"]["sales"]
                rolled_up_stats["new_face_sold"] += child_hierarchy["stats"]["new_face_sold"]
                rolled_up_stats["premium"] += child_hierarchy["stats"]["premium"]
        
        return {
            **user,
            "stats": rolled_up_stats,  # This now includes own + all subordinates
            "children": children
        }
    
    hierarchy = await build_hierarchy(current_user['id'])
    return hierarchy

@api_router.get("/users/team-members")
async def get_all_team_members_alt(current_user: dict = Depends(get_current_user)):
    team_id = current_user.get('team_id')
    subordinate_ids = await get_all_subordinates(current_user['id'], team_id)
    if current_user['id'] in subordinate_ids:
        subordinate_ids.remove(current_user['id'])  # Remove self
    
    members = await db.users.find({"id": {"$in": subordinate_ids}}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return members

# Invite Routes
@api_router.post("/invites")
async def create_invite(invite_data: InviteCreate, current_user: dict = Depends(get_current_user)):
    # Validate role hierarchy
    role_hierarchy = {
        "state_manager": ["regional_manager"],
        "regional_manager": ["district_manager"],
        "district_manager": ["agent"],
        "agent": []
    }
    
    if invite_data.role not in role_hierarchy.get(current_user['role'], []):
        raise HTTPException(status_code=403, detail="Cannot invite this role")
    
    invite_code = generate_invite_code()
    
    invite = Invite(
        invite_code=invite_code,
        name=invite_data.name,
        email=invite_data.email,
        role=invite_data.role,
        manager_id=current_user['id'],
        team_id=current_user.get('team_id')  # Multi-tenancy
    )
    
    doc = invite.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.invites.insert_one(doc)
    
    return {"message": "Invite created", "invite_code": invite_code, "invite": invite.model_dump()}

@api_router.get("/invites/my")
async def get_my_invites(current_user: dict = Depends(get_current_user)):
    invites = await db.invites.find({"manager_id": current_user['id']}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return invites

@api_router.delete("/invites/{invite_id}")
async def delete_invite(invite_id: str, current_user: dict = Depends(get_current_user)):
    invite = await db.invites.find_one({"id": invite_id, "manager_id": current_user['id']})
    if not invite:
        raise HTTPException(status_code=404, detail="Invite not found")
    
    await db.invites.delete_one({"id": invite_id})
    return {"message": "Invite deleted"}

@api_router.get("/invites/verify/{invite_code}")
async def verify_invite(invite_code: str):
    invite = await db.invites.find_one({"invite_code": invite_code, "status": "pending"}, {"_id": 0})
    if not invite:
        raise HTTPException(status_code=404, detail="Invalid or expired invite code")
    
    return invite

# User Management
@api_router.delete("/users/{user_id}")
async def remove_user(user_id: str, current_user: dict = Depends(get_current_user)):
    # SCOPED TO TEAM
    team_id = current_user.get('team_id')
    subordinates = await get_all_subordinates(current_user['id'], team_id)
    
    if user_id not in subordinates or user_id == current_user['id']:
        raise HTTPException(status_code=403, detail="Not authorized to remove this user")
    
    await db.users.delete_one({"id": user_id})
    await db.activities.delete_many({"user_id": user_id})
    
    return {"message": "User removed"}

# Debug endpoint to check for duplicates
@api_router.get("/debug/user-activities/{user_id}")
async def debug_user_activities(user_id: str, current_user: dict = Depends(get_current_user)):
    """Debug endpoint to check a user's activities for duplicates"""
    # SCOPED TO TEAM
    team_id = current_user.get('team_id')
    subordinates = await get_all_subordinates(current_user['id'], team_id)
    
    if user_id not in subordinates and user_id != current_user['id']:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get user info
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    
    # Get all activities
    activities = await db.activities.find({"user_id": user_id}, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Check for duplicates by date
    date_counts = {}
    for activity in activities:
        date = activity['date']
        if date not in date_counts:
            date_counts[date] = []
        date_counts[date].append({
            "id": activity['id'],
            "presentations": activity.get('presentations', 0),
            "premium": activity.get('premium', 0),
            "edited_by": activity.get('edited_by'),
            "created_at": activity.get('created_at')
        })
    
    duplicates = {date: acts for date, acts in date_counts.items() if len(acts) > 1}
    
    return {
        "user": user,
        "total_activities": len(activities),
        "activities_by_date": date_counts,
        "duplicates": duplicates,
        "has_duplicates": len(duplicates) > 0
    }

# Cleanup endpoint to remove duplicates for a user
@api_router.post("/debug/cleanup-user-duplicates/{user_id}")
async def cleanup_user_duplicates(user_id: str, current_user: dict = Depends(get_current_user)):
    """Remove duplicate activities for a user, keeping the most recent one"""
    # SCOPED TO TEAM
    team_id = current_user.get('team_id')
    subordinates = await get_all_subordinates(current_user['id'], team_id)
    
    if user_id not in subordinates and user_id != current_user['id']:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get all activities for this user
    activities = await db.activities.find({"user_id": user_id}, {"_id": 0}).to_list(1000)
    
    # Group by date
    date_groups = {}
    for activity in activities:
        date = activity['date']
        if date not in date_groups:
            date_groups[date] = []
        date_groups[date].append(activity)
    
    # Find and remove duplicates
    deleted_count = 0
    cleaned_dates = []
    
    for date, acts in date_groups.items():
        if len(acts) > 1:
            # Sort by edited_at (most recent first), then created_at
            sorted_acts = sorted(acts, key=lambda x: (x.get('edited_at') or x.get('created_at'), x.get('created_at')), reverse=True)
            
            # Keep the first (most recent), delete the rest
            keep = sorted_acts[0]
            delete_ids = [a['id'] for a in sorted_acts[1:]]
            
            result = await db.activities.delete_many({"id": {"$in": delete_ids}})
            deleted_count += result.deleted_count
            cleaned_dates.append({
                "date": date,
                "kept": keep['id'],
                "deleted": delete_ids,
                "deleted_count": result.deleted_count
            })
    
    return {
        "user_id": user_id,
        "total_deleted": deleted_count,
        "dates_cleaned": cleaned_dates,
        "message": f"Removed {deleted_count} duplicate activities"
    }

# Cleanup ALL duplicates in the system (manager only)
@api_router.post("/debug/cleanup-all-duplicates")
async def cleanup_all_duplicates(current_user: dict = Depends(get_current_user)):
    """Remove all duplicate activities in the system (keeps most recent)"""
    # Get all subordinates - SCOPED TO TEAM
    team_id = current_user.get('team_id')
    subordinate_ids = await get_all_subordinates(current_user['id'], team_id)
    
    # Get all activities for subordinates
    activities = await db.activities.find({"user_id": {"$in": subordinate_ids}}, {"_id": 0}).to_list(10000)
    
    # Group by user_id and date
    groups = {}
    for activity in activities:
        key = (activity['user_id'], activity['date'])
        if key not in groups:
            groups[key] = []
        groups[key].append(activity)
    
    # Find and remove duplicates
    total_deleted = 0
    users_cleaned = {}
    
    for (user_id, date), acts in groups.items():
        if len(acts) > 1:
            # Sort by edited_at (most recent first), then created_at
            sorted_acts = sorted(acts, key=lambda x: (x.get('edited_at') or x.get('created_at'), x.get('created_at')), reverse=True)
            
            # Keep the first (most recent), delete the rest
            keep = sorted_acts[0]
            delete_ids = [a['id'] for a in sorted_acts[1:]]
            
            result = await db.activities.delete_many({"id": {"$in": delete_ids}})
            total_deleted += result.deleted_count
            
            if user_id not in users_cleaned:
                user = await db.users.find_one({"id": user_id}, {"_id": 0, "name": 1})
                users_cleaned[user_id] = {
                    "name": user['name'] if user else "Unknown",
                    "dates_cleaned": []
                }
            users_cleaned[user_id]["dates_cleaned"].append(date)
    
    return {
        "total_deleted": total_deleted,
        "users_affected": len(users_cleaned),
        "details": users_cleaned,
        "message": f"Removed {total_deleted} duplicate activities across {len(users_cleaned)} users"
    }

# Delete ALL activities for a specific user (nuclear option)
@api_router.delete("/debug/delete-all-user-activities/{user_id}")
async def delete_all_user_activities(user_id: str, current_user: dict = Depends(get_current_user)):
    """Delete ALL activities for a specific user (use with caution!)"""
    # SCOPED TO TEAM
    team_id = current_user.get('team_id')
    subordinates = await get_all_subordinates(current_user['id'], team_id)
    
    if user_id not in subordinates and user_id != current_user['id']:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get user info first
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "name": 1, "email": 1})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Count activities before deletion
    count = await db.activities.count_documents({"user_id": user_id})
    
    # Delete ALL activities
    result = await db.activities.delete_many({"user_id": user_id})
    
    return {
        "user": user,
        "activities_deleted": result.deleted_count,
        "message": f"Deleted all {result.deleted_count} activities for {user['name']}"
    }

# Leaderboard
@api_router.get("/leaderboard/{period}")
async def get_leaderboard(period: str, current_user: dict = Depends(get_current_user), user_date: str = None):
    # Check feature access
    await check_feature_access(current_user, "leaderboard")
    
    from datetime import timedelta
    
    team_id = current_user.get('team_id')
    
    # Use Central Time for date calculations
    if user_date:
        today = datetime.strptime(user_date, '%Y-%m-%d').date()
    else:
        central_tz = pytz_timezone('America/Chicago')
        today = datetime.now(central_tz).date()
    
    # Calculate start_date based on period
    if period == "daily":
        start_date = today
    elif period == "weekly":
        # Start of current week (Monday)
        start_date = today - timedelta(days=today.weekday())
    elif period == "monthly":
        # Start of current month
        start_date = today.replace(day=1)
    elif period == "quarterly":
        # Start of current quarter
        quarter = (today.month - 1) // 3
        start_date = today.replace(month=quarter * 3 + 1, day=1)
    elif period == "yearly":
        # Start of current year
        start_date = today.replace(month=1, day=1)
    else:
        raise HTTPException(status_code=400, detail="Invalid period")
    
    # End date is always today (inclusive)
    end_date = today
    
    # LEADERBOARD: Show ALL team members regardless of role or hierarchy position
    # A leaderboard is a team-wide ranking - everyone should see the same top performers
    # ALL users (including super_admin) are scoped to their assigned team
    if not team_id:
        return {"presentations": [], "referrals": [], "testimonials": [], "new_face_sold": [], "premium": []}
    
    all_users_query = {"team_id": team_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
    all_user_ids = [u["id"] for u in await db.users.find(all_users_query, {"_id": 0, "id": 1}).to_list(1000)]
    
    # Get all users info
    users = await db.users.find({"id": {"$in": all_user_ids}}, {"_id": 0, "password_hash": 0}).to_list(1000)
    user_dict = {u['id']: u for u in users}
    
    # Get all activities for the period for the team
    # Query uses >= start_date AND <= end_date
    act_query = {
        "user_id": {"$in": all_user_ids}, 
        "date": {"$gte": start_date.isoformat(), "$lte": end_date.isoformat()},
        "team_id": team_id
    }
    activities = await db.activities.find(act_query, {"_id": 0}).to_list(10000)
    
    # Aggregate by user
    user_stats = {}
    for activity in activities:
        uid = activity['user_id']
        if uid not in user_stats:
            user_stats[uid] = {
                "user_id": uid,
                "name": user_dict.get(uid, {}).get('name', 'Unknown'),
                "presentations": 0,
                "referrals": 0,
                "testimonials": 0,
                "new_face_sold": 0,
                "premium": 0.0
            }
        user_stats[uid]['presentations'] += activity.get('presentations', 0)
        user_stats[uid]['referrals'] += activity.get('referrals', 0)
        user_stats[uid]['testimonials'] += activity.get('testimonials', 0)
        user_stats[uid]['new_face_sold'] += activity.get('new_face_sold', 0)
        user_stats[uid]['premium'] += activity.get('premium', 0)
    
    # Create leaderboards for each category - Top 5 from ENTIRE organization
    leaderboard = {
        "presentations": sorted(user_stats.values(), key=lambda x: x['presentations'], reverse=True)[:5],
        "referrals": sorted(user_stats.values(), key=lambda x: x['referrals'], reverse=True)[:5],
        "testimonials": sorted(user_stats.values(), key=lambda x: x['testimonials'], reverse=True)[:5],
        "new_face_sold": sorted(user_stats.values(), key=lambda x: x['new_face_sold'], reverse=True)[:5],
        "premium": sorted(user_stats.values(), key=lambda x: x['premium'], reverse=True)[:5]
    }
    
    return leaderboard


# ============================================
# ============================================
# PMA DocuSphere - Document Library Endpoints
# ============================================

@api_router.get("/docusphere/folders")
async def get_docusphere_folders(current_user: dict = Depends(get_current_user)):
    """Get folders for user's team only"""
    # Check feature access
    await check_feature_access(current_user, "docusphere")
    
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail="You must be assigned to a team to access DocuSphere")
    
    # Only return folders for the user's team
    folders = await db.docusphere_folders.find(
        {"team_id": team_id}, 
        {"_id": 0}
    ).sort("name", 1).to_list(500)
    return folders

@api_router.post("/docusphere/folders")
async def create_docusphere_folder(
    folder_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Create a new folder (State Manager only, within their team)"""
    # Check feature access
    await check_feature_access(current_user, "docusphere")
    
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only State Managers can create folders")
    
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail="You must be assigned to a team")
    
    folder = {
        "id": str(uuid.uuid4()),
        "name": folder_data.get('name'),
        "parent_id": folder_data.get('parent_id'),
        "team_id": team_id,
        "created_by": current_user['id'],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.docusphere_folders.insert_one(folder)
    return {"message": "Folder created", "id": folder['id']}

@api_router.delete("/docusphere/folders/{folder_id}")
async def delete_docusphere_folder(folder_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a folder and all its contents (State Manager only, within their team)"""
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only State Managers can delete folders")
    
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail="You must be assigned to a team")
    
    # Verify folder belongs to user's team
    folder = await db.docusphere_folders.find_one({"id": folder_id, "team_id": team_id})
    if not folder:
        raise HTTPException(status_code=404, detail="Folder not found or access denied")
    
    # Get all subfolders recursively (within same team)
    async def get_subfolder_ids(parent_id):
        ids = [parent_id]
        subfolders = await db.docusphere_folders.find(
            {"parent_id": parent_id, "team_id": team_id}, 
            {"id": 1}
        ).to_list(100)
        for sf in subfolders:
            ids.extend(await get_subfolder_ids(sf['id']))
        return ids
    
    folder_ids = await get_subfolder_ids(folder_id)
    
    # Delete all documents in these folders (within same team)
    await db.docusphere_documents.delete_many({"folder_id": {"$in": folder_ids}, "team_id": team_id})
    
    # Delete all folders
    await db.docusphere_folders.delete_many({"id": {"$in": folder_ids}, "team_id": team_id})
    
    return {"message": "Folder and contents deleted"}

@api_router.get("/docusphere/documents")
async def get_docusphere_documents(current_user: dict = Depends(get_current_user)):
    """Get documents for user's team only (without file data)"""
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail="You must be assigned to a team to access DocuSphere")
    
    # Only return documents for the user's team
    documents = await db.docusphere_documents.find(
        {"team_id": team_id}, 
        {"_id": 0, "file_data": 0}
    ).sort("filename", 1).to_list(1000)
    return documents

@api_router.post("/docusphere/documents")
async def upload_docusphere_document(
    file: UploadFile = File(...),
    folder_id: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Upload a document (State Manager only, within their team)"""
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only State Managers can upload documents")
    
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail="You must be assigned to a team")
    
    # If folder specified, verify it belongs to user's team
    if folder_id:
        folder = await db.docusphere_folders.find_one({"id": folder_id, "team_id": team_id})
        if not folder:
            raise HTTPException(status_code=404, detail="Folder not found or access denied")
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")
    
    file_content = await file.read()
    
    if len(file_content) > 15 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size must be less than 15MB")
    
    file_base64 = base64.b64encode(file_content).decode('utf-8')
    
    document = {
        "id": str(uuid.uuid4()),
        "filename": file.filename,
        "file_data": file_base64,
        "file_size": len(file_content),
        "folder_id": folder_id,
        "team_id": team_id,
        "uploaded_by": current_user['id'],
        "uploaded_by_name": current_user['name'],
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.docusphere_documents.insert_one(document)
    
    return {"message": "Document uploaded", "id": document['id'], "filename": document['filename']}

@api_router.get("/docusphere/documents/{doc_id}/download")
async def download_docusphere_document(doc_id: str, current_user: dict = Depends(get_current_user)):
    """Download a document (user's team only)"""
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail="You must be assigned to a team")
    
    # Only allow download if document belongs to user's team
    document = await db.docusphere_documents.find_one(
        {"id": doc_id, "team_id": team_id}, 
        {"_id": 0}
    )
    if not document:
        raise HTTPException(status_code=404, detail="Document not found or access denied")
    
    file_content = base64.b64decode(document['file_data'])
    
    return Response(
        content=file_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={document['filename']}"}
    )

@api_router.delete("/docusphere/documents/{doc_id}")
async def delete_docusphere_document(doc_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a document (State Manager only, within their team)"""
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only State Managers can delete documents")
    
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail="You must be assigned to a team")
    
    # Only delete if document belongs to user's team
    result = await db.docusphere_documents.delete_one({"id": doc_id, "team_id": team_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document not found or access denied")
    
    return {"message": "Document deleted"}


# ============================================
# Recruiting Pipeline Endpoints (State Manager and Regional Manager)
# ============================================

@api_router.get("/recruiting")
async def get_recruits(current_user: dict = Depends(get_current_user)):
    """Get recruits - STRICTLY scoped to current user's team_id.
    
    NO cross-team visibility allowed under any circumstance.
    """
    # Check feature access
    await check_feature_access(current_user, "recruiting")
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can access recruiting")
    
    team_id = current_user.get('team_id')
    
    # CRITICAL: No team_id = no data. Prevents any cross-team leakage.
    if not team_id:
        return []
    
    # STRICT team filter - only exact match, no NULL/missing allowed
    strict_team_filter = {"team_id": team_id}
    
    if current_user['role'] == 'regional_manager':
        # Regional Manager sees their own recruits + their District Managers' recruits
        # ALL scoped to their team_id
        subordinates = await get_all_subordinates(current_user['id'], team_id)
        subordinate_ids = subordinates
        subordinate_ids.append(current_user['id'])
        
        query = {"$and": [
            {"$or": [{"rm_id": current_user['id']}, {"dm_id": {"$in": subordinate_ids}}]},
            strict_team_filter
        ]}
        recruits = await db.recruits.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    elif current_user['role'] == 'district_manager':
        # District Manager only sees their own recruits, scoped to team
        query = {"$and": [
            {"dm_id": current_user['id']},
            strict_team_filter
        ]}
        recruits = await db.recruits.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    else:
        # super_admin or State Manager sees ALL recruits in THEIR team only
        # STRICT: No cross-team visibility, no NULL/missing team_id included
        query = strict_team_filter
        recruits = await db.recruits.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    return recruits

@api_router.post("/recruiting")
async def create_recruit(recruit_data: dict, current_user: dict = Depends(get_current_user)):
    """Create a new recruit (State Manager, Regional Manager, or District Manager)"""
    # Check feature access
    await check_feature_access(current_user, "recruiting")
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can manage recruiting")
    
    # Handle RM assignment
    rm_id = recruit_data.get('rm_id', '')
    rm_name = recruit_data.get('rm', '')
    if current_user['role'] == 'regional_manager':
        rm_id = current_user['id']
        rm_name = current_user['name']
    
    # Handle DM assignment
    dm_id = recruit_data.get('dm_id', '')
    dm_name = recruit_data.get('dm', '')
    if current_user['role'] == 'district_manager':
        dm_id = current_user['id']
        dm_name = current_user['name']
        # Also get the DM's manager (Regional Manager) and assign them
        dm_user = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
        if dm_user and dm_user.get('manager_id'):
            rm_user = await db.users.find_one({"id": dm_user['manager_id']}, {"_id": 0})
            if rm_user:
                rm_id = rm_user['id']
                rm_name = rm_user['name']
    
    recruit = {
        "id": str(uuid.uuid4()),
        "team_id": current_user.get('team_id'),  # Multi-tenancy
        "name": recruit_data.get('name', ''),
        "phone": recruit_data.get('phone', ''),
        "email": recruit_data.get('email', ''),
        "source": recruit_data.get('source', ''),
        "state": recruit_data.get('state', ''),
        "rm": rm_name,
        "rm_id": rm_id,
        "dm": dm_name,
        "dm_id": dm_id,
        "text_email": recruit_data.get('text_email', False),
        "vertafore": recruit_data.get('vertafore', False),
        "study_materials": recruit_data.get('study_materials', False),
        "fingerprint": recruit_data.get('fingerprint', False),
        "testing_date": recruit_data.get('testing_date', ''),
        "pass_fail": recruit_data.get('pass_fail', ''),
        "npa_license": recruit_data.get('npa_license', False),
        "comments": recruit_data.get('comments', ''),
        "pipeline_status": recruit_data.get('pipeline_status', 'active'),
        "created_by": current_user['id'],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.recruits.insert_one(recruit)
    return {"message": "Recruit created", "id": recruit['id']}

@api_router.put("/recruiting/{recruit_id}")
async def update_recruit(recruit_id: str, recruit_data: dict, current_user: dict = Depends(get_current_user)):
    """Update a recruit (State Manager, Regional Manager or District Manager for their own recruits)"""
    # Check feature access
    await check_feature_access(current_user, "recruiting")
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can manage recruiting")
    
    existing = await db.recruits.find_one({"id": recruit_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Recruit not found")
    
    # Regional Managers can only edit their own recruits
    if current_user['role'] == 'regional_manager' and existing.get('rm_id') != current_user['id']:
        raise HTTPException(status_code=403, detail="You can only edit your own recruits")
    
    # District Managers can only edit their own recruits
    if current_user['role'] == 'district_manager' and existing.get('dm_id') != current_user['id']:
        raise HTTPException(status_code=403, detail="You can only edit your own recruits")
    
    update_data = {
        "name": recruit_data.get('name', existing.get('name')),
        "phone": recruit_data.get('phone', existing.get('phone')),
        "email": recruit_data.get('email', existing.get('email')),
        "source": recruit_data.get('source', existing.get('source')),
        "state": recruit_data.get('state', existing.get('state')),
        "rm": recruit_data.get('rm', existing.get('rm')),
        "rm_id": recruit_data.get('rm_id', existing.get('rm_id')),
        "dm": recruit_data.get('dm', existing.get('dm')),
        "dm_id": recruit_data.get('dm_id', existing.get('dm_id')),
        "text_email": recruit_data.get('text_email', existing.get('text_email')),
        "vertafore": recruit_data.get('vertafore', existing.get('vertafore')),
        "study_materials": recruit_data.get('study_materials', existing.get('study_materials')),
        "fingerprint": recruit_data.get('fingerprint', existing.get('fingerprint')),
        "testing_date": recruit_data.get('testing_date', existing.get('testing_date')),
        "pass_fail": recruit_data.get('pass_fail', existing.get('pass_fail')),
        "npa_license": recruit_data.get('npa_license', existing.get('npa_license')),
        "comments": recruit_data.get('comments', existing.get('comments')),
        "pipeline_status": recruit_data.get('pipeline_status', existing.get('pipeline_status', 'active')),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.recruits.update_one({"id": recruit_id}, {"$set": update_data})
    return {"message": "Recruit updated"}

@api_router.delete("/recruiting/{recruit_id}")
async def delete_recruit(recruit_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a recruit (State Manager, Regional Manager or District Manager for their own recruits)"""
    # Check feature access
    await check_feature_access(current_user, "recruiting")
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can manage recruiting")
    
    existing = await db.recruits.find_one({"id": recruit_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Recruit not found")
    
    # Regional Managers can only delete their own recruits
    if current_user['role'] == 'regional_manager' and existing.get('rm_id') != current_user['id']:
        raise HTTPException(status_code=403, detail="You can only delete your own recruits")
    
    # District Managers can only delete their own recruits
    if current_user['role'] == 'district_manager' and existing.get('dm_id') != current_user['id']:
        raise HTTPException(status_code=403, detail="You can only delete your own recruits")
    
    await db.recruits.delete_one({"id": recruit_id})
    return {"message": "Recruit deleted"}


# ============================================
# Interview Management Endpoints
# ============================================

@api_router.get("/interviews")
async def get_interviews(current_user: dict = Depends(get_current_user)):
    """Get interviews - STRICTLY scoped to current user's team_id.
    
    NO cross-team visibility allowed under any circumstance.
    """
    # Check feature access
    await check_feature_access(current_user, "interviews")
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can access interviews")
    
    team_id = current_user.get('team_id')
    
    # CRITICAL: No team_id = no data. Prevents any cross-team leakage.
    if not team_id:
        return []
    
    # Base filters
    archived_filter = {"$or": [{"archived": {"$exists": False}}, {"archived": False}]}
    # STRICT team filter - only exact match, no NULL/missing allowed
    strict_team_filter = {"team_id": team_id}
    
    # Query for interviews shared with current user
    shared_with_filter = {"shared_with": current_user['id']}
    
    if current_user['role'] in ['super_admin', 'state_manager']:
        # State manager and super_admin see ALL interviews in THEIR team only
        query = {"$and": [archived_filter, strict_team_filter]}
        interviews = await db.interviews.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    elif current_user['role'] == 'regional_manager':
        subordinate_ids = await get_all_subordinates(current_user['id'], team_id)
        all_ids = list(set(subordinate_ids))
        all_ids.append(current_user['id'])
        
        query = {"$and": [
            {"$or": [
                {"interviewer_id": {"$in": all_ids}},
                shared_with_filter
            ]},
            archived_filter,
            strict_team_filter
        ]}
        interviews = await db.interviews.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    else:
        # District managers see only their own interviews + shared with them
        query = {"$and": [
            {"$or": [
                {"interviewer_id": current_user['id']},
                shared_with_filter
            ]},
            archived_filter,
            strict_team_filter
        ]}
        interviews = await db.interviews.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Mark which interviews are shared
    for interview in interviews:
        interview['is_shared'] = interview.get('interviewer_id') != current_user['id'] and current_user['id'] in (interview.get('shared_with') or [])
    
    return interviews

@api_router.get("/interviews/stats")
async def get_interview_stats(current_user: dict = Depends(get_current_user)):
    """Get interview statistics - STRICTLY scoped to current user's team_id."""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can access interviews")
    
    from datetime import datetime, timedelta
    import pytz
    
    central = pytz.timezone('America/Chicago')
    now = datetime.now(central)
    
    team_id = current_user.get('team_id')
    if not team_id:
        return {"total": 0, "this_week": 0, "this_month": 0, "this_year": 0, "moving_forward": 0, "not_moving_forward": 0, "second_interview_scheduled": 0, "completed": 0}
    
    # Calculate date ranges
    today = now.date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    year_start = today.replace(month=1, day=1)
    
    # STRICT team filter - only exact match
    strict_team_filter = {"team_id": team_id}
    
    if current_user['role'] in ['super_admin', 'state_manager']:
        base_query = strict_team_filter
    elif current_user['role'] == 'regional_manager':
        subordinate_ids = await get_all_subordinates(current_user['id'], team_id)
        all_ids = list(set(subordinate_ids))
        all_ids.append(current_user['id'])
        base_query = {"$and": [{"interviewer_id": {"$in": all_ids}}, strict_team_filter]}
    else:
        base_query = {"$and": [{"interviewer_id": current_user['id']}, strict_team_filter]}
    
    # Get all interviews for stats
    all_interviews = await db.interviews.find(base_query, {"_id": 0}).to_list(10000)
    
    # Calculate stats
    total = len(all_interviews)
    this_week = sum(1 for i in all_interviews if i.get('interview_date', '')[:10] >= str(week_start))
    this_month = sum(1 for i in all_interviews if i.get('interview_date', '')[:10] >= str(month_start))
    this_year = sum(1 for i in all_interviews if i.get('interview_date', '')[:10] >= str(year_start))
    
    # Status breakdown
    moving_forward = sum(1 for i in all_interviews if i.get('status') == 'moving_forward')
    not_moving = sum(1 for i in all_interviews if i.get('status') == 'not_moving_forward')
    second_scheduled = sum(1 for i in all_interviews if i.get('status') == 'second_interview_scheduled')
    completed = sum(1 for i in all_interviews if i.get('status') == 'completed')
    
    return {
        "total": total,
        "this_week": this_week,
        "this_month": this_month,
        "this_year": this_year,
        "moving_forward": moving_forward,
        "not_moving_forward": not_moving,
        "second_interview_scheduled": second_scheduled,
        "completed": completed
    }

@api_router.get("/interviews/regional-breakdown")
async def get_interview_regional_breakdown(current_user: dict = Depends(get_current_user)):
    """Get interview statistics broken down by region/manager - State Manager only"""
    if current_user['role'] not in ['super_admin', 'state_manager']:
        raise HTTPException(status_code=403, detail="Only State Managers can view regional breakdown")
    
    from datetime import datetime, timedelta
    import pytz
    
    central = pytz.timezone('America/Chicago')
    now = datetime.now(central)
    
    team_id = current_user.get('team_id')
    
    # Calculate date ranges
    today = now.date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    year_start = today.replace(month=1, day=1)
    
    # Get all interviews (scoped by team for non-super_admin)
    interview_query = {}
    if current_user['role'] != 'super_admin' and team_id:
        interview_query["team_id"] = team_id
    all_interviews = await db.interviews.find(interview_query, {"_id": 0}).to_list(10000)
    
    # Get all regional managers - ONLY include real accounts with company email, scoped by team
    rm_query = {"role": "regional_manager"}
    if current_user['role'] != 'super_admin' and team_id:
        rm_query["team_id"] = team_id
    regional_managers = await db.users.find(
        rm_query,
        {"_id": 0, "id": 1, "name": 1, "email": 1, "is_active": 1}
    ).to_list(100)
    
    # Filter to ONLY include accounts with @pmagent.net email domain (real employees)
    regional_managers = [
        rm for rm in regional_managers 
        if '@pmagent.net' in (rm.get('email', '') or '').lower()
        and rm.get('is_active', True) != False
    ]
    
    # Get all district managers with their manager_id (to link to RM) - also only real accounts, scoped by team
    dm_query = {"role": "district_manager"}
    if current_user['role'] != 'super_admin' and team_id:
        dm_query["team_id"] = team_id
    district_managers = await db.users.find(
        dm_query,
        {"_id": 0, "id": 1, "name": 1, "manager_id": 1, "email": 1, "is_active": 1}
    ).to_list(100)
    
    # Filter to ONLY include accounts with @pmagent.net email domain
    district_managers = [
        dm for dm in district_managers 
        if '@pmagent.net' in (dm.get('email', '') or '').lower()
        and dm.get('is_active', True) != False
    ]
    
    # Build a map of DM -> RM
    dm_to_rm = {}
    for dm in district_managers:
        dm_to_rm[dm['id']] = dm.get('manager_id', '')
    
    # Get State Manager IDs to exclude their interviews from regional breakdown
    sm_query = {"role": "state_manager"}
    if current_user['role'] != 'super_admin' and team_id:
        sm_query["team_id"] = team_id
    state_managers = await db.users.find(
        sm_query,
        {"_id": 0, "id": 1}
    ).to_list(100)
    state_manager_ids = [sm['id'] for sm in state_managers]
    
    # Filter out State Manager interviews from all_interviews for regional breakdown
    all_interviews_excluding_sm = [i for i in all_interviews if i.get('interviewer_id') not in state_manager_ids]
    
    # Build regional breakdown
    regional_data = []
    
    for rm in regional_managers:
        rm_id = rm['id']
        rm_name = rm.get('name', 'Unknown')
        
        # Get all DMs under this RM
        rm_dm_ids = [dm['id'] for dm in district_managers if dm.get('manager_id') == rm_id]
        
        # All interviewers in this region (RM + their DMs)
        region_interviewer_ids = [rm_id] + rm_dm_ids
        
        # Filter interviews by this region (using filtered list that excludes SM interviews)
        region_interviews = [i for i in all_interviews_excluding_sm if i.get('interviewer_id') in region_interviewer_ids]
        
        # Calculate stats
        total = len(region_interviews)
        this_week = sum(1 for i in region_interviews if i.get('interview_date', '')[:10] >= str(week_start))
        this_month = sum(1 for i in region_interviews if i.get('interview_date', '')[:10] >= str(month_start))
        this_year = sum(1 for i in region_interviews if i.get('interview_date', '')[:10] >= str(year_start))
        
        # Moving forward rate
        moving_forward = sum(1 for i in region_interviews if i.get('status') in ['moving_forward', 'second_interview_scheduled', 'completed'])
        
        # Build DM breakdown within this region
        dm_breakdown = []
        for dm in district_managers:
            if dm.get('manager_id') == rm_id:
                dm_interviews = [i for i in all_interviews_excluding_sm if i.get('interviewer_id') == dm['id']]
                dm_breakdown.append({
                    "id": dm['id'],
                    "name": dm.get('name', 'Unknown'),
                    "total": len(dm_interviews),
                    "this_week": sum(1 for i in dm_interviews if i.get('interview_date', '')[:10] >= str(week_start)),
                    "this_month": sum(1 for i in dm_interviews if i.get('interview_date', '')[:10] >= str(month_start)),
                    "this_year": sum(1 for i in dm_interviews if i.get('interview_date', '')[:10] >= str(year_start))
                })
        
        # RM's own interviews (not through DMs)
        rm_own_interviews = [i for i in all_interviews_excluding_sm if i.get('interviewer_id') == rm_id]
        
        regional_data.append({
            "rm_id": rm_id,
            "rm_name": rm_name,
            "total": total,
            "this_week": this_week,
            "this_month": this_month,
            "this_year": this_year,
            "moving_forward": moving_forward,
            "moving_forward_rate": round((moving_forward / total * 100), 1) if total > 0 else 0,
            "rm_own_interviews": len(rm_own_interviews),
            "dm_count": len(rm_dm_ids),
            "dm_breakdown": sorted(dm_breakdown, key=lambda x: x['this_week'], reverse=True)
        })
    
    # Sort by this_week descending (most active regions first)
    regional_data.sort(key=lambda x: x['this_week'], reverse=True)
    
    return {
        "regional_breakdown": regional_data,
        "date_ranges": {
            "week_start": str(week_start),
            "month_start": str(month_start),
            "year_start": str(year_start)
        }
    }

@api_router.post("/interviews")
async def create_interview(interview_data: dict, current_user: dict = Depends(get_current_user)):
    """Create a new interview (1st interview submission)"""
    # Check feature access
    await check_feature_access(current_user, "interviews")
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can conduct interviews")
    
    from uuid import uuid4
    from datetime import datetime, timezone
    
    interview = {
        "id": str(uuid4()),
        "team_id": current_user.get('team_id'),  # Multi-tenancy
        "candidate_name": interview_data.get('candidate_name', ''),
        "candidate_location": interview_data.get('candidate_location', ''),
        "candidate_phone": interview_data.get('candidate_phone', ''),
        "interview_date": interview_data.get('interview_date', ''),
        "interviewer_id": current_user['id'],
        "interviewer_name": current_user['name'],
        
        # Interview form fields
        "hobbies_interests": interview_data.get('hobbies_interests', ''),
        "must_have_commission": interview_data.get('must_have_commission', False),
        "must_have_travel": interview_data.get('must_have_travel', False),
        "must_have_background": interview_data.get('must_have_background', False),
        "must_have_car": interview_data.get('must_have_car', False),
        "work_history": interview_data.get('work_history', ''),
        "what_would_change": interview_data.get('what_would_change', ''),
        "why_left_recent": interview_data.get('why_left_recent', ''),
        "other_interviews": interview_data.get('other_interviews', ''),
        "top_3_looking_for": interview_data.get('top_3_looking_for', ''),
        "why_important": interview_data.get('why_important', ''),
        "situation_6_12_months": interview_data.get('situation_6_12_months', ''),
        "family_impact": interview_data.get('family_impact', ''),
        "competitiveness_scale": interview_data.get('competitiveness_scale', 5),
        "competitiveness_example": interview_data.get('competitiveness_example', ''),
        "work_ethic_scale": interview_data.get('work_ethic_scale', 5),
        "work_ethic_example": interview_data.get('work_ethic_example', ''),
        "career_packet_sent": interview_data.get('career_packet_sent', False),
        "candidate_strength": interview_data.get('candidate_strength', 3),
        "red_flags_notes": interview_data.get('red_flags_notes', ''),
        
        # Status and tracking
        "status": interview_data.get('status', 'new'),  # new, moving_forward, not_moving_forward, second_interview_scheduled, completed
        "second_interview_date": None,
        "second_interview_notes": '',
        "added_to_recruiting": False,
        "recruit_id": None,
        
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.interviews.insert_one(interview)
    interview.pop('_id', None)
    return interview

@api_router.put("/interviews/{interview_id}")
async def update_interview(interview_id: str, interview_data: dict, current_user: dict = Depends(get_current_user)):
    """Update an interview"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can update interviews")
    
    existing = await db.interviews.find_one({"id": interview_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Interview not found")
    
    # Non-state managers can only update their own interviews (except for 2nd interview scheduling)
    if current_user['role'] != 'state_manager':
        if existing.get('interviewer_id') != current_user['id']:
            # Allow if only updating second_interview_date
            allowed_fields = ['second_interview_date', 'second_interview_notes', 'status']
            update_keys = set(interview_data.keys())
            if not update_keys.issubset(set(allowed_fields)):
                raise HTTPException(status_code=403, detail="You can only update your own interviews")
    
    from datetime import datetime, timezone
    interview_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.interviews.update_one(
        {"id": interview_id},
        {"$set": interview_data}
    )
    
    updated = await db.interviews.find_one({"id": interview_id}, {"_id": 0})
    return updated

@api_router.delete("/interviews/{interview_id}")
async def delete_interview(interview_id: str, current_user: dict = Depends(get_current_user)):
    """Archive an interview (soft delete) - State Manager only. Stats are preserved."""
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only State Managers can delete interviews")
    
    existing = await db.interviews.find_one({"id": interview_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Interview not found")
    
    # Soft delete - mark as archived instead of deleting
    from datetime import datetime, timezone
    await db.interviews.update_one(
        {"id": interview_id},
        {"$set": {
            "archived": True,
            "archived_at": datetime.now(timezone.utc).isoformat(),
            "archived_by": current_user['id']
        }}
    )
    return {"message": "Interview archived"}

@api_router.post("/interviews/{interview_id}/add-to-recruiting")
async def add_interview_to_recruiting(interview_id: str, current_user: dict = Depends(get_current_user)):
    """Add a completed interview candidate to the recruiting pipeline"""
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only State Managers can add to recruiting pipeline")
    
    interview = await db.interviews.find_one({"id": interview_id})
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    
    if interview.get('added_to_recruiting'):
        raise HTTPException(status_code=400, detail="Already added to recruiting pipeline")
    
    from uuid import uuid4
    from datetime import datetime, timezone
    
    # Create recruit from interview data
    recruit = {
        "id": str(uuid4()),
        "name": interview.get('candidate_name', ''),
        "phone": interview.get('candidate_phone', ''),
        "email": '',
        "source": 'Interview',
        "state": interview.get('candidate_location', ''),
        "rm": '',
        "rm_id": '',
        "dm": '',
        "dm_id": '',
        "text_email": False,
        "vertafore": False,
        "study_materials": False,
        "fingerprint": False,
        "testing_date": '',
        "pass_fail": '',
        "npa_license": False,
        "comments": f"From interview on {interview.get('interview_date', '')}. Interviewer: {interview.get('interviewer_name', '')}",
        "pipeline_status": "active",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user['id']
    }
    
    await db.recruits.insert_one(recruit)
    
    # Update interview to mark as added
    await db.interviews.update_one(
        {"id": interview_id},
        {"$set": {
            "added_to_recruiting": True,
            "recruit_id": recruit['id'],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    recruit.pop('_id', None)
    return {"message": "Added to recruiting pipeline", "recruit": recruit}

@api_router.post("/interviews/{interview_id}/share")
async def share_interview(interview_id: str, share_data: dict, current_user: dict = Depends(get_current_user)):
    """Share an interview with specific team members"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can share interviews")
    
    existing = await db.interviews.find_one({"id": interview_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Interview not found")
    
    # Only the interviewer or state manager can share
    if existing.get('interviewer_id') != current_user['id'] and current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only the interviewer or state manager can share this interview")
    
    shared_with = share_data.get('shared_with', [])
    
    # Get the names of shared users for display
    shared_users = await db.users.find(
        {"id": {"$in": shared_with}}, 
        {"_id": 0, "id": 1, "name": 1}
    ).to_list(100)
    shared_with_names = [u.get('name', '') for u in shared_users]
    
    await db.interviews.update_one(
        {"id": interview_id},
        {"$set": {
            "shared_with": shared_with,
            "shared_with_names": shared_with_names,
            "shared_by": current_user['id'],
            "shared_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Interview shared with {len(shared_with)} team member(s)"}


# ============================================
# SNA (Sales New Agent) Tracker Endpoints
# ============================================
# Automatically tracks new agents from their first production for 90 days toward $30,000 premium goal

SNA_GOAL = 30000
SNA_TRACKING_DAYS = 90

@api_router.get("/sna-tracker")
async def get_sna_agents(current_user: dict = Depends(get_current_user)):
    """Get all SNA (new agents) being tracked - automatically based on first production"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager']:
        raise HTTPException(status_code=403, detail="Only State and Regional Managers can access SNA tracker")
    
    # Check if SNA sub-tab is enabled for this team (Phase 2 enforcement)
    await check_subtab_access(current_user, 'sna')
    
    team_id = current_user.get('team_id')
    if not team_id:
        return {"active": [], "graduated": [], "goal": SNA_GOAL, "tracking_days": SNA_TRACKING_DAYS}
    
    # Get all agents/DMs (potential SNAs) - only real accounts with @pmagent.net, scoped to team
    # ALL users (including super_admin) are scoped to their assigned team
    if current_user['role'] in ['super_admin', 'state_manager']:
        query = {"role": {"$in": ["agent", "district_manager"]}, "team_id": team_id}
        potential_snas = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    else:
        # Regional managers see only their subordinates (scoped to team)
        subordinate_ids = await get_all_subordinates(current_user['id'], team_id)
        potential_snas = await db.users.find(
            {"role": {"$in": ["agent", "district_manager"]}, "id": {"$in": subordinate_ids}, "team_id": team_id},
            {"_id": 0, "password_hash": 0}
        ).to_list(1000)
    
    # Filter to only real accounts with @pmagent.net email
    potential_snas = [u for u in potential_snas if '@pmagent.net' in (u.get('email', '') or '').lower()]
    
    # Filter out excluded agents
    potential_snas = [u for u in potential_snas if not u.get('sna_excluded', False)]
    
    # Calculate progress for each potential SNA
    sna_data = []
    graduated_data = []
    
    for user in potential_snas:
        # Find their FIRST production (first activity with premium > 0), scoped to team
        act_query = {"user_id": user['id'], "premium": {"$gt": 0}, "team_id": team_id}
        first_production = await db.activities.find_one(
            act_query,
            {"_id": 0},
            sort=[("date", 1)]  # Oldest first
        )
        
        # Skip users who haven't produced yet
        if not first_production:
            continue
        
        sna_start = first_production.get('date', '')
        if not sna_start:
            continue
            
        # Calculate days since first production
        try:
            start_date = datetime.strptime(sna_start[:10], '%Y-%m-%d').date()
        except:
            continue
        
        today = datetime.now(timezone.utc).date()
        days_in = (today - start_date).days
        days_remaining = max(0, SNA_TRACKING_DAYS - days_in)
        
        # Get total premium since first production (scoped to team)
        act_query2 = {"user_id": user['id'], "date": {"$gte": sna_start[:10]}, "team_id": team_id}
        activities = await db.activities.find(act_query2, {"_id": 0}).to_list(10000)
        
        total_premium = sum(a.get('premium', 0) for a in activities)
        
        # Calculate pace based on 90 days
        expected_by_now = (SNA_GOAL / SNA_TRACKING_DAYS) * days_in if days_in > 0 else 0
        on_pace = total_premium >= expected_by_now
        
        # Calculate daily needed to hit goal
        if days_remaining > 0:
            daily_needed = (SNA_GOAL - total_premium) / days_remaining
            weekly_needed = daily_needed * 7
        else:
            daily_needed = 0
            weekly_needed = 0
        
        # Get manager info
        manager = await db.users.find_one({"id": user.get('manager_id', '')}, {"_id": 0, "password_hash": 0})
        
        agent_data = {
            "id": user['id'],
            "name": user.get('name', ''),
            "email": user.get('email', ''),
            "role": user.get('role', ''),
            "manager_name": manager.get('name', '') if manager else '',
            "sna_start_date": sna_start,
            "days_in": days_in,
            "days_remaining": days_remaining,
            "total_premium": total_premium,
            "goal": SNA_GOAL,
            "expected_by_now": round(expected_by_now, 2),
            "on_pace": on_pace,
            "weekly_needed": round(weekly_needed, 2),
            "daily_needed": round(daily_needed, 2),
            "progress_percent": round((total_premium / SNA_GOAL) * 100, 1),
            "completed": total_premium >= SNA_GOAL,
            "graduated_date": user.get('sna_graduated_date', None)
        }
        
        # Check if graduated (hit goal OR past 90 days)
        if total_premium >= SNA_GOAL:
            agent_data['status'] = 'completed'
            graduated_data.append(agent_data)
        elif days_in > SNA_TRACKING_DAYS:
            agent_data['status'] = 'expired'
            graduated_data.append(agent_data)
        else:
            agent_data['status'] = 'active'
            sna_data.append(agent_data)
    
    # Sort active by days remaining (most urgent first), graduated by total premium
    sna_data.sort(key=lambda x: x['days_remaining'])
    graduated_data.sort(key=lambda x: x['total_premium'], reverse=True)
    
    return {
        "active": sna_data,
        "graduated": graduated_data,
        "goal": SNA_GOAL,
        "tracking_days": SNA_TRACKING_DAYS
    }

@api_router.post("/sna-tracker/{user_id}/exclude")
async def exclude_from_sna_tracking(user_id: str, current_user: dict = Depends(get_current_user)):
    """Exclude/remove a user from SNA tracking"""
    if current_user['role'] not in ['state_manager', 'regional_manager']:
        raise HTTPException(status_code=403, detail="Only State and Regional Managers can manage SNA tracking")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "sna_excluded": True,
            "sna_excluded_by": current_user['id'],
            "sna_excluded_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Removed {user.get('name', '')} from SNA tracking"}

@api_router.post("/sna-tracker/{user_id}/include")
async def include_in_sna_tracking(user_id: str, current_user: dict = Depends(get_current_user)):
    """Re-include a previously excluded user in SNA tracking"""
    if current_user['role'] not in ['state_manager', 'regional_manager']:
        raise HTTPException(status_code=403, detail="Only State and Regional Managers can manage SNA tracking")
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"sna_excluded": False}}
    )
    
    return {"message": "User re-added to SNA tracking"}


# ============================================
# NPA (New Producing Agent) Tracker Endpoints
# ============================================
# Agents are manually added and tracked toward $1,000 premium goal

NPA_GOAL = 1000

@api_router.get("/npa-tracker")
async def get_npa_agents(current_user: dict = Depends(get_current_user)):
    """Get all manually added NPA agents and their progress"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can access NPA tracker")
    
    # Check if NPA sub-tab is enabled for this team (Phase 2 enforcement)
    await check_subtab_access(current_user, 'npa')
    
    team_id = current_user.get('team_id')
    if not team_id:
        return {"active": [], "achieved": [], "goal": NPA_GOAL}
    
    # Get NPA agents based on role, filtered by team
    # ALL users (including super_admin) are scoped to their assigned team
    if current_user['role'] in ['super_admin', 'state_manager']:
        query = {"team_id": team_id}
        npa_agents = await db.npa_agents.find(query, {"_id": 0}).to_list(1000)
    else:
        # Managers see only their own added agents (within team)
        subordinate_ids = await get_all_subordinates(current_user['id'], team_id)
        subordinate_ids.append(current_user['id'])
        query = {"added_by": {"$in": subordinate_ids}, "team_id": team_id}
        npa_agents = await db.npa_agents.find(query, {"_id": 0}).to_list(1000)
    
    npa_data = []
    achieved_data = []
    
    for agent in npa_agents:
        # If linked to a user, calculate their actual premium from activities
        user_id = agent.get('user_id', '')
        if user_id:
            # Get all activities for this user and sum premium (scoped to team)
            act_query = {"user_id": user_id, "premium": {"$gt": 0}, "team_id": team_id}
            activities = await db.activities.find(act_query, {"_id": 0, "premium": 1}).to_list(10000)
            total_premium = sum(a.get('premium', 0) for a in activities)
        else:
            # Use manually entered premium
            total_premium = agent.get('total_premium', 0)
        
        achieved_npa = total_premium >= NPA_GOAL
        
        # Check if just achieved NPA and update achievement date
        if achieved_npa and not agent.get('achievement_date'):
            await db.npa_agents.update_one(
                {"id": agent.get('id')},
                {"$set": {"achievement_date": datetime.now(timezone.utc).strftime('%Y-%m-%d')}}
            )
        
        agent_info = {
            "id": agent.get('id'),
            "user_id": user_id,
            "name": agent.get('name', ''),
            "phone": agent.get('phone', ''),
            "email": agent.get('email', ''),
            "start_date": agent.get('start_date', ''),
            "upline_dm": agent.get('upline_dm', ''),
            "upline_rm": agent.get('upline_rm', ''),
            "total_premium": total_premium,
            "goal": NPA_GOAL,
            "progress_percent": round((total_premium / NPA_GOAL) * 100, 1) if NPA_GOAL > 0 else 0,
            "achieved_npa": achieved_npa,
            "achievement_date": agent.get('achievement_date', None),
            "added_by": agent.get('added_by', ''),
            "added_by_name": agent.get('added_by_name', ''),
            "created_at": agent.get('created_at', ''),
            "notes": agent.get('notes', '')
        }
        
        if achieved_npa:
            achieved_data.append(agent_info)
        else:
            npa_data.append(agent_info)
    
    # Sort active by progress descending, achieved by achievement date descending
    npa_data.sort(key=lambda x: x['total_premium'], reverse=True)
    achieved_data.sort(key=lambda x: x.get('achievement_date') or '', reverse=True)
    
    return {
        "active": npa_data,
        "achieved": achieved_data,
        "goal": NPA_GOAL
    }

@api_router.post("/npa-tracker")
async def add_npa_agent(data: dict, current_user: dict = Depends(get_current_user)):
    """Add a new agent to NPA tracking"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can add NPA agents")
    
    npa_agent = {
        "id": str(uuid.uuid4()),
        "team_id": current_user.get('team_id'),  # Multi-tenancy
        "user_id": data.get('user_id', ''),  # Link to team member if selected from list
        "name": data.get('name', ''),
        "phone": data.get('phone', ''),
        "email": data.get('email', ''),
        "start_date": data.get('start_date', datetime.now(timezone.utc).strftime('%Y-%m-%d')),
        "upline_dm": data.get('upline_dm', ''),
        "upline_rm": data.get('upline_rm', ''),
        "total_premium": float(data.get('total_premium', 0)),
        "notes": data.get('notes', ''),
        "added_by": current_user['id'],
        "added_by_name": current_user.get('name', ''),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "achievement_date": None
    }
    
    # Check if they've achieved NPA status
    if npa_agent['total_premium'] >= NPA_GOAL:
        npa_agent['achievement_date'] = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    await db.npa_agents.insert_one(npa_agent)
    
    return {"message": f"Added {npa_agent['name']} to NPA tracking", "id": npa_agent['id']}

@api_router.put("/npa-tracker/{agent_id}")
async def update_npa_agent(agent_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Update an NPA agent's information or premium"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can update NPA agents")
    
    existing = await db.npa_agents.find_one({"id": agent_id})
    if not existing:
        raise HTTPException(status_code=404, detail="NPA agent not found")
    
    # Build update dict
    update_data = {}
    if 'name' in data:
        update_data['name'] = data['name']
    if 'phone' in data:
        update_data['phone'] = data['phone']
    if 'email' in data:
        update_data['email'] = data['email']
    if 'start_date' in data:
        update_data['start_date'] = data['start_date']
    if 'upline_dm' in data:
        update_data['upline_dm'] = data['upline_dm']
    if 'upline_rm' in data:
        update_data['upline_rm'] = data['upline_rm']
    if 'total_premium' in data:
        new_premium = float(data['total_premium'])
        update_data['total_premium'] = new_premium
        # Check if just achieved NPA status
        if new_premium >= NPA_GOAL and not existing.get('achievement_date'):
            update_data['achievement_date'] = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    if 'notes' in data:
        update_data['notes'] = data['notes']
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    update_data['updated_by'] = current_user['id']
    
    await db.npa_agents.update_one({"id": agent_id}, {"$set": update_data})
    
    return {"message": "NPA agent updated successfully"}

@api_router.delete("/npa-tracker/{agent_id}")
async def delete_npa_agent(agent_id: str, current_user: dict = Depends(get_current_user)):
    """Remove an agent from NPA tracking"""
    if current_user['role'] not in ['state_manager', 'regional_manager']:
        raise HTTPException(status_code=403, detail="Only State and Regional Managers can delete NPA agents")
    
    result = await db.npa_agents.delete_one({"id": agent_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="NPA agent not found")
    
    return {"message": "NPA agent removed from tracking"}


# ============================================
# PMA Bonus PDF Management Endpoints
# ============================================

@api_router.get("/pma-bonuses")
async def get_pma_bonuses(current_user: dict = Depends(get_current_user)):
    """Get all PMA bonus PDFs"""
    await check_feature_access(current_user, "pma_bonuses")
    bonuses = await db.pma_bonuses.find({}, {"_id": 0, "file_data": 0}).sort("uploaded_at", -1).to_list(100)
    return bonuses

@api_router.post("/pma-bonuses")
async def upload_pma_bonus(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload a PMA bonus PDF (State Manager only)"""
    await check_feature_access(current_user, "pma_bonuses")
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only State Managers can upload bonus PDFs")
    
    # Validate file type
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")
    
    # Read file content
    file_content = await file.read()
    
    # Check file size (max 10MB)
    if len(file_content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size must be less than 10MB")
    
    # Store as base64
    file_base64 = base64.b64encode(file_content).decode('utf-8')
    
    bonus_doc = {
        "id": str(uuid.uuid4()),
        "filename": file.filename,
        "file_data": file_base64,
        "file_size": len(file_content),
        "uploaded_by": current_user['id'],
        "uploaded_by_name": current_user['name'],
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.pma_bonuses.insert_one(bonus_doc)
    
    return {
        "message": "PDF uploaded successfully",
        "id": bonus_doc['id'],
        "filename": bonus_doc['filename']
    }

@api_router.get("/pma-bonuses/{bonus_id}/download")
async def download_pma_bonus(bonus_id: str, current_user: dict = Depends(get_current_user)):
    """Download a PMA bonus PDF"""
    await check_feature_access(current_user, "pma_bonuses")
    bonus = await db.pma_bonuses.find_one({"id": bonus_id}, {"_id": 0})
    if not bonus:
        raise HTTPException(status_code=404, detail="Bonus PDF not found")
    
    # Decode base64 to bytes
    file_content = base64.b64decode(bonus['file_data'])
    
    return Response(
        content=file_content,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={bonus['filename']}"
        }
    )

@api_router.delete("/pma-bonuses/{bonus_id}")
async def delete_pma_bonus(bonus_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a PMA bonus PDF (State Manager only)"""
    await check_feature_access(current_user, "pma_bonuses")
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only State Managers can delete bonus PDFs")
    
    result = await db.pma_bonuses.delete_one({"id": bonus_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bonus PDF not found")
    
    return {"message": "PDF deleted successfully"}


# ============================================
# Team Reorganization & Archiving Endpoints
# ============================================

class UserReassignment(BaseModel):
    role: Optional[str] = None
    manager_id: Optional[str] = None

@api_router.put("/users/{user_id}/reassign")
async def reassign_user(user_id: str, reassignment: UserReassignment, current_user: dict = Depends(get_current_user)):
    """Reassign a user's role and/or manager (for promotions/transfers)"""
    # Only managers can reassign
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can reassign team members")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Build update dict
    update_data = {}
    if reassignment.role is not None:
        # Validate role
        valid_roles = ['agent', 'district_manager', 'regional_manager', 'state_manager']
        if reassignment.role not in valid_roles:
            raise HTTPException(status_code=400, detail="Invalid role")
        update_data['role'] = reassignment.role
    
    if reassignment.manager_id is not None:
        # Validate manager exists
        if reassignment.manager_id != "":
            manager = await db.users.find_one({"id": reassignment.manager_id}, {"_id": 0})
            if not manager:
                raise HTTPException(status_code=404, detail="Manager not found")
        update_data['manager_id'] = reassignment.manager_id if reassignment.manager_id != "" else None
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No changes provided")
    
    # Update user
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    return {"message": "User reassigned successfully", "updates": update_data}

@api_router.put("/users/{user_id}/archive")
async def archive_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Archive a user (when they quit) - preserves all data but removes from active hierarchy"""
    # Only state managers can archive
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only state managers can archive users")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.get('status') == 'archived':
        raise HTTPException(status_code=400, detail="User is already archived")
    
    # Check if user has subordinates (exclude archived)
    subordinates = await db.users.find(
        {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
        {"_id": 0}
    ).to_list(1000)
    subordinate_count = len(subordinates)
    
    # Archive the user (keep all historical data)
    await db.users.update_one({"id": user_id}, {"$set": {"status": "archived", "archived_at": datetime.now(timezone.utc).isoformat()}})
    
    return {
        "message": "User archived successfully",
        "subordinates_count": subordinate_count,
        "warning": f"This user has {subordinate_count} subordinate(s) who need to be reassigned" if subordinate_count > 0 else None
    }

@api_router.put("/users/{user_id}/unarchive")
async def unarchive_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Unarchive a user (restore to active)"""
    if current_user['role'] != 'state_manager':
        raise HTTPException(status_code=403, detail="Only state managers can unarchive users")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.get('status') != 'archived':
        raise HTTPException(status_code=400, detail="User is not archived")
    
    await db.users.update_one({"id": user_id}, {"$set": {"status": "active"}, "$unset": {"archived_at": ""}})
    
    return {"message": "User unarchived successfully"}

@api_router.get("/users/archived/list")
async def get_archived_users(current_user: dict = Depends(get_current_user)):
    """Get list of archived users within the user's team and hierarchy"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can view archived users")
    
    team_id = current_user.get('team_id')
    if not team_id:
        return []
    
    # ALL users (including super_admin) are scoped to their assigned team on product pages
    archived = await db.users.find(
        {"status": "archived", "team_id": team_id}, 
        {"_id": 0, "password_hash": 0}
    ).sort("archived_at", -1).to_list(1000)
    
    # Get activity stats for each archived user
    for user in archived:
        activities = await db.activities.find({"user_id": user['id'], "team_id": team_id}, {"_id": 0}).to_list(10000)
        total_stats = {
            "presentations": sum(a.get('presentations', 0) for a in activities),
            "appointments": sum(a.get('appointments', 0) for a in activities),
            "sales": sum(a.get('sales', 0) for a in activities),
            "premium": sum(a.get('premium', 0) for a in activities)
        }
        user['total_stats'] = total_stats
    
    return archived

@api_router.get("/users/active/list")
async def get_active_users_for_reassignment(current_user: dict = Depends(get_current_user)):
    """Get active users within the user's team and hierarchy for team reorganization.
    
    ALL users (including super_admin) are scoped to their assigned team on product pages.
    Cross-team visibility is ONLY available in Admin endpoints where team_id is explicitly passed.
    
    - super_admin/state_manager: Can see ALL users in their team (they're at the top of hierarchy)
    - regional_manager/district_manager: Can see only their downline
    """
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can access this")
    
    team_id = current_user.get('team_id')
    if not team_id:
        return []
    
    # super_admin uses same team scoping as state_manager on product pages
    if current_user['role'] in ['super_admin', 'state_manager']:
        # State manager and super_admin see ALL users in their team (they're at the top)
        users = await db.users.find(
            {
                "team_id": team_id,
                "$or": [{"status": "active"}, {"status": {"$exists": False}}]
            },
            {"_id": 0, "password_hash": 0}
        ).to_list(10000)
    else:
        # For regional/district managers: filter by team_id AND hierarchy
        # Get only users within their downline hierarchy (same team)
        subordinate_ids = await get_all_subordinates(current_user['id'], team_id)
        
        # Also include self
        all_ids = list(set(subordinate_ids))
        if current_user['id'] not in all_ids:
            all_ids.append(current_user['id'])
        
        users = await db.users.find(
            {
                "id": {"$in": all_ids},
                "team_id": team_id,
                "$or": [{"status": "active"}, {"status": {"$exists": False}}]
            },
            {"_id": 0, "password_hash": 0}
        ).to_list(10000)
    
    # Add manager name to each user
    for user in users:
        if user.get('manager_id'):
            manager = await db.users.find_one({"id": user['manager_id']}, {"_id": 0, "name": 1})
            user['manager_name'] = manager['name'] if manager else "Unknown"
        else:
            user['manager_name'] = "None"
    
    return users



# ============================================
# Weekly Averages & Analytics Endpoints
# ============================================

@api_router.get("/analytics/personal-averages")
async def get_personal_averages(current_user: dict = Depends(get_current_user)):
    """Get personal weekly averages for different time periods"""
    # Check feature access
    await check_feature_access(current_user, "analytics")
    
    from datetime import timedelta
    
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    # Calculate date ranges
    periods = {
        "last_4_weeks": today - timedelta(weeks=4),
        "last_8_weeks": today - timedelta(weeks=8),
        "last_12_weeks": today - timedelta(weeks=12),
        "ytd": today.replace(month=1, day=1)
    }
    
    result = {}
    
    for period_name, start_date in periods.items():
        # Get activities for this period
        activities = await db.activities.find({
            "user_id": current_user['id'],
            "date": {"$gte": start_date.isoformat()}
        }, {"_id": 0}).to_list(10000)
        
        # Calculate totals
        totals = {
            "presentations": sum(a.get('presentations', 0) for a in activities),
            "appointments": sum(a.get('appointments', 0) for a in activities),
            "sales": sum(a.get('sales', 0) for a in activities),
            "premium": sum(a.get('premium', 0) for a in activities)
        }
        
        # Calculate weeks in period
        days_in_period = (today - start_date).days
        weeks_in_period = max(days_in_period / 7, 1)
        
        # Calculate averages
        averages = {
            "presentations": round(totals["presentations"] / weeks_in_period, 1),
            "appointments": round(totals["appointments"] / weeks_in_period, 1),
            "sales": round(totals["sales"] / weeks_in_period, 1),
            "premium": round(totals["premium"] / weeks_in_period, 2)
        }
        
        result[period_name] = {
            "averages": averages,
            "totals": totals,
            "weeks": round(weeks_in_period, 1)
        }
    
    return result

@api_router.get("/analytics/team-averages")
async def get_team_averages(current_user: dict = Depends(get_current_user)):
    """Get team averages for managers"""
    # Check feature access
    await check_feature_access(current_user, "analytics")
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can access team analytics")
    
    team_id = current_user.get('team_id')
    
    from datetime import timedelta
    
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    # Get all subordinates - SCOPED TO TEAM
    async def get_all_subordinates_scoped(user_id: str):
        ids = []
        query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        subordinates = await db.users.find(query, {"_id": 0, "id": 1}).to_list(1000)
        for sub in subordinates:
            ids.append(sub['id'])
            ids.extend(await get_all_subordinates_scoped(sub['id']))
        return ids
    
    team_ids = await get_all_subordinates_scoped(current_user['id'])
    team_ids.append(current_user['id'])  # Include manager's own numbers
    
    # Calculate for each period
    periods = {
        "last_4_weeks": today - timedelta(weeks=4),
        "last_8_weeks": today - timedelta(weeks=8),
        "last_12_weeks": today - timedelta(weeks=12),
        "ytd": today.replace(month=1, day=1)
    }
    
    result = {}
    
    for period_name, start_date in periods.items():
        # Get activities for entire team - SCOPED TO TEAM
        act_query = {"user_id": {"$in": team_ids}, "date": {"$gte": start_date.isoformat()}}
        if team_id:
            act_query["team_id"] = team_id
        activities = await db.activities.find(act_query, {"_id": 0}).to_list(100000)
        
        # Calculate team totals
        team_totals = {
            "presentations": sum(a.get('presentations', 0) for a in activities),
            "appointments": sum(a.get('appointments', 0) for a in activities),
            "sales": sum(a.get('sales', 0) for a in activities),
            "premium": sum(a.get('premium', 0) for a in activities)
        }
        
        days_in_period = (today - start_date).days
        weeks_in_period = max(days_in_period / 7, 1)
        
        # Calculate team averages (per team member per week)
        team_member_count = len(team_ids) if team_ids else 1
        
        team_avg_per_member = {
            "presentations": round(team_totals["presentations"] / weeks_in_period / team_member_count, 1),
            "appointments": round(team_totals["appointments"] / weeks_in_period / team_member_count, 1),
            "sales": round(team_totals["sales"] / weeks_in_period / team_member_count, 1),
            "premium": round(team_totals["premium"] / weeks_in_period / team_member_count, 2)
        }
        
        result[period_name] = {
            "team_averages_per_member": team_avg_per_member,
            "team_totals": team_totals,
            "team_size": team_member_count,
            "weeks": round(weeks_in_period, 1)
        }
    
    return result

@api_router.get("/analytics/individual-member-averages")
async def get_individual_member_averages(current_user: dict = Depends(get_current_user), period: str = "last_4_weeks"):
    """Get individual averages for each team member"""
    # Check feature access
    await check_feature_access(current_user, "analytics")
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can access this")
    
    team_id = current_user.get('team_id')
    
    from datetime import timedelta
    
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    # Calculate date range
    period_map = {
        "last_4_weeks": today - timedelta(weeks=4),
        "last_8_weeks": today - timedelta(weeks=8),
        "last_12_weeks": today - timedelta(weeks=12),
        "ytd": today.replace(month=1, day=1)
    }
    
    start_date = period_map.get(period, today - timedelta(weeks=4))
    
    # Get all subordinates - SCOPED TO TEAM
    async def get_all_subordinates_with_info(user_id: str):
        result = []
        query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        subordinates = await db.users.find(query, {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1}).to_list(1000)
        for sub in subordinates:
            result.append(sub)
            result.extend(await get_all_subordinates_with_info(sub['id']))
        return result
    
    team_members = await get_all_subordinates_with_info(current_user['id'])
    
    # Calculate averages for each member
    days_in_period = (today - start_date).days
    weeks_in_period = max(days_in_period / 7, 1)
    
    member_averages = []
    
    for member in team_members:
        # SCOPE activities by team_id too
        act_query = {"user_id": member['id'], "date": {"$gte": start_date.isoformat()}}
        if team_id:
            act_query["team_id"] = team_id
        activities = await db.activities.find(act_query, {"_id": 0}).to_list(10000)
        
        totals = {
            "presentations": sum(a.get('presentations', 0) for a in activities),
            "appointments": sum(a.get('appointments', 0) for a in activities),
            "sales": sum(a.get('sales', 0) for a in activities),
            "premium": sum(a.get('premium', 0) for a in activities)
        }
        
        averages = {
            "presentations": round(totals["presentations"] / weeks_in_period, 1),
            "appointments": round(totals["appointments"] / weeks_in_period, 1),
            "sales": round(totals["sales"] / weeks_in_period, 1),
            "premium": round(totals["premium"] / weeks_in_period, 2)
        }
        
        member_averages.append({
            "id": member['id'],
            "name": member['name'],
            "email": member['email'],
            "role": member['role'],
            "averages": averages,
            "totals": totals
        })
    
    return {
        "period": period,
        "weeks": round(weeks_in_period, 1),
        "members": member_averages
    }

@api_router.get("/analytics/manager-team-averages")
async def get_manager_team_averages(current_user: dict = Depends(get_current_user), period: str = "last_4_weeks"):
    """Get average performance for each direct report manager's team"""
    # Check feature access
    await check_feature_access(current_user, "analytics")
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can access this")
    
    team_id = current_user.get('team_id')
    
    from datetime import timedelta
    
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    # Calculate date range
    period_map = {
        "last_4_weeks": today - timedelta(weeks=4),
        "last_8_weeks": today - timedelta(weeks=8),
        "last_12_weeks": today - timedelta(weeks=12),
        "ytd": today.replace(month=1, day=1)
    }
    
    start_date = period_map.get(period, today - timedelta(weeks=4))
    days_in_period = (today - start_date).days
    weeks_in_period = max(days_in_period / 7, 1)
    
    # Get direct report managers (only managers, not agents) - SCOPED TO TEAM
    dm_query = {
        "manager_id": current_user['id'],
        "role": {"$in": ["state_manager", "regional_manager", "district_manager"]},
        "$or": [{"status": "active"}, {"status": {"$exists": False}}]
    }
    if team_id:
        dm_query["team_id"] = team_id
    direct_managers = await db.users.find(dm_query, {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1}).to_list(1000)
    
    # For each manager, calculate their team's averages - SCOPED TO TEAM
    async def get_team_ids(manager_id: str):
        ids = []
        query = {"manager_id": manager_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        subordinates = await db.users.find(query, {"_id": 0, "id": 1}).to_list(1000)
        for sub in subordinates:
            ids.append(sub['id'])
            ids.extend(await get_team_ids(sub['id']))
        return ids
    
    manager_results = []
    
    for manager in direct_managers:
        # Get all team member IDs under this manager (including the manager themselves)
        team_member_ids = await get_team_ids(manager['id'])
        team_member_ids.append(manager['id'])  # Include manager's own numbers
        team_size = len(team_member_ids)
        
        # Get activities for this manager's entire team - SCOPED TO TEAM
        act_query = {"user_id": {"$in": team_member_ids}, "date": {"$gte": start_date.isoformat()}}
        if team_id:
            act_query["team_id"] = team_id
        activities = await db.activities.find(act_query, {"_id": 0}).to_list(100000)
        
        # Calculate totals
        totals = {
            "presentations": sum(a.get('presentations', 0) for a in activities),
            "appointments": sum(a.get('appointments', 0) for a in activities),
            "sales": sum(a.get('sales', 0) for a in activities),
            "premium": sum(a.get('premium', 0) for a in activities)
        }
        
        # Calculate averages per member per week
        if team_size > 0:
            averages_per_member = {
                "presentations": round(totals["presentations"] / weeks_in_period / team_size, 1),
                "appointments": round(totals["appointments"] / weeks_in_period / team_size, 1),
                "sales": round(totals["sales"] / weeks_in_period / team_size, 1),
                "premium": round(totals["premium"] / weeks_in_period / team_size, 2)
            }
        else:
            averages_per_member = {
                "presentations": 0,
                "appointments": 0,
                "sales": 0,
                "premium": 0
            }
        
        # Check if this manager has subordinate managers
        has_subordinates = await db.users.count_documents({
            "manager_id": manager['id'],
            "role": {"$in": ["state_manager", "regional_manager", "district_manager"]},
            "$or": [{"status": "active"}, {"status": {"$exists": False}}]
        })
        
        manager_results.append({
            "id": manager['id'],
            "name": manager['name'],
            "email": manager['email'],
            "role": manager['role'],
            "team_size": team_size,
            "averages": averages_per_member,
            "totals": totals,
            "has_subordinate_managers": has_subordinates > 0
        })
    
    return {
        "period": period,
        "weeks": round(weeks_in_period, 1),
        "managers": manager_results
    }


@api_router.get("/analytics/manager-subordinates")
async def get_manager_subordinate_averages(manager_id: str, period: str, current_user: dict = Depends(get_current_user)):
    """Get subordinate managers for a specific manager (for hierarchy drill-down)"""
    # Check feature access
    await check_feature_access(current_user, "analytics")
    
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can access this")
    
    team_id = current_user.get('team_id')
    
    from datetime import timedelta
    
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    # Calculate date range
    period_map = {
        "last_4_weeks": today - timedelta(weeks=4),
        "last_8_weeks": today - timedelta(weeks=8),
        "last_12_weeks": today - timedelta(weeks=12),
        "ytd": today.replace(month=1, day=1)
    }
    
    start_date = period_map.get(period, today - timedelta(weeks=4))
    days_in_period = (today - start_date).days
    weeks_in_period = max(days_in_period / 7, 1)
    
    # Get direct subordinate managers only - SCOPED TO TEAM
    sub_query = {
        "manager_id": manager_id,
        "role": {"$in": ["state_manager", "regional_manager", "district_manager"]},
        "$or": [{"status": "active"}, {"status": {"$exists": False}}]
    }
    if team_id:
        sub_query["team_id"] = team_id
    subordinate_managers = await db.users.find(sub_query, {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1}).to_list(1000)
    
    # For each subordinate manager, calculate their team's averages - SCOPED TO TEAM
    async def get_team_member_ids(mgr_id: str):
        ids = []
        query = {"manager_id": mgr_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        subordinates = await db.users.find(query, {"_id": 0, "id": 1}).to_list(1000)
        for sub in subordinates:
            ids.append(sub['id'])
            ids.extend(await get_team_member_ids(sub['id']))
        return ids
    
    manager_results = []
    
    for manager in subordinate_managers:
        # Get all team member IDs under this manager (including the manager themselves)
        team_member_ids = await get_team_member_ids(manager['id'])
        team_member_ids.append(manager['id'])  # Include manager's own numbers
        team_size = len(team_member_ids)
        
        # Get activities for this manager's entire team - SCOPED TO TEAM
        act_query = {"user_id": {"$in": team_member_ids}, "date": {"$gte": start_date.isoformat()}}
        if team_id:
            act_query["team_id"] = team_id
        activities = await db.activities.find(act_query, {"_id": 0}).to_list(100000)
        
        # Calculate totals
        totals = {
            "presentations": sum(a.get('presentations', 0) for a in activities),
            "appointments": sum(a.get('appointments', 0) for a in activities),
            "sales": sum(a.get('sales', 0) for a in activities),
            "premium": sum(a.get('premium', 0) for a in activities)
        }
        
        # Calculate averages per member per week
        if team_size > 0:
            averages_per_member = {
                "presentations": round(totals["presentations"] / weeks_in_period / team_size, 1),
                "appointments": round(totals["appointments"] / weeks_in_period / team_size, 1),
                "sales": round(totals["sales"] / weeks_in_period / team_size, 1),
                "premium": round(totals["premium"] / weeks_in_period / team_size, 2)
            }
        else:
            averages_per_member = {
                "presentations": 0,
                "appointments": 0,
                "sales": 0,
                "premium": 0
            }
        
        # Check if this manager has subordinate managers
        has_subordinates = await db.users.count_documents({
            "manager_id": manager['id'],
            "role": {"$in": ["state_manager", "regional_manager", "district_manager"]},
            "$or": [{"status": "active"}, {"status": {"$exists": False}}]
        })
        
        manager_results.append({
            "id": manager['id'],
            "name": manager['name'],
            "email": manager['email'],
            "role": manager['role'],
            "team_size": team_size,
            "averages": averages_per_member,
            "totals": totals,
            "has_subordinate_managers": has_subordinates > 0
        })
    
    return {
        "period": period,
        "weeks": round(weeks_in_period, 1),
        "managers": manager_results
    }


# ============================================
# True Field Averages (State Manager Only)
# ============================================

@api_router.get("/analytics/true-field-averages")
async def get_true_field_averages(period: str = "last_4_weeks", current_user: dict = Depends(get_current_user)):
    """Get true field averages - only counting people who logged activity in the period (State Manager only)"""
    # Check feature access
    await check_feature_access(current_user, "analytics")
    
    if current_user['role'] not in ['super_admin', 'state_manager']:
        raise HTTPException(status_code=403, detail="Only State Managers can access true field averages")
    
    team_id = current_user.get('team_id')
    
    from datetime import timedelta
    
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    
    # Define period ranges
    periods = {
        "last_week": today - timedelta(weeks=1),
        "last_2_weeks": today - timedelta(weeks=2),
        "last_4_weeks": today - timedelta(weeks=4),
        "last_8_weeks": today - timedelta(weeks=8),
        "last_12_weeks": today - timedelta(weeks=12),
        "ytd": today.replace(month=1, day=1)
    }
    
    if period not in periods:
        period = "last_4_weeks"
    
    start_date = periods[period]
    
    # Get activities in the period - SCOPED TO TEAM
    query = {"date": {"$gte": start_date.isoformat()}}
    if team_id:
        query["team_id"] = team_id
    
    activities = await db.activities.find(query, {"_id": 0}).to_list(100000)
    
    # Get unique users who logged activity in this period
    active_user_ids = set()
    user_totals = {}
    
    for activity in activities:
        user_id = activity.get('user_id')
        if not user_id:
            continue
            
        active_user_ids.add(user_id)
        
        if user_id not in user_totals:
            user_totals[user_id] = {
                "appointments": 0,
                "presentations": 0,
                "sales": 0,
                "premium": 0,
                "days_active": set()
            }
        
        user_totals[user_id]["appointments"] += activity.get('appointments', 0)
        user_totals[user_id]["presentations"] += activity.get('presentations', 0) or activity.get('sits', 0)
        user_totals[user_id]["sales"] += activity.get('sales', 0)
        user_totals[user_id]["premium"] += activity.get('premium', 0)
        user_totals[user_id]["days_active"].add(activity.get('date'))
    
    # Calculate weeks in period
    days_in_period = (today - start_date).days
    weeks_in_period = max(days_in_period / 7, 1)
    
    # Calculate true averages (only from people who were active)
    active_count = len(active_user_ids)
    
    if active_count > 0:
        total_appointments = sum(u["appointments"] for u in user_totals.values())
        total_presentations = sum(u["presentations"] for u in user_totals.values())
        total_sales = sum(u["sales"] for u in user_totals.values())
        total_premium = sum(u["premium"] for u in user_totals.values())
        
        # Average per active person per week
        true_averages = {
            "appointments": round(total_appointments / weeks_in_period / active_count, 2),
            "presentations": round(total_presentations / weeks_in_period / active_count, 2),
            "sales": round(total_sales / weeks_in_period / active_count, 2),
            "premium": round(total_premium / weeks_in_period / active_count, 2)
        }
        
        # Totals for the period
        totals = {
            "appointments": total_appointments,
            "presentations": total_presentations,
            "sales": total_sales,
            "premium": round(total_premium, 2)
        }
    else:
        true_averages = {
            "appointments": 0,
            "presentations": 0,
            "sales": 0,
            "premium": 0
        }
        totals = {
            "appointments": 0,
            "presentations": 0,
            "sales": 0,
            "premium": 0
        }
    
    # Get user details for breakdown
    active_users_details = []
    if active_user_ids:
        users = await db.users.find(
            {"id": {"$in": list(active_user_ids)}},
            {"_id": 0, "id": 1, "name": 1, "role": 1}
        ).to_list(1000)
        
        user_map = {u['id']: u for u in users}
        
        for user_id, data in user_totals.items():
            user_info = user_map.get(user_id, {"name": "Unknown", "role": "unknown"})
            active_users_details.append({
                "id": user_id,
                "name": user_info.get('name', 'Unknown'),
                "role": user_info.get('role', 'unknown'),
                "appointments": data["appointments"],
                "presentations": data["presentations"],
                "sales": data["sales"],
                "premium": round(data["premium"], 2),
                "days_active": len(data["days_active"]),
                "avg_appointments_per_week": round(data["appointments"] / weeks_in_period, 2),
                "avg_presentations_per_week": round(data["presentations"] / weeks_in_period, 2),
                "avg_sales_per_week": round(data["sales"] / weeks_in_period, 2),
                "avg_premium_per_week": round(data["premium"] / weeks_in_period, 2)
            })
        
        # Sort by premium descending
        active_users_details.sort(key=lambda x: x["premium"], reverse=True)
    
    return {
        "period": period,
        "start_date": start_date.isoformat(),
        "end_date": today.isoformat(),
        "weeks": round(weeks_in_period, 1),
        "active_field_count": active_count,
        "true_averages": true_averages,
        "totals": totals,
        "active_users": active_users_details
    }


# ============================================
# Goal Progress & Pace Calculator Endpoints
# ============================================

class GoalSettings(BaseModel):
    goal_premium: float
    stretch_goal_premium: float

class TeamGoalSettings(BaseModel):
    goal_premium: float
    stretch_goal_premium: float

@api_router.post("/goals/individual")
async def set_individual_goals(goals: GoalSettings, current_user: dict = Depends(get_current_user)):
    """Set individual premium goals for the year"""
    goal_data = {
        "user_id": current_user['id'],
        "team_id": current_user.get('team_id'),  # Multi-tenancy
        "year": datetime.now(pytz_timezone('America/Chicago')).year,
        "goal_premium": goals.goal_premium,
        "stretch_goal_premium": goals.stretch_goal_premium,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Upsert goal (update if exists, insert if not)
    await db.goals.update_one(
        {"user_id": current_user['id'], "year": goal_data["year"]},
        {"$set": goal_data},
        upsert=True
    )
    
    return {"message": "Goals updated successfully", "goals": goal_data}

@api_router.get("/goals/individual/progress")
async def get_individual_goal_progress(current_user: dict = Depends(get_current_user)):
    """Get individual goal progress and pace calculations"""
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    current_year = today.year
    team_id = current_user.get('team_id')
    
    # Get goals for current year
    goal = await db.goals.find_one(
        {"user_id": current_user['id'], "year": current_year},
        {"_id": 0}
    )
    
    if not goal:
        return {
            "has_goals": False,
            "message": "No goals set for this year"
        }
    
    # Get YTD premium (scoped to team)
    year_start = today.replace(month=1, day=1)
    act_query = {"user_id": current_user['id'], "date": {"$gte": year_start.isoformat()}}
    if team_id:
        act_query["team_id"] = team_id
    activities = await db.activities.find(act_query, {"_id": 0}).to_list(10000)
    
    ytd_premium = sum(a.get('premium', 0) for a in activities)
    
    # Calculate weeks elapsed and remaining
    days_elapsed = (today - year_start).days
    weeks_elapsed = days_elapsed / 7
    total_weeks_in_year = 52
    weeks_remaining = total_weeks_in_year - weeks_elapsed
    
    # Goal calculations
    goal_premium = goal['goal_premium']
    stretch_goal_premium = goal['stretch_goal_premium']
    
    # Expected at this point (pro-rated based on weeks elapsed)
    expected_goal = (goal_premium / total_weeks_in_year) * weeks_elapsed
    expected_stretch = (stretch_goal_premium / total_weeks_in_year) * weeks_elapsed
    
    # Calculate progress
    goal_progress = {
        "current": ytd_premium,
        "goal": goal_premium,
        "expected": expected_goal,
        "percentage": round((ytd_premium / goal_premium * 100), 1) if goal_premium > 0 else 0,
        "ahead_behind": ytd_premium - expected_goal,
        "on_pace": ytd_premium >= expected_goal,
        "weekly_needed": round((goal_premium - ytd_premium) / weeks_remaining, 2) if weeks_remaining > 0 else 0
    }
    
    stretch_progress = {
        "current": ytd_premium,
        "goal": stretch_goal_premium,
        "expected": expected_stretch,
        "percentage": round((ytd_premium / stretch_goal_premium * 100), 1) if stretch_goal_premium > 0 else 0,
        "ahead_behind": ytd_premium - expected_stretch,
        "on_pace": ytd_premium >= expected_stretch,
        "weekly_needed": round((stretch_goal_premium - ytd_premium) / weeks_remaining, 2) if weeks_remaining > 0 else 0
    }
    
    return {
        "has_goals": True,
        "year": current_year,
        "weeks_elapsed": round(weeks_elapsed, 1),
        "weeks_remaining": round(weeks_remaining, 1),
        "goal": goal_progress,
        "stretch_goal": stretch_progress
    }

@api_router.post("/goals/team")
async def set_team_goals(goals: TeamGoalSettings, current_user: dict = Depends(get_current_user)):
    """Set team premium goals (managers only)"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can set team goals")
    
    goal_data = {
        "user_id": current_user['id'],
        "year": datetime.now(pytz_timezone('America/Chicago')).year,
        "goal_premium": goals.goal_premium,
        "stretch_goal_premium": goals.stretch_goal_premium,
        "is_team_goal": True,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.team_goals.update_one(
        {"user_id": current_user['id'], "year": goal_data["year"]},
        {"$set": goal_data},
        upsert=True
    )
    
    return {"message": "Team goals updated successfully", "goals": goal_data}

@api_router.get("/goals/team/progress")
async def get_team_goal_progress(current_user: dict = Depends(get_current_user)):
    """Get team goal progress (managers only)"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can view team goals")
    
    team_id = current_user.get('team_id')
    
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    current_year = today.year
    
    # Get team goals
    team_goal = await db.team_goals.find_one(
        {"user_id": current_user['id'], "year": current_year},
        {"_id": 0}
    )
    
    if not team_goal:
        return {
            "has_goals": False,
            "message": "No team goals set for this year"
        }
    
    # Get all subordinates - SCOPED TO TEAM
    async def get_all_subordinates(user_id: str):
        ids = []
        query = {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]}
        if team_id:
            query["team_id"] = team_id
        subordinates = await db.users.find(query, {"_id": 0, "id": 1}).to_list(1000)
        for sub in subordinates:
            ids.append(sub['id'])
            ids.extend(await get_all_subordinates(sub['id']))
        return ids
    
    team_ids = await get_all_subordinates(current_user['id'])
    team_ids.append(current_user['id'])  # Include manager
    
    # Get YTD premium for entire team - SCOPED TO TEAM
    year_start = today.replace(month=1, day=1)
    act_query = {"user_id": {"$in": team_ids}, "date": {"$gte": year_start.isoformat()}}
    if team_id:
        act_query["team_id"] = team_id
    activities = await db.activities.find(act_query, {"_id": 0}).to_list(100000)
    
    ytd_premium = sum(a.get('premium', 0) for a in activities)
    
    # Calculate weeks
    days_elapsed = (today - year_start).days
    weeks_elapsed = days_elapsed / 7
    total_weeks_in_year = 52
    weeks_remaining = total_weeks_in_year - weeks_elapsed
    
    # Goal calculations
    goal_premium = team_goal['goal_premium']
    stretch_goal_premium = team_goal['stretch_goal_premium']
    
    expected_goal = (goal_premium / total_weeks_in_year) * weeks_elapsed
    expected_stretch = (stretch_goal_premium / total_weeks_in_year) * weeks_elapsed
    
    goal_progress = {
        "current": ytd_premium,
        "goal": goal_premium,
        "expected": expected_goal,
        "percentage": round((ytd_premium / goal_premium * 100), 1) if goal_premium > 0 else 0,
        "ahead_behind": ytd_premium - expected_goal,
        "on_pace": ytd_premium >= expected_goal,
        "weekly_needed": round((goal_premium - ytd_premium) / weeks_remaining, 2) if weeks_remaining > 0 else 0
    }
    
    stretch_progress = {
        "current": ytd_premium,
        "goal": stretch_goal_premium,
        "expected": expected_stretch,
        "percentage": round((ytd_premium / stretch_goal_premium * 100), 1) if stretch_goal_premium > 0 else 0,
        "ahead_behind": ytd_premium - expected_stretch,
        "on_pace": ytd_premium >= expected_stretch,
        "weekly_needed": round((stretch_goal_premium - ytd_premium) / weeks_remaining, 2) if weeks_remaining > 0 else 0
    }
    
    return {
        "has_goals": True,
        "year": current_year,
        "team_size": len(team_ids),
        "weeks_elapsed": round(weeks_elapsed, 1),
        "weeks_remaining": round(weeks_remaining, 1),
        "goal": goal_progress,
        "stretch_goal": stretch_progress
    }

@api_router.get("/goals/team/members")
async def get_team_members_goals(current_user: dict = Depends(get_current_user)):
    """Get all team members' individual goals and progress (managers only)"""
    if current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Only managers can view team member goals")
    
    central_tz = pytz_timezone('America/Chicago')
    today = datetime.now(central_tz).date()
    current_year = today.year
    year_start = today.replace(month=1, day=1)
    
    # Get all subordinates
    async def get_all_subordinates_with_info(user_id: str):
        result = []
        subordinates = await db.users.find(
            {"manager_id": user_id, "$or": [{"status": "active"}, {"status": {"$exists": False}}]},
            {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1}
        ).to_list(1000)
        for sub in subordinates:
            result.append(sub)
            result.extend(await get_all_subordinates_with_info(sub['id']))
        return result
    
    team_members = await get_all_subordinates_with_info(current_user['id'])
    
    # Calculate weeks
    days_elapsed = (today - year_start).days
    weeks_elapsed = days_elapsed / 7
    total_weeks_in_year = 52
    weeks_remaining = total_weeks_in_year - weeks_elapsed
    
    members_progress = []
    
    for member in team_members:
        # Get member's goals
        goal = await db.goals.find_one(
            {"user_id": member['id'], "year": current_year},
            {"_id": 0}
        )
        
        # Get YTD premium
        activities = await db.activities.find({
            "user_id": member['id'],
            "date": {"$gte": year_start.isoformat()}
        }, {"_id": 0}).to_list(10000)
        
        ytd_premium = sum(a.get('premium', 0) for a in activities)
        
        if goal:
            goal_premium = goal['goal_premium']
            stretch_goal_premium = goal['stretch_goal_premium']
            expected_goal = (goal_premium / total_weeks_in_year) * weeks_elapsed
            
            members_progress.append({
                "id": member['id'],
                "name": member['name'],
                "email": member['email'],
                "role": member['role'],
                "has_goals": True,
                "ytd_premium": ytd_premium,
                "goal_premium": goal_premium,
                "stretch_goal_premium": stretch_goal_premium,
                "percentage": round((ytd_premium / goal_premium * 100), 1) if goal_premium > 0 else 0,
                "on_pace": ytd_premium >= expected_goal,
                "ahead_behind": ytd_premium - expected_goal
            })
        else:
            members_progress.append({
                "id": member['id'],
                "name": member['name'],
                "email": member['email'],
                "role": member['role'],
                "has_goals": False,
                "ytd_premium": ytd_premium
            })
    
    return {
        "year": current_year,
        "weeks_elapsed": round(weeks_elapsed, 1),
        "members": members_progress
    }

# ==================== SUITABILITY FORMS ====================

# Suitability Form configuration (can be updated without rebuilding)
SUITABILITY_FORM_CONFIG = {
    "income_ranges": [
        {"value": "25k-50k", "label": "$25,000 - $50,000"},
        {"value": "50k-75k", "label": "$50,000 - $75,000"},
        {"value": "75k-100k", "label": "$75,000 - $100,000"},
        {"value": "100k+", "label": "$100,000+"}
    ],
    "savings_ranges": [
        {"value": "0-500", "label": "$0 - $500"},
        {"value": "500-1000", "label": "$500 - $1,000"},
        {"value": "1000+", "label": "$1,000+"}
    ],
    "net_worth_ranges": [
        {"value": "0-50k", "label": "$0 - $50,000"},
        {"value": "50k-250k", "label": "$50,000 - $250,000"},
        {"value": "250k-500k", "label": "$250,000 - $500,000"},
        {"value": "500k+", "label": "$500,000+"}
    ]
}

class SuitabilityFormCreate(BaseModel):
    client_name: str
    client_phone: str
    client_address: str
    annual_income: str
    monthly_savings: str
    liquid_net_worth: str
    sale_made: bool
    agents: List[str]
    presentation_date: str
    presentation_location: str
    life_licensed: bool = True
    regional_assigned: Optional[str] = None
    notes: Optional[str] = ""
    results: Optional[str] = ""

@api_router.get("/suitability-forms/config")
async def get_suitability_form_config(current_user: dict = Depends(get_current_user)):
    """Get form configuration for dropdowns"""
    await check_feature_access(current_user, "suitability")
    return SUITABILITY_FORM_CONFIG

@api_router.get("/suitability-forms")
async def get_suitability_forms(
    current_user: dict = Depends(get_current_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    view_all: bool = False
):
    """Get suitability forms - users see their own, managers can see all"""
    await check_feature_access(current_user, "suitability")
    team_id = current_user.get('team_id')
    query = {}
    
    # Team filtering for multi-tenancy
    if team_id:
        query["team_id"] = team_id
    
    # Date filtering
    if start_date and end_date:
        query["presentation_date"] = {"$gte": start_date, "$lte": end_date}
    
    # Access control - users see their own, managers can see all with view_all flag
    if not view_all or current_user['role'] == 'agent':
        query["submitted_by"] = current_user['id']
    elif current_user['role'] in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        # Managers can view all forms from their team (scoped to team)
        if view_all:
            team_ids = await get_all_subordinates(current_user['id'], team_id)
            team_ids.append(current_user['id'])
            query["submitted_by"] = {"$in": team_ids}
    
    forms = await db.suitability_forms.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return forms

@api_router.post("/suitability-forms")
async def create_suitability_form(form_data: SuitabilityFormCreate, current_user: dict = Depends(get_current_user)):
    """Create a new suitability form"""
    await check_feature_access(current_user, "suitability")
    form_dict = {
        "id": str(uuid.uuid4()),
        "team_id": current_user.get('team_id'),  # Multi-tenancy
        **form_data.dict(),
        "submitted_by": current_user['id'],
        "submitted_by_name": current_user['name'],
        "submitted_by_email": current_user['email'],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.suitability_forms.insert_one(form_dict)
    form_dict.pop('_id', None)
    return {"message": "Suitability form submitted successfully", "form": form_dict}

@api_router.put("/suitability-forms/{form_id}")
async def update_suitability_form(form_id: str, form_data: dict, current_user: dict = Depends(get_current_user)):
    """Update a suitability form"""
    await check_feature_access(current_user, "suitability")
    existing = await db.suitability_forms.find_one({"id": form_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Form not found")
    
    # Only allow owner or managers to update
    if existing['submitted_by'] != current_user['id'] and current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Not authorized to update this form")
    
    form_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.suitability_forms.update_one({"id": form_id}, {"$set": form_data})
    
    updated = await db.suitability_forms.find_one({"id": form_id}, {"_id": 0})
    return updated

@api_router.delete("/suitability-forms/{form_id}")
async def delete_suitability_form(form_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a suitability form"""
    await check_feature_access(current_user, "suitability")
    existing = await db.suitability_forms.find_one({"id": form_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Form not found")
    
    # Only allow owner or managers to delete
    if existing['submitted_by'] != current_user['id'] and current_user['role'] not in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        raise HTTPException(status_code=403, detail="Not authorized to delete this form")
    
    await db.suitability_forms.delete_one({"id": form_id})
    return {"message": "Form deleted successfully"}

@api_router.get("/suitability-forms/export")
async def export_suitability_forms(
    current_user: dict = Depends(get_current_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "csv"
):
    """Export suitability forms to CSV"""
    await check_feature_access(current_user, "suitability")
    
    team_id = current_user.get('team_id')
    query = {}
    
    if start_date and end_date:
        query["presentation_date"] = {"$gte": start_date, "$lte": end_date}
    
    # Managers can export all, others export their own - SCOPED TO TEAM
    if current_user['role'] in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        team_ids = await get_all_subordinates(current_user['id'], team_id)
        query["submitted_by"] = {"$in": team_ids}
    else:
        query["submitted_by"] = current_user['id']
    
    forms = await db.suitability_forms.find(query, {"_id": 0}).sort("presentation_date", -1).to_list(10000)
    
    if not forms:
        raise HTTPException(status_code=404, detail="No forms found for the specified criteria")
    
    # Build CSV with proper Excel formatting
    import io
    import csv
    
    output = io.StringIO()
    # Use excel dialect and quote all fields for better Excel compatibility
    writer = csv.writer(output, dialect='excel', quoting=csv.QUOTE_ALL)
    
    # Header
    writer.writerow([
        "Client Name", "Phone", "Address", "Annual Income", "Monthly Savings",
        "Liquid Net Worth", "Sale Made", "Agents Associated/Bankers Agent #", "Presentation Date",
        "Presentation Location", "Life Licensed", "Regional Assigned", "Notes", "Results", "Submitted By", "Submitted Date"
    ])
    
    # Get labels for ranges
    income_labels = {r['value']: r['label'] for r in SUITABILITY_FORM_CONFIG['income_ranges']}
    savings_labels = {r['value']: r['label'] for r in SUITABILITY_FORM_CONFIG['savings_ranges']}
    net_worth_labels = {r['value']: r['label'] for r in SUITABILITY_FORM_CONFIG['net_worth_ranges']}
    
    for form in forms:
        writer.writerow([
            form.get('client_name', '') or '',
            form.get('client_phone', '') or '',
            form.get('client_address', '') or '',
            income_labels.get(form.get('annual_income', ''), form.get('annual_income', '')),
            savings_labels.get(form.get('monthly_savings', ''), form.get('monthly_savings', '')),
            net_worth_labels.get(form.get('liquid_net_worth', ''), form.get('liquid_net_worth', '')),
            "Yes" if form.get('sale_made') else "No",
            "; ".join(form.get('agents', [])) if form.get('agents') else '',
            form.get('presentation_date', '') or '',
            form.get('presentation_location', '') or '',
            "Yes" if form.get('life_licensed', True) else "No",
            form.get('regional_assigned', '') or '',
            (form.get('notes', '') or '').replace('\n', ' ').replace('\r', ' '),
            (form.get('results', '') or '').replace('\n', ' ').replace('\r', ' '),
            form.get('submitted_by_name', '') or '',
            form.get('created_at', '')[:10] if form.get('created_at') else ''
        ])
    
    csv_content = output.getvalue()
    output.close()
    
    # Add BOM for Excel UTF-8 recognition
    csv_with_bom = '\ufeff' + csv_content
    
    return Response(
        content=csv_with_bom,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=suitability_forms_{start_date or 'all'}_{end_date or 'all'}.csv"}
    )

@api_router.get("/suitability-forms/friday-report")
async def get_friday_report_export(
    current_user: dict = Depends(get_current_user),
    week_offset: int = 0
):
    """Export Friday Report - Professional Excel format grouped by agent"""
    await check_feature_access(current_user, "suitability")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    team_id = current_user.get('team_id')
    
    # Calculate week boundaries (Monday to Sunday)
    today = datetime.now(timezone.utc).date()
    start_of_week = today - timedelta(days=today.weekday() + (week_offset * 7))
    end_of_week = start_of_week + timedelta(days=6)
    
    start_date = start_of_week.isoformat()
    end_date = end_of_week.isoformat()
    
    query = {
        "presentation_date": {"$gte": start_date, "$lte": end_date}
    }
    
    # Managers see all team forms - SCOPED TO TEAM
    if current_user['role'] in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        team_ids = await get_all_subordinates(current_user['id'], team_id)
        query["submitted_by"] = {"$in": team_ids}
    else:
        query["submitted_by"] = current_user['id']
    
    forms = await db.suitability_forms.find(query, {"_id": 0}).sort([("submitted_by_name", 1), ("presentation_date", 1)]).to_list(10000)
    
    if not forms:
        raise HTTPException(status_code=404, detail="No forms found for this week")
    
    # Group forms by agent
    from collections import defaultdict
    by_agent = defaultdict(list)
    for form in forms:
        agent_name = form.get('submitted_by_name', 'Unknown')
        by_agent[agent_name].append(form)
    
    # Get labels for ranges
    income_labels = {r['value']: r['label'] for r in SUITABILITY_FORM_CONFIG['income_ranges']}
    savings_labels = {r['value']: r['label'] for r in SUITABILITY_FORM_CONFIG['savings_ranges']}
    net_worth_labels = {r['value']: r['label'] for r in SUITABILITY_FORM_CONFIG['net_worth_ranges']}
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Friday Report"
    
    # Define styles
    title_font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    
    subheader_font = Font(name='Arial', size=11, bold=True)
    subheader_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    
    agent_header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    agent_header_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    
    data_font = Font(name='Arial', size=10)
    
    summary_label_font = Font(name='Arial', size=11, bold=True)
    summary_value_font = Font(name='Arial', size=11)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    column_widths = [18, 14, 25, 16, 14, 16, 10, 20, 12, 15, 12, 12, 30, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    row = 1
    
    # Title Section
    ws.merge_cells(f'A{row}:N{row}')
    cell = ws.cell(row=row, column=1, value="FRIDAY SUITABILITY REPORT")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30
    row += 1
    
    ws.merge_cells(f'A{row}:N{row}')
    cell = ws.cell(row=row, column=1, value=f"Week: {start_date} to {end_date}")
    cell.font = Font(name='Arial', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    ws.merge_cells(f'A{row}:N{row}')
    cell = ws.cell(row=row, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC")
    cell.font = Font(name='Arial', size=10, italic=True)
    cell.alignment = Alignment(horizontal='center')
    row += 2
    
    # Summary Section
    total_forms = len(forms)
    total_sales = sum(1 for f in forms if f.get('sale_made'))
    conversion_rate = round((total_sales / total_forms * 100), 1) if total_forms > 0 else 0
    
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1, value="WEEKLY SUMMARY")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = header_fill
    row += 1
    
    summary_data = [
        ("Total Forms", total_forms),
        ("Total Sales Made", total_sales),
        ("Conversion Rate", f"{conversion_rate}%"),
        ("Number of Agents", len(by_agent))
    ]
    
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = summary_label_font
        ws.cell(row=row, column=2, value=value).font = summary_value_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        row += 1
    
    row += 1
    
    # Agent Summary Table
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1, value="AGENT SUMMARY")
    cell.font = header_font
    cell.fill = header_fill
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = header_fill
    row += 1
    
    agent_headers = ["Agent Name", "Forms", "Sales", "Conv. Rate"]
    for col, header in enumerate(agent_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    for agent_name in sorted(by_agent.keys()):
        agent_forms = by_agent[agent_name]
        agent_sales = sum(1 for f in agent_forms if f.get('sale_made'))
        agent_rate = round((agent_sales / len(agent_forms) * 100), 1) if agent_forms else 0
        
        ws.cell(row=row, column=1, value=agent_name).font = data_font
        ws.cell(row=row, column=2, value=len(agent_forms)).font = data_font
        ws.cell(row=row, column=3, value=agent_sales).font = data_font
        ws.cell(row=row, column=4, value=f"{agent_rate}%").font = data_font
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            if col > 1:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        row += 1
    
    row += 2
    
    # Detailed Forms by Agent
    detail_headers = [
        "Client Name", "Phone", "Address", "Annual Income", "Monthly Savings",
        "Net Worth", "Sale", "Agents/Bankers #", "Date", "Location", 
        "Licensed", "Regional", "Notes", "Results"
    ]
    
    for agent_name in sorted(by_agent.keys()):
        agent_forms = by_agent[agent_name]
        agent_sales = sum(1 for f in agent_forms if f.get('sale_made'))
        
        # Agent Header
        ws.merge_cells(f'A{row}:N{row}')
        cell = ws.cell(row=row, column=1, value=f"{agent_name} ({len(agent_forms)} forms, {agent_sales} sales)")
        cell.font = agent_header_font
        cell.fill = agent_header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 22
        for col in range(1, 15):
            ws.cell(row=row, column=col).fill = agent_header_fill
        row += 1
        
        # Column Headers
        for col, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[row].height = 20
        row += 1
        
        # Data rows
        for form in agent_forms:
            data = [
                form.get('client_name', '') or '',
                form.get('client_phone', '') or '',
                form.get('client_address', '') or '',
                income_labels.get(form.get('annual_income', ''), form.get('annual_income', '')),
                savings_labels.get(form.get('monthly_savings', ''), form.get('monthly_savings', '')),
                net_worth_labels.get(form.get('liquid_net_worth', ''), form.get('liquid_net_worth', '')),
                "Yes" if form.get('sale_made') else "No",
                "; ".join(form.get('agents', [])) if form.get('agents') else '',
                form.get('presentation_date', '') or '',
                form.get('presentation_location', '') or '',
                "Yes" if form.get('life_licensed', True) else "No",
                form.get('regional_assigned', '') or '',
                (form.get('notes', '') or '').replace('\n', ' ').replace('\r', ' '),
                (form.get('results', '') or '').replace('\n', ' ').replace('\r', ' ')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Color sale column
                if col == 7:
                    if value == "Yes":
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            row += 1
        
        row += 1  # Blank row between agents
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Friday_Report_{start_date}_to_{end_date}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/suitability-forms/weekly-report")
async def get_weekly_suitability_report(
    current_user: dict = Depends(get_current_user),
    week_offset: int = 0
):
    """Get weekly suitability forms report"""
    await check_feature_access(current_user, "suitability")
    
    team_id = current_user.get('team_id')
    
    # Calculate week boundaries
    today = datetime.now(timezone.utc).date()
    start_of_week = today - timedelta(days=today.weekday() + (week_offset * 7))
    end_of_week = start_of_week + timedelta(days=6)
    
    start_date = start_of_week.isoformat()
    end_date = end_of_week.isoformat()
    
    query = {
        "presentation_date": {"$gte": start_date, "$lte": end_date}
    }
    
    # Managers see all team forms - SCOPED TO TEAM
    if current_user['role'] in ['super_admin', 'state_manager', 'regional_manager', 'district_manager']:
        team_ids = await get_all_subordinates(current_user['id'], team_id)
        query["submitted_by"] = {"$in": team_ids}
    else:
        query["submitted_by"] = current_user['id']
    
    forms = await db.suitability_forms.find(query, {"_id": 0}).sort("presentation_date", -1).to_list(1000)
    
    # Calculate stats
    total_forms = len(forms)
    sales_made = sum(1 for f in forms if f.get('sale_made'))
    
    # Group by agent
    by_agent = {}
    for form in forms:
        name = form.get('submitted_by_name', 'Unknown')
        if name not in by_agent:
            by_agent[name] = {"total": 0, "sales": 0}
        by_agent[name]["total"] += 1
        if form.get('sale_made'):
            by_agent[name]["sales"] += 1
    
    return {
        "week_start": start_date,
        "week_end": end_date,
        "total_forms": total_forms,
        "sales_made": sales_made,
        "conversion_rate": round((sales_made / total_forms * 100), 1) if total_forms > 0 else 0,
        "by_agent": by_agent,
        "forms": forms
    }

# ==================== FACT FINDER ====================

class FactFinderHealthExpenses(BaseModel):
    choose_physician: Optional[int] = None
    coverage_traveling: Optional[int] = None
    personal_agent: Optional[int] = None
    affordability: Optional[int] = None
    critical_illness: Optional[int] = None

class FactFinderRetirementIncome(BaseModel):
    safety_principal: Optional[int] = None
    transferring_assets: Optional[int] = None
    minimizing_taxes: Optional[int] = None
    accessibility_money: Optional[int] = None
    rate_return: Optional[int] = None
    outliving_assets: Optional[int] = None

class FactFinderFinalExpenses(BaseModel):
    funeral_costs: Optional[int] = None
    survivor_income: Optional[int] = None
    legacy_giving: Optional[int] = None
    charitable_giving: Optional[int] = None
    living_benefits: Optional[int] = None

class FactFinderExtendedCare(BaseModel):
    remaining_independent: Optional[int] = None
    protecting_assets: Optional[int] = None
    care_location_choices: Optional[int] = None
    not_burdening_family: Optional[int] = None
    how_remembered: Optional[int] = None

class FactFinderClientInfo(BaseModel):
    first_name: Optional[str] = ""
    last_name: Optional[str] = ""
    birth_date: Optional[str] = ""
    spouse_first: Optional[str] = ""
    spouse_last: Optional[str] = ""
    spouse_birth_date: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    state: Optional[str] = ""
    zip_code: Optional[str] = ""
    employer: Optional[str] = ""
    employer_retired: Optional[bool] = False
    spouse_employer: Optional[str] = ""
    spouse_employer_retired: Optional[bool] = False
    email: Optional[str] = ""
    phone: Optional[str] = ""

class FactFinderCreate(BaseModel):
    client_info: FactFinderClientInfo
    health_expenses: FactFinderHealthExpenses
    retirement_income: FactFinderRetirementIncome
    final_expenses: FactFinderFinalExpenses
    extended_care: FactFinderExtendedCare
    producer_name_1: Optional[str] = ""
    producer_name_2: Optional[str] = ""
    agent_number_1: Optional[str] = ""
    agent_number_2: Optional[str] = ""
    notes: Optional[str] = ""
    status: Optional[str] = "draft"

@api_router.get("/fact-finders")
async def get_fact_finders(
    current_user: dict = Depends(get_current_user),
    month: Optional[str] = None,
    search: Optional[str] = None,
    created_by: Optional[str] = None
):
    """Get fact finders for user's team"""
    await check_feature_access(current_user, "fact_finder")
    
    team_id = current_user.get('team_id')
    query = {}
    
    # Team scoping - ALL users (including super_admin) are scoped to their assigned team
    if not team_id:
        raise HTTPException(status_code=403, detail="You must be assigned to a team")
    query["team_id"] = team_id
    
    # Access control - regular users see their own, managers see team
    if current_user['role'] not in ['super_admin', 'state_manager']:
        query["created_by"] = current_user['id']
    
    # Filters
    if month:
        query["month_key"] = month
    
    if created_by:
        query["created_by"] = created_by
    
    if search:
        query["$or"] = [
            {"client_info.first_name": {"$regex": search, "$options": "i"}},
            {"client_info.last_name": {"$regex": search, "$options": "i"}},
            {"client_info.email": {"$regex": search, "$options": "i"}},
            {"client_info.city": {"$regex": search, "$options": "i"}}
        ]
    
    fact_finders = await db.fact_finders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Add creator name
    user_ids = list(set(ff.get('created_by') for ff in fact_finders if ff.get('created_by')))
    users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    user_map = {u['id']: u['name'] for u in users}
    
    for ff in fact_finders:
        ff['creator_name'] = user_map.get(ff.get('created_by'), 'Unknown')
    
    return fact_finders

@api_router.post("/fact-finders")
async def create_fact_finder(data: FactFinderCreate, current_user: dict = Depends(get_current_user)):
    """Create a new fact finder"""
    await check_feature_access(current_user, "fact_finder")
    
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail="You must be assigned to a team")
    
    now = datetime.now(timezone.utc)
    
    fact_finder = {
        "id": str(uuid.uuid4()),
        "team_id": team_id,
        "created_by": current_user['id'],
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        "month_key": now.strftime("%Y-%m"),
        "client_info": data.client_info.model_dump(),
        "health_expenses": data.health_expenses.model_dump(),
        "retirement_income": data.retirement_income.model_dump(),
        "final_expenses": data.final_expenses.model_dump(),
        "extended_care": data.extended_care.model_dump(),
        "producer_name_1": data.producer_name_1,
        "producer_name_2": data.producer_name_2,
        "agent_number_1": data.agent_number_1,
        "agent_number_2": data.agent_number_2,
        "notes": data.notes,
        "status": data.status
    }
    
    await db.fact_finders.insert_one(fact_finder)
    fact_finder.pop('_id', None)
    
    return {"message": "Fact Finder created", "fact_finder": fact_finder}

@api_router.get("/fact-finders/{fact_finder_id}")
async def get_fact_finder(fact_finder_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific fact finder"""
    await check_feature_access(current_user, "fact_finder")
    
    fact_finder = await db.fact_finders.find_one({"id": fact_finder_id}, {"_id": 0})
    if not fact_finder:
        raise HTTPException(status_code=404, detail="Fact Finder not found")
    
    # Access check
    team_id = current_user.get('team_id')
    if current_user['role'] != 'super_admin':
        if fact_finder.get('team_id') != team_id:
            raise HTTPException(status_code=403, detail="Access denied")
        if current_user['role'] not in ['state_manager'] and fact_finder.get('created_by') != current_user['id']:
            raise HTTPException(status_code=403, detail="Access denied")
    
    return fact_finder

@api_router.put("/fact-finders/{fact_finder_id}")
async def update_fact_finder(fact_finder_id: str, data: FactFinderCreate, current_user: dict = Depends(get_current_user)):
    """Update a fact finder"""
    await check_feature_access(current_user, "fact_finder")
    
    existing = await db.fact_finders.find_one({"id": fact_finder_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Fact Finder not found")
    
    # Access check
    team_id = current_user.get('team_id')
    if current_user['role'] != 'super_admin':
        if existing.get('team_id') != team_id:
            raise HTTPException(status_code=403, detail="Access denied")
        if current_user['role'] not in ['state_manager'] and existing.get('created_by') != current_user['id']:
            raise HTTPException(status_code=403, detail="Only the creator can edit")
    
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "client_info": data.client_info.model_dump(),
        "health_expenses": data.health_expenses.model_dump(),
        "retirement_income": data.retirement_income.model_dump(),
        "final_expenses": data.final_expenses.model_dump(),
        "extended_care": data.extended_care.model_dump(),
        "producer_name_1": data.producer_name_1,
        "producer_name_2": data.producer_name_2,
        "agent_number_1": data.agent_number_1,
        "agent_number_2": data.agent_number_2,
        "notes": data.notes,
        "status": data.status
    }
    
    await db.fact_finders.update_one({"id": fact_finder_id}, {"$set": update_data})
    
    updated = await db.fact_finders.find_one({"id": fact_finder_id}, {"_id": 0})
    return updated

@api_router.delete("/fact-finders/{fact_finder_id}")
async def delete_fact_finder(fact_finder_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a fact finder (creator or state_manager only)"""
    await check_feature_access(current_user, "fact_finder")
    
    existing = await db.fact_finders.find_one({"id": fact_finder_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Fact Finder not found")
    
    # Access check
    team_id = current_user.get('team_id')
    if current_user['role'] != 'super_admin':
        if existing.get('team_id') != team_id:
            raise HTTPException(status_code=403, detail="Access denied")
        if current_user['role'] not in ['state_manager'] and existing.get('created_by') != current_user['id']:
            raise HTTPException(status_code=403, detail="Only the creator or State Manager can delete")
    
    await db.fact_finders.delete_one({"id": fact_finder_id})
    return {"message": "Fact Finder deleted"}

@api_router.post("/fact-finders/{fact_finder_id}/duplicate")
async def duplicate_fact_finder(fact_finder_id: str, current_user: dict = Depends(get_current_user)):
    """Duplicate a fact finder"""
    await check_feature_access(current_user, "fact_finder")
    
    existing = await db.fact_finders.find_one({"id": fact_finder_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Fact Finder not found")
    
    # Access check
    team_id = current_user.get('team_id')
    if current_user['role'] != 'super_admin':
        if existing.get('team_id') != team_id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    now = datetime.now(timezone.utc)
    
    new_fact_finder = {
        **existing,
        "id": str(uuid.uuid4()),
        "created_by": current_user['id'],
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        "month_key": now.strftime("%Y-%m"),
        "status": "draft"
    }
    
    # Update client name to indicate it's a copy
    if new_fact_finder.get('client_info', {}).get('last_name'):
        new_fact_finder['client_info']['last_name'] += " (Copy)"
    
    await db.fact_finders.insert_one(new_fact_finder)
    new_fact_finder.pop('_id', None)
    
    return {"message": "Fact Finder duplicated", "fact_finder": new_fact_finder}

@api_router.get("/fact-finders/{fact_finder_id}/pdf")
async def export_fact_finder_pdf(fact_finder_id: str, current_user: dict = Depends(get_current_user)):
    """Export fact finder as PDF"""
    await check_feature_access(current_user, "fact_finder")
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    import io
    
    fact_finder = await db.fact_finders.find_one({"id": fact_finder_id}, {"_id": 0})
    if not fact_finder:
        raise HTTPException(status_code=404, detail="Fact Finder not found")
    
    # Access check
    team_id = current_user.get('team_id')
    if current_user['role'] != 'super_admin':
        if fact_finder.get('team_id') != team_id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    client_info = fact_finder.get('client_info', {})
    health = fact_finder.get('health_expenses', {})
    retirement = fact_finder.get('retirement_income', {})
    final = fact_finder.get('final_expenses', {})
    extended = fact_finder.get('extended_care', {})
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=30, bottomMargin=30, leftMargin=40, rightMargin=40)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, textColor=colors.HexColor('#8B4513'))
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=11, spaceBefore=10, spaceAfter=5, textColor=colors.HexColor('#1e40af'))
    
    # Header
    elements.append(Paragraph("PMA USA - Fact Finder", title_style))
    elements.append(Paragraph("You may have spent a lifetime accumulating assets", ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=12, alignment=1, textColor=colors.HexColor('#8B4513'), spaceAfter=15)))
    
    def rating_display(value):
        """Convert rating value to visual display"""
        if value is None:
            return "○ ○ ○ ○ ○"
        circles = ["○"] * 5
        if 1 <= value <= 5:
            circles[value - 1] = "●"
        return " ".join(circles)
    
    def build_rating_table(title, items, data):
        """Build a rating section table"""
        header = [title, '1', '2', '3', '4', '5']
        rows = [header]
        for label, key in items:
            val = data.get(key)
            row = [label]
            for i in range(1, 6):
                row.append("●" if val == i else "○")
            rows.append(row)
        
        table = Table(rows, colWidths=[180, 25, 25, 25, 25, 25])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4a574')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#8B7355')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FFF8DC')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        return table
    
    # Health Expenses
    health_items = [
        ("Choose your physician", "choose_physician"),
        ("Coverage when traveling", "coverage_traveling"),
        ("Personal agent", "personal_agent"),
        ("Affordability", "affordability"),
        ("Coverage for Critical Illness", "critical_illness"),
    ]
    
    # Retirement Income
    retirement_items = [
        ("Safety of principal", "safety_principal"),
        ("Transferring assets", "transferring_assets"),
        ("Minimizing taxes", "minimizing_taxes"),
        ("Accessibility of money", "accessibility_money"),
        ("Rate of return", "rate_return"),
        ("Outliving assets", "outliving_assets"),
    ]
    
    # Final Expenses
    final_items = [
        ("Funeral costs", "funeral_costs"),
        ("Survivor income", "survivor_income"),
        ("Legacy giving", "legacy_giving"),
        ("Charitable giving", "charitable_giving"),
        ("Living benefits of Life Insurance", "living_benefits"),
    ]
    
    # Extended Care
    extended_items = [
        ("Remaining independent", "remaining_independent"),
        ("Protecting assets", "protecting_assets"),
        ("Having choices in care location", "care_location_choices"),
        ("Not burdening friends or family", "not_burdening_family"),
        ("How you are remembered", "how_remembered"),
    ]
    
    # Two-column layout for rating sections
    left_col = []
    right_col = []
    
    left_col.append(build_rating_table("Health Expenses:", health_items, health))
    left_col.append(Spacer(1, 10))
    left_col.append(build_rating_table("Final Expenses:", final_items, final))
    
    right_col.append(build_rating_table("Retirement Income:", retirement_items, retirement))
    right_col.append(Spacer(1, 10))
    right_col.append(build_rating_table("Extended Care:", extended_items, extended))
    
    # Create side-by-side layout
    from reportlab.platypus import KeepTogether
    
    # Add rating sections as a 2-column table
    rating_table = Table([[
        left_col[0], right_col[0]
    ]], colWidths=[280, 280])
    rating_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(rating_table)
    elements.append(Spacer(1, 10))
    
    rating_table2 = Table([[
        left_col[2], right_col[2]
    ]], colWidths=[280, 280])
    rating_table2.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(rating_table2)
    elements.append(Spacer(1, 15))
    
    # Producer Info
    producer_data = [
        ["Producer's name:", fact_finder.get('producer_name_1', ''), "Bankers Life Agent Number:", fact_finder.get('agent_number_1', '')],
        ["Producer's name:", fact_finder.get('producer_name_2', ''), "Bankers Life Agent Number:", fact_finder.get('agent_number_2', '')],
    ]
    producer_table = Table(producer_data, colWidths=[90, 150, 140, 150])
    producer_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(producer_table)
    elements.append(Spacer(1, 15))
    
    # Client Information Section
    elements.append(Paragraph("Client Information", section_style))
    
    client_data = [
        ["First Name:", client_info.get('first_name', ''), "Last Name:", client_info.get('last_name', ''), "Birth Date:", client_info.get('birth_date', '')],
        ["Spouse First:", client_info.get('spouse_first', ''), "Last:", client_info.get('spouse_last', ''), "Birth Date:", client_info.get('spouse_birth_date', '')],
        ["Address:", client_info.get('address', ''), "City:", client_info.get('city', ''), "State:", client_info.get('state', ''), "Zip:", client_info.get('zip_code', '')],
        ["Employer:", client_info.get('employer', ''), "Retired:", "☑" if client_info.get('employer_retired') else "☐", "Spouse Employer:", client_info.get('spouse_employer', ''), "Retired:", "☑" if client_info.get('spouse_employer_retired') else "☐"],
        ["E-mail:", client_info.get('email', ''), "", "", "Phone #:", client_info.get('phone', '')],
    ]
    
    client_table = Table(client_data, colWidths=[60, 100, 50, 100, 55, 50, 35, 60])
    client_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFFACD')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#8B7355')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(client_table)
    elements.append(Spacer(1, 15))
    
    # Notes
    if fact_finder.get('notes'):
        elements.append(Paragraph("Notes", section_style))
        notes_table = Table([[fact_finder.get('notes', '')]], colWidths=[530])
        notes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFFACD')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#8B7355')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(notes_table)
    
    doc.build(elements)
    pdf_content = buffer.getvalue()
    
    # Build filename
    last_name = client_info.get('last_name', 'Unknown').replace(' ', '_')
    first_name = client_info.get('first_name', 'Client').replace(' ', '_')
    date_str = datetime.now().strftime('%Y-%m-%d')
    filename = f"FactFinder_{last_name}_{first_name}_{date_str}.pdf"
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/fact-finders/months/list")
async def get_fact_finder_months(current_user: dict = Depends(get_current_user)):
    """Get list of months that have fact finders"""
    await check_feature_access(current_user, "fact_finder")
    
    team_id = current_user.get('team_id')
    if not team_id:
        raise HTTPException(status_code=403, detail="You must be assigned to a team")
    
    # ALL users (including super_admin) are scoped to their assigned team
    query = {"team_id": team_id}
    
    if current_user['role'] not in ['super_admin', 'state_manager']:
        query["created_by"] = current_user['id']
    
    pipeline = [
        {"$match": query},
        {"$group": {"_id": "$month_key", "count": {"$sum": 1}}},
        {"$sort": {"_id": -1}}
    ]
    
    months = await db.fact_finders.aggregate(pipeline).to_list(100)
    return [{"month": m["_id"], "count": m["count"]} for m in months]

# Health check endpoint accessible via /api/health
@api_router.get("/health")
async def api_health_check():
    """Health check endpoint accessible via /api/health"""
    try:
        await db.command('ping')
        return {
            "status": "healthy",
            "service": "crm-sales-tracker-backend",
            "database": "connected"
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail="Service unavailable")

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)



# Health check endpoint (not under /api prefix for Kubernetes)
@app.get("/health")
async def health_check():
    """Health check endpoint for Kubernetes/deployment systems"""
    try:
        # Quick ping to MongoDB to verify connection
        await db.command('ping')
        return {
            "status": "healthy",
            "service": "crm-sales-tracker-backend",
            "database": "connected"
        }
    except Exception as e:
        logger.error(f"Health check failed: {str(e)}")
        raise HTTPException(status_code=503, detail="Service unavailable")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()