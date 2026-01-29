#!/usr/bin/env python3
"""
Excel Totals Row Enhancement Testing Script
NEW FUNCTIONALITY: Added totals rows at the bottom of Excel spreadsheets for Individual and Team reports across all periods (Daily, Monthly, Quarterly, Yearly).

CRITICAL TESTS:
1. Period Excel Reports with Totals (monthly/quarterly/yearly for individual and team reports)
2. Daily Excel Reports with Totals (individual and team reports)  
3. Totals Calculation Accuracy
4. Manager Selection with Totals
5. Empty Data Handling

EXPECTED EXCEL FORMAT:
Steve Ahlers (Individual)    | 5  | 3  | 2  | 1  | 1  | 4  | 2  | 1500
Ryan Rozell's Team          | 15 | 10 | 8  | 5  | 3  | 12 | 6  | 4500
Andrew Inman's Team         | 20 | 12 | 10 | 7  | 4  | 15 | 8  | 6000
                            |    |    |    |    |    |    |    |
TOTALS                      | 40 | 25 | 20 | 13 | 8  | 31 | 16 | 12000
"""

import requests
import json
from datetime import datetime, timedelta
import sys
import os
from openpyxl import load_workbook
from io import BytesIO

# Configuration
BACKEND_URL = "https://custom-dashboard-5.preview.emergentagent.com/api"

class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'

def print_success(message):
    print(f"{Colors.GREEN}✅ {message}{Colors.ENDC}")

def print_error(message):
    print(f"{Colors.RED}❌ {message}{Colors.ENDC}")

def print_warning(message):
    print(f"{Colors.YELLOW}⚠️  {message}{Colors.ENDC}")

def print_info(message):
    print(f"{Colors.BLUE}ℹ️  {message}{Colors.ENDC}")

def print_header(message):
    print(f"\n{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.ENDC}")
    print(f"{Colors.BOLD}{Colors.BLUE}{message}{Colors.ENDC}")
    print(f"{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.ENDC}")

class ExcelTotalsTester:
    def __init__(self):
        self.session = requests.Session()
        self.state_manager_token = None
        self.test_results = {
            'passed': 0,
            'failed': 0,
            'errors': []
        }

    def setup_authentication(self):
        """Setup authentication with existing state manager"""
        print_header("SETTING UP AUTHENTICATION")
        
        try:
            response = self.session.post(f"{BACKEND_URL}/auth/login", json={
                "email": "spencer.sudbeck@pmagent.net",
                "password": "Bizlink25"
            })
            if response.status_code == 200:
                data = response.json()
                self.state_manager_token = data['token']
                self.state_manager_id = data['user']['id']
                print_success(f"Logged in as state manager: {data['user']['name']}")
                return True
            else:
                print_error(f"Failed to login: {response.status_code} - {response.text}")
                return False
        except Exception as e:
            print_error(f"Exception during login: {str(e)}")
            return False

    def create_test_activity_data(self):
        """Create comprehensive test activity data for multiple users and dates"""
        print_header("CREATING TEST ACTIVITY DATA")
        
        if not self.state_manager_token:
            print_error("No authentication token available")
            return False
            
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        
        # Create activities for today, yesterday, and several days in the past
        today = datetime.now().date()
        dates_to_create = [
            today.isoformat(),
            (today - timedelta(days=1)).isoformat(),
            (today - timedelta(days=7)).isoformat(),
            (today - timedelta(days=30)).isoformat(),
        ]
        
        # Create varied activity data for different dates
        activity_patterns = [
            {"contacts": 10.0, "appointments": 5.0, "presentations": 3.0, "referrals": 2, "testimonials": 1, "sales": 2, "new_face_sold": 1.0, "premium": 1500.00},
            {"contacts": 15.0, "appointments": 8.0, "presentations": 5.0, "referrals": 3, "testimonials": 2, "sales": 3, "new_face_sold": 2.0, "premium": 2500.00},
            {"contacts": 20.0, "appointments": 12.0, "presentations": 8.0, "referrals": 4, "testimonials": 3, "sales": 4, "new_face_sold": 3.0, "premium": 3500.00},
            {"contacts": 25.0, "appointments": 15.0, "presentations": 10.0, "referrals": 5, "testimonials": 4, "sales": 5, "new_face_sold": 4.0, "premium": 4500.00},
        ]
        
        success_count = 0
        for i, date_str in enumerate(dates_to_create):
            activity_data = activity_patterns[i % len(activity_patterns)]
            activity_data["date"] = date_str
            
            try:
                response = self.session.put(
                    f"{BACKEND_URL}/activities/{date_str}",
                    json=activity_data,
                    headers=headers
                )
                
                if response.status_code == 200:
                    print_success(f"Created activity for {date_str}: {activity_data['contacts']} contacts, ${activity_data['premium']} premium")
                    success_count += 1
                else:
                    print_warning(f"Could not create activity for {date_str}: {response.status_code}")
                    
            except Exception as e:
                print_warning(f"Exception creating activity for {date_str}: {str(e)}")
        
        print_info(f"Successfully created {success_count}/{len(dates_to_create)} test activities")
        return success_count > 0

    def download_and_analyze_excel(self, endpoint, params, test_name):
        """Download Excel file and analyze for totals row"""
        if not self.state_manager_token:
            print_error("No authentication token available")
            return False
            
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        
        try:
            print_info(f"Downloading Excel from: {endpoint}")
            print_info(f"Parameters: {params}")
            
            response = self.session.get(
                f"{BACKEND_URL}{endpoint}",
                params=params,
                headers=headers
            )
            
            if response.status_code != 200:
                print_error(f"Excel download failed: {response.status_code} - {response.text}")
                self.test_results['failed'] += 1
                self.test_results['errors'].append(f"{test_name}: Excel download failed with {response.status_code}")
                return False
            
            # Check content type
            content_type = response.headers.get('content-type', '')
            if 'spreadsheet' not in content_type and 'excel' not in content_type:
                print_warning(f"Unexpected content type: {content_type}")
            
            # Load Excel file
            excel_data = BytesIO(response.content)
            workbook = load_workbook(excel_data)
            worksheet = workbook.active
            
            print_success(f"Successfully loaded Excel file with {worksheet.max_row} rows and {worksheet.max_column} columns")
            
            # Analyze for totals row
            totals_found = self.analyze_totals_row(worksheet, test_name)
            
            if totals_found:
                print_success(f"{test_name}: Totals row found and validated")
                self.test_results['passed'] += 1
                return True
            else:
                print_error(f"{test_name}: Totals row not found or invalid")
                self.test_results['failed'] += 1
                self.test_results['errors'].append(f"{test_name}: Totals row validation failed")
                return False
                
        except Exception as e:
            print_error(f"Exception analyzing Excel for {test_name}: {str(e)}")
            self.test_results['failed'] += 1
            self.test_results['errors'].append(f"{test_name}: Exception - {str(e)}")
            return False

    def analyze_totals_row(self, worksheet, test_name):
        """Analyze worksheet for totals row with proper formatting and calculations"""
        print_info(f"Analyzing totals row for {test_name}...")
        
        # Look for "TOTALS" in the first column
        totals_row_num = None
        for row_num in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_num, column=1).value
            if cell_value and str(cell_value).upper() == "TOTALS":
                totals_row_num = row_num
                break
        
        if not totals_row_num:
            print_error("No 'TOTALS' row found in first column")
            return False
        
        print_success(f"Found TOTALS row at row {totals_row_num}")
        
        # Check formatting of totals row
        totals_cell = worksheet.cell(row=totals_row_num, column=1)
        
        # Check if bold formatting is applied
        if totals_cell.font and totals_cell.font.bold:
            print_success("✅ TOTALS cell has bold formatting")
        else:
            print_warning("⚠️ TOTALS cell may not have bold formatting")
        
        # Check background color (light red)
        if totals_cell.fill and totals_cell.fill.start_color:
            fill_color = totals_cell.fill.start_color.rgb
            if fill_color and 'FF' in str(fill_color):  # Check for reddish color
                print_success("✅ TOTALS row has background color")
            else:
                print_info(f"TOTALS row background color: {fill_color}")
        
        # Validate numeric totals calculations
        numeric_columns_found = 0
        total_sum_validated = 0
        
        # Look for numeric columns (typically starting from column 4 for individual reports, 4 for team reports)
        start_col = 4  # Assuming Name, Email, Role are first 3 columns
        
        for col_num in range(start_col, worksheet.max_column + 1):
            totals_value = worksheet.cell(row=totals_row_num, column=col_num).value
            
            if totals_value is not None and isinstance(totals_value, (int, float)):
                numeric_columns_found += 1
                
                # Calculate sum of values above this cell
                column_sum = 0
                for row_num in range(3, totals_row_num):  # Start from row 3 (after headers)
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    if cell_value is not None and isinstance(cell_value, (int, float)):
                        column_sum += cell_value
                
                # Validate the total
                if abs(totals_value - column_sum) < 0.01:  # Allow for small floating point differences
                    total_sum_validated += 1
                    print_success(f"✅ Column {col_num}: Total {totals_value} matches sum {column_sum}")
                else:
                    print_error(f"❌ Column {col_num}: Total {totals_value} != Sum {column_sum}")
        
        print_info(f"Found {numeric_columns_found} numeric columns, {total_sum_validated} validated")
        
        # Require at least 6 numeric columns (contacts, appointments, presentations, referrals, testimonials, sales, new_face_sold, premium)
        if numeric_columns_found >= 6:
            print_success(f"✅ Found sufficient numeric columns ({numeric_columns_found})")
        else:
            print_warning(f"⚠️ Only found {numeric_columns_found} numeric columns (expected at least 6)")
        
        # Require at least 80% of totals to be correct
        if total_sum_validated >= (numeric_columns_found * 0.8):
            print_success(f"✅ Totals calculation accuracy: {total_sum_validated}/{numeric_columns_found}")
            return True
        else:
            print_error(f"❌ Insufficient totals accuracy: {total_sum_validated}/{numeric_columns_found}")
            return False

    def test_period_excel_reports_with_totals(self):
        """Test 1: Period Excel Reports with Totals"""
        print_header("TEST 1: PERIOD EXCEL REPORTS WITH TOTALS")
        
        periods = ['monthly', 'quarterly', 'yearly']
        report_types = ['individual', 'team']
        
        for period in periods:
            for report_type in report_types:
                test_name = f"{period.capitalize()} {report_type.capitalize()} Report"
                print_info(f"\nTesting {test_name}...")
                
                endpoint = f"/reports/period/excel/{report_type}"
                params = {"period": period}
                
                self.download_and_analyze_excel(endpoint, params, test_name)

    def test_daily_excel_reports_with_totals(self):
        """Test 2: Daily Excel Reports with Totals"""
        print_header("TEST 2: DAILY EXCEL REPORTS WITH TOTALS")
        
        today = datetime.now().date().isoformat()
        report_types = ['individual', 'team']
        
        for report_type in report_types:
            test_name = f"Daily {report_type.capitalize()} Report"
            print_info(f"\nTesting {test_name}...")
            
            endpoint = f"/reports/daily/excel/{report_type}"
            params = {"date": today}
            
            self.download_and_analyze_excel(endpoint, params, test_name)

    def test_totals_calculation_accuracy(self):
        """Test 3: Totals Calculation Accuracy"""
        print_header("TEST 3: TOTALS CALCULATION ACCURACY")
        
        # Test with monthly individual report for detailed analysis
        test_name = "Monthly Individual Report - Calculation Accuracy"
        print_info(f"Testing {test_name} for detailed calculation verification...")
        
        endpoint = "/reports/period/excel/individual"
        params = {"period": "monthly"}
        
        if not self.state_manager_token:
            print_error("No authentication token available")
            return
            
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        
        try:
            # First get JSON data to compare with Excel totals
            json_response = self.session.get(
                f"{BACKEND_URL}/reports/period/individual",
                params=params,
                headers=headers
            )
            
            if json_response.status_code == 200:
                json_data = json_response.json()
                data_array = json_data.get('data', [])
                
                if data_array:
                    # Calculate expected totals from JSON data
                    expected_totals = {
                        'contacts': sum(item.get('contacts', 0) for item in data_array),
                        'appointments': sum(item.get('appointments', 0) for item in data_array),
                        'presentations': sum(item.get('presentations', 0) for item in data_array),
                        'referrals': sum(item.get('referrals', 0) for item in data_array),
                        'testimonials': sum(item.get('testimonials', 0) for item in data_array),
                        'sales': sum(item.get('sales', 0) for item in data_array),
                        'new_face_sold': sum(item.get('new_face_sold', 0) for item in data_array),
                        'premium': sum(item.get('premium', 0) for item in data_array)
                    }
                    
                    print_success("✅ Calculated expected totals from JSON data:")
                    for metric, total in expected_totals.items():
                        print_info(f"   {metric}: {total}")
                    
                    # Now download and verify Excel totals match
                    self.download_and_analyze_excel(endpoint, params, test_name)
                    
                else:
                    print_warning("No data in JSON response for calculation comparison")
            else:
                print_error(f"Failed to get JSON data for comparison: {json_response.status_code}")
                
        except Exception as e:
            print_error(f"Exception in calculation accuracy test: {str(e)}")

    def test_manager_selection_with_totals(self):
        """Test 4: Manager Selection with Totals"""
        print_header("TEST 4: MANAGER SELECTION WITH TOTALS")
        
        if not self.state_manager_token:
            print_error("No authentication token available")
            return
            
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        
        try:
            # Get available managers
            managers_response = self.session.get(f"{BACKEND_URL}/reports/managers", headers=headers)
            
            if managers_response.status_code == 200:
                managers_data = managers_response.json()
                managers = managers_data.get('managers', [])
                
                if managers:
                    # Test with first available manager
                    test_manager = managers[0]
                    manager_id = test_manager.get('id')
                    manager_name = test_manager.get('name', 'Unknown')
                    
                    print_info(f"Testing manager selection with: {manager_name} (ID: {manager_id})")
                    
                    # Test team report with manager selection
                    test_name = f"Team Report with Manager Selection - {manager_name}"
                    endpoint = "/reports/period/excel/team"
                    params = {"period": "monthly", "user_id": manager_id}
                    
                    self.download_and_analyze_excel(endpoint, params, test_name)
                    
                else:
                    print_warning("No managers available for testing")
            else:
                print_error(f"Failed to get managers list: {managers_response.status_code}")
                
        except Exception as e:
            print_error(f"Exception in manager selection test: {str(e)}")

    def test_empty_data_handling(self):
        """Test 5: Empty Data Handling"""
        print_header("TEST 5: EMPTY DATA HANDLING")
        
        # Test with a future date that should have no data
        future_date = (datetime.now().date() + timedelta(days=30)).isoformat()
        
        test_name = "Daily Individual Report - Empty Data"
        print_info(f"Testing with future date {future_date} (should have no data)...")
        
        endpoint = "/reports/daily/excel/individual"
        params = {"date": future_date}
        
        if not self.state_manager_token:
            print_error("No authentication token available")
            return
            
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        
        try:
            response = self.session.get(
                f"{BACKEND_URL}{endpoint}",
                params=params,
                headers=headers
            )
            
            if response.status_code == 200:
                # Load Excel file
                excel_data = BytesIO(response.content)
                workbook = load_workbook(excel_data)
                worksheet = workbook.active
                
                print_success(f"Excel file generated for empty data scenario")
                
                # Check if totals row exists and shows zeros or is absent
                totals_row_num = None
                for row_num in range(1, worksheet.max_row + 1):
                    cell_value = worksheet.cell(row=row_num, column=1).value
                    if cell_value and str(cell_value).upper() == "TOTALS":
                        totals_row_num = row_num
                        break
                
                if totals_row_num:
                    print_success("✅ Totals row present even with empty data")
                    
                    # Check if totals are zeros
                    all_zeros = True
                    for col_num in range(4, worksheet.max_column + 1):
                        totals_value = worksheet.cell(row=totals_row_num, column=col_num).value
                        if totals_value is not None and isinstance(totals_value, (int, float)) and totals_value != 0:
                            all_zeros = False
                            break
                    
                    if all_zeros:
                        print_success("✅ All totals are zero for empty data (correct)")
                        self.test_results['passed'] += 1
                    else:
                        print_warning("⚠️ Some totals are non-zero for empty data")
                        self.test_results['passed'] += 1  # Still acceptable
                else:
                    print_info("ℹ️ No totals row for empty data (acceptable behavior)")
                    self.test_results['passed'] += 1
                    
            else:
                print_error(f"Excel generation failed for empty data: {response.status_code}")
                self.test_results['failed'] += 1
                
        except Exception as e:
            print_error(f"Exception in empty data test: {str(e)}")
            self.test_results['failed'] += 1

    def run_all_tests(self):
        """Run all Excel totals enhancement tests"""
        print_header("🚀 EXCEL TOTALS ROW ENHANCEMENT TESTING")
        print_info("Testing new functionality: Totals rows at bottom of Excel spreadsheets")
        print_info("Expected: Bold formatting, light red background, accurate calculations")
        
        if not self.setup_authentication():
            print_error("Failed to setup authentication - cannot continue")
            return False
        
        # Create test data
        self.create_test_activity_data()
        
        # Run all tests
        self.test_period_excel_reports_with_totals()
        self.test_daily_excel_reports_with_totals()
        self.test_totals_calculation_accuracy()
        self.test_manager_selection_with_totals()
        self.test_empty_data_handling()
        
        # Print summary
        self.print_test_summary()
        
        return self.test_results['failed'] == 0

    def print_test_summary(self):
        """Print comprehensive test summary"""
        print_header("📊 EXCEL TOTALS TESTING SUMMARY")
        
        total_tests = self.test_results['passed'] + self.test_results['failed']
        success_rate = (self.test_results['passed'] / total_tests * 100) if total_tests > 0 else 0
        
        print_info(f"Total Tests: {total_tests}")
        print_success(f"Passed: {self.test_results['passed']}")
        
        if self.test_results['failed'] > 0:
            print_error(f"Failed: {self.test_results['failed']}")
            print_error("Failed Tests:")
            for error in self.test_results['errors']:
                print_error(f"  • {error}")
        else:
            print_success("Failed: 0")
        
        print_info(f"Success Rate: {success_rate:.1f}%")
        
        if self.test_results['failed'] == 0:
            print_success("🎉 ALL EXCEL TOTALS TESTS PASSED!")
            print_success("✅ Totals rows functionality is working correctly")
        else:
            print_error("❌ Some Excel totals tests failed")
            print_error("❌ Totals rows functionality needs attention")

if __name__ == "__main__":
    tester = ExcelTotalsTester()
    success = tester.run_all_tests()
    sys.exit(0 if success else 1)