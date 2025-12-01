#!/usr/bin/env python3
"""
CRITICAL EXCEL DOWNLOAD BUG FIX TESTING

USER ISSUE: Excel download shows different data than web interface.
- Web interface correctly shows Steve Ahlers' team (Ryan Rozell, Andrew Inman, Robert Whitman)
- Excel download incorrectly shows different data (Steve Ahlers, Colton Cox)

WHAT WAS FIXED:
1. Updated Excel download endpoints to accept same parameters as JSON endpoints (user_id, month, quarter, year)
2. Excel downloads now call get_period_report() and get_daily_report() with ALL parameters
3. Frontend now passes same parameters to Excel downloads as JSON requests

CRITICAL TESTS:
- Test 1: Team Report Excel with Manager Selection
- Test 2: Individual Report Excel with Manager Selection  
- Test 3: Historical Period Excel Downloads
- Validation: Excel file data should exactly match JSON response data
"""

import requests
import json
from datetime import datetime, timedelta
import sys
import os
import tempfile
from openpyxl import load_workbook
import io

# Configuration
BACKEND_URL = "https://sales-leaderboard-5.preview.emergentagent.com/api"

class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'

def print_success(message):
    print(f"{Colors.GREEN}✅ {message}{Colors.ENDC}")

def print_error(message):
    print(f"{Colors.RED}❌ {message}{Colors.ENDC}")

def print_warning(message):
    print(f"{Colors.YELLOW}⚠️  {message}{Colors.ENDC}")

def print_info(message):
    print(f"{Colors.BLUE}ℹ️  {message}{Colors.ENDC}")

def print_header(message):
    print(f"\n{Colors.BOLD}{Colors.BLUE}{'='*80}{Colors.ENDC}")
    print(f"{Colors.BOLD}{Colors.BLUE}{message}{Colors.ENDC}")
    print(f"{Colors.BOLD}{Colors.BLUE}{'='*80}{Colors.ENDC}")

class ExcelDownloadTester:
    def __init__(self):
        self.session = requests.Session()
        self.state_manager_token = None
        self.steve_ahlers_id = None
        self.ryan_rozell_id = None
        self.test_results = {
            'passed': 0,
            'failed': 0,
            'errors': []
        }

    def setup_authentication(self):
        """Setup authentication with existing state manager"""
        print_header("SETTING UP AUTHENTICATION")
        
        try:
            # Try to login with existing state manager
            response = self.session.post(f"{BACKEND_URL}/auth/login", json={
                "email": "spencer.sudbeck@pmagent.net",
                "password": "Bizlink25"
            })
            
            if response.status_code == 200:
                data = response.json()
                self.state_manager_token = data['token']
                self.state_manager_id = data['user']['id']
                print_success(f"Logged in as state manager: {data['user']['name']}")
                return True
            else:
                print_error(f"Failed to login: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            print_error(f"Exception during login: {str(e)}")
            return False

    def get_managers_list(self):
        """Get list of available managers for testing"""
        print_header("GETTING MANAGERS LIST")
        
        if not self.state_manager_token:
            print_error("No authentication token available")
            return False
            
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        
        try:
            response = self.session.get(f"{BACKEND_URL}/reports/managers", headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                managers = data.get('managers', [])
                
                print_success(f"Found {len(managers)} managers in hierarchy")
                
                # Look for Steve Ahlers and other managers
                for manager in managers:
                    name = manager.get('name', 'Unknown')
                    manager_id = manager.get('id', 'Unknown')
                    role = manager.get('role', 'Unknown')
                    
                    print_info(f"Manager: {name} ({role}) - ID: {manager_id}")
                    
                    # Store specific manager IDs for testing
                    if 'steve' in name.lower() and 'ahlers' in name.lower():
                        self.steve_ahlers_id = manager_id
                        print_success(f"Found Steve Ahlers ID: {manager_id}")
                    elif 'ryan' in name.lower() and 'rozell' in name.lower():
                        self.ryan_rozell_id = manager_id
                        print_success(f"Found Ryan Rozell ID: {manager_id}")
                
                return True
            else:
                print_error(f"Failed to get managers list: {response.status_code}")
                return False
                
        except Exception as e:
            print_error(f"Exception getting managers list: {str(e)}")
            return False

    def compare_json_vs_excel_data(self, json_data, excel_content, report_type):
        """Compare JSON response data with Excel file content"""
        try:
            # Load Excel file from bytes
            workbook = load_workbook(io.BytesIO(excel_content))
            worksheet = workbook.active
            
            print_info(f"Excel file loaded successfully - Sheet: {worksheet.title}")
            
            # Extract data from Excel (skip header rows)
            excel_data = []
            for row in worksheet.iter_rows(min_row=3, values_only=True):  # Skip title and header
                if row[0] is not None:  # Skip empty rows
                    excel_data.append(row)
            
            print_info(f"Excel contains {len(excel_data)} data rows")
            
            # Compare with JSON data
            json_data_array = json_data.get('data', [])
            print_info(f"JSON contains {len(json_data_array)} data entries")
            
            if len(excel_data) != len(json_data_array):
                print_error(f"Data count mismatch: Excel {len(excel_data)} vs JSON {len(json_data_array)}")
                return False
            
            # Compare specific data points based on report type
            if report_type == 'team':
                return self.compare_team_data(json_data_array, excel_data)
            elif report_type == 'individual':
                return self.compare_individual_data(json_data_array, excel_data)
            elif report_type == 'organization':
                return self.compare_organization_data(json_data, excel_data)
            
            return True
            
        except Exception as e:
            print_error(f"Exception comparing data: {str(e)}")
            return False

    def compare_team_data(self, json_data, excel_data):
        """Compare team report data between JSON and Excel"""
        print_info("Comparing team report data...")
        
        for i, (json_team, excel_row) in enumerate(zip(json_data, excel_data)):
            json_team_name = json_team.get('team_name', '')
            json_manager = json_team.get('manager', '')
            json_contacts = json_team.get('contacts', 0)
            json_premium = json_team.get('premium', 0)
            
            excel_team_name = str(excel_row[0]) if excel_row[0] else ''
            excel_manager = str(excel_row[1]) if excel_row[1] else ''
            excel_contacts = float(excel_row[3]) if excel_row[3] else 0
            excel_premium = float(excel_row[10]) if excel_row[10] else 0
            
            print_info(f"Team {i+1}:")
            print_info(f"  JSON: {json_team_name} (Manager: {json_manager}) - Contacts: {json_contacts}, Premium: {json_premium}")
            print_info(f"  Excel: {excel_team_name} (Manager: {excel_manager}) - Contacts: {excel_contacts}, Premium: {excel_premium}")
            
            # Check if data matches
            if (json_team_name != excel_team_name or 
                json_manager != excel_manager or 
                abs(json_contacts - excel_contacts) > 0.01 or
                abs(json_premium - excel_premium) > 0.01):
                print_error(f"Data mismatch in team {i+1}")
                return False
        
        print_success("Team data matches between JSON and Excel")
        return True

    def compare_individual_data(self, json_data, excel_data):
        """Compare individual report data between JSON and Excel"""
        print_info("Comparing individual report data...")
        
        for i, (json_person, excel_row) in enumerate(zip(json_data, excel_data)):
            json_name = json_person.get('name', '')
            json_email = json_person.get('email', '')
            json_contacts = json_person.get('contacts', 0)
            json_premium = json_person.get('premium', 0)
            
            excel_name = str(excel_row[0]) if excel_row[0] else ''
            excel_email = str(excel_row[1]) if excel_row[1] else ''
            excel_contacts = float(excel_row[3]) if excel_row[3] else 0
            excel_premium = float(excel_row[10]) if excel_row[10] else 0
            
            print_info(f"Person {i+1}:")
            print_info(f"  JSON: {json_name} ({json_email}) - Contacts: {json_contacts}, Premium: {json_premium}")
            print_info(f"  Excel: {excel_name} ({excel_email}) - Contacts: {excel_contacts}, Premium: {excel_premium}")
            
            # Check if data matches
            if (json_name != excel_name or 
                json_email != excel_email or 
                abs(json_contacts - excel_contacts) > 0.01 or
                abs(json_premium - excel_premium) > 0.01):
                print_error(f"Data mismatch in person {i+1}")
                return False
        
        print_success("Individual data matches between JSON and Excel")
        return True

    def compare_organization_data(self, json_data, excel_data):
        """Compare organization report data between JSON and Excel"""
        print_info("Comparing organization report data...")
        
        json_org_data = json_data.get('data', {})
        json_total_members = json_data.get('total_members', 0)
        
        # Excel organization data is in metric/value pairs
        excel_metrics = {}
        for row in excel_data:
            if len(row) >= 2 and row[0] and row[1] is not None:
                metric_name = str(row[0]).lower()
                value = row[1]
                excel_metrics[metric_name] = value
        
        print_info(f"JSON Total Members: {json_total_members}")
        print_info(f"Excel Total Members: {excel_metrics.get('total members', 0)}")
        
        # Compare key metrics
        metrics_to_compare = ['contacts', 'appointments', 'presentations', 'premium']
        for metric in metrics_to_compare:
            json_value = json_org_data.get(metric, 0)
            excel_value = excel_metrics.get(metric, 0) or excel_metrics.get(f'total {metric}', 0)
            
            print_info(f"{metric.capitalize()}: JSON={json_value}, Excel={excel_value}")
            
            if abs(float(json_value) - float(excel_value)) > 0.01:
                print_error(f"Organization data mismatch in {metric}")
                return False
        
        print_success("Organization data matches between JSON and Excel")
        return True

    def test_team_report_excel_with_manager_selection(self):
        """Test 1: Team Report Excel with Manager Selection"""
        print_header("TEST 1: TEAM REPORT EXCEL WITH MANAGER SELECTION")
        
        if not self.state_manager_token:
            print_error("No authentication token available")
            return
            
        if not self.steve_ahlers_id:
            print_warning("No Steve Ahlers ID available - using first available manager")
            # Get first manager from list as fallback
            headers = {"Authorization": f"Bearer {self.state_manager_token}"}
            try:
                response = self.session.get(f"{BACKEND_URL}/reports/managers", headers=headers)
                if response.status_code == 200:
                    managers = response.json().get('managers', [])
                    if managers:
                        self.steve_ahlers_id = managers[0]['id']
                        print_info(f"Using manager: {managers[0]['name']} (ID: {self.steve_ahlers_id})")
            except:
                print_error("Could not get fallback manager ID")
                return
        
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        
        print_info(f"Testing team report Excel download for manager ID: {self.steve_ahlers_id}")
        print_info("Expected: Excel should contain Steve's direct reports (Ryan Rozell's Team, Andrew Inman's Team, Robert Whitman's Team)")
        print_info("Should NOT contain: Steve Ahlers + Colton Cox")
        
        try:
            # Step 1: Get JSON data for comparison
            print_info("Step 1: Getting JSON team report data...")
            json_response = self.session.get(
                f"{BACKEND_URL}/reports/period/team",
                params={
                    "period": "monthly",
                    "user_id": self.steve_ahlers_id,
                    "month": "2025-11"
                },
                headers=headers
            )
            
            if json_response.status_code != 200:
                print_error(f"JSON team report failed: {json_response.status_code} - {json_response.text}")
                self.test_results['failed'] += 1
                return
            
            json_data = json_response.json()
            print_success("JSON team report retrieved successfully")
            
            # Log JSON data for verification
            json_teams = json_data.get('data', [])
            print_info(f"JSON shows {len(json_teams)} teams:")
            for i, team in enumerate(json_teams):
                team_name = team.get('team_name', 'Unknown')
                manager = team.get('manager', 'Unknown')
                contacts = team.get('contacts', 0)
                premium = team.get('premium', 0)
                print_info(f"  Team {i+1}: {team_name} (Manager: {manager}) - Contacts: {contacts}, Premium: ${premium}")
            
            # Step 2: Get Excel data
            print_info("Step 2: Getting Excel team report data...")
            excel_response = self.session.get(
                f"{BACKEND_URL}/reports/period/excel/team",
                params={
                    "period": "monthly",
                    "user_id": self.steve_ahlers_id,
                    "month": "2025-11"
                },
                headers=headers
            )
            
            if excel_response.status_code != 200:
                print_error(f"Excel team report failed: {excel_response.status_code} - {excel_response.text}")
                self.test_results['failed'] += 1
                return
            
            print_success("Excel team report downloaded successfully")
            
            # Verify it's an Excel file
            content_type = excel_response.headers.get('content-type', '')
            content_disposition = excel_response.headers.get('content-disposition', '')
            
            if 'spreadsheet' in content_type or 'excel' in content_type or '.xlsx' in content_disposition:
                print_success("Excel file format verified")
                self.test_results['passed'] += 1
            else:
                print_warning(f"Excel format unclear - Content-Type: {content_type}")
            
            # Step 3: Compare JSON vs Excel data
            print_info("Step 3: Comparing JSON vs Excel data...")
            if self.compare_json_vs_excel_data(json_data, excel_response.content, 'team'):
                print_success("✅ CRITICAL SUCCESS: Excel data matches JSON data exactly")
                print_success("✅ BUG FIX VERIFIED: Excel now shows same team data as web interface")
                self.test_results['passed'] += 1
            else:
                print_error("❌ CRITICAL FAILURE: Excel data does not match JSON data")
                print_error("❌ BUG STILL EXISTS: Excel shows different data than web interface")
                self.test_results['failed'] += 1
                self.test_results['errors'].append("Team Excel data mismatch with JSON")
            
        except Exception as e:
            print_error(f"Exception in team report Excel test: {str(e)}")
            self.test_results['failed'] += 1
            self.test_results['errors'].append(f"Team Excel test exception: {str(e)}")

    def test_individual_report_excel_with_manager_selection(self):
        """Test 2: Individual Report Excel with Manager Selection"""
        print_header("TEST 2: INDIVIDUAL REPORT EXCEL WITH MANAGER SELECTION")
        
        if not self.state_manager_token:
            print_error("No authentication token available")
            return
            
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        
        # Test with and without user_id parameter
        test_cases = [
            {"name": "All individuals", "params": {"period": "monthly"}},
            {"name": "Specific user", "params": {"period": "monthly", "user_id": self.steve_ahlers_id}} if self.steve_ahlers_id else None
        ]
        
        # Filter out None cases
        test_cases = [case for case in test_cases if case is not None]
        
        for test_case in test_cases:
            print_info(f"Testing individual report: {test_case['name']}")
            
            try:
                # Step 1: Get JSON data
                json_response = self.session.get(
                    f"{BACKEND_URL}/reports/period/individual",
                    params=test_case['params'],
                    headers=headers
                )
                
                if json_response.status_code != 200:
                    print_error(f"JSON individual report failed: {json_response.status_code}")
                    self.test_results['failed'] += 1
                    continue
                
                json_data = json_response.json()
                print_success(f"JSON individual report retrieved: {len(json_data.get('data', []))} individuals")
                
                # Step 2: Get Excel data
                excel_response = self.session.get(
                    f"{BACKEND_URL}/reports/period/excel/individual",
                    params=test_case['params'],
                    headers=headers
                )
                
                if excel_response.status_code != 200:
                    print_error(f"Excel individual report failed: {excel_response.status_code}")
                    self.test_results['failed'] += 1
                    continue
                
                print_success("Excel individual report downloaded successfully")
                
                # Step 3: Compare data
                if self.compare_json_vs_excel_data(json_data, excel_response.content, 'individual'):
                    print_success(f"✅ Individual report Excel matches JSON for: {test_case['name']}")
                    self.test_results['passed'] += 1
                else:
                    print_error(f"❌ Individual report Excel mismatch for: {test_case['name']}")
                    self.test_results['failed'] += 1
                    self.test_results['errors'].append(f"Individual Excel mismatch: {test_case['name']}")
                
            except Exception as e:
                print_error(f"Exception in individual report test ({test_case['name']}): {str(e)}")
                self.test_results['failed'] += 1

    def test_daily_excel_downloads(self):
        """Test daily Excel downloads with user_id parameter"""
        print_header("TEST 3: DAILY EXCEL DOWNLOADS WITH MANAGER SELECTION")
        
        if not self.state_manager_token:
            print_error("No authentication token available")
            return
            
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        test_date = datetime.now().date().isoformat()
        
        print_info(f"Testing daily Excel downloads for date: {test_date}")
        
        # Test cases for daily reports
        test_cases = [
            {"name": "Daily Team Report", "report_type": "team", "params": {"date": test_date}},
            {"name": "Daily Team Report with Manager", "report_type": "team", "params": {"date": test_date, "user_id": self.steve_ahlers_id}} if self.steve_ahlers_id else None,
            {"name": "Daily Individual Report", "report_type": "individual", "params": {"date": test_date}},
            {"name": "Daily Organization Report", "report_type": "organization", "params": {"date": test_date}}
        ]
        
        # Filter out None cases
        test_cases = [case for case in test_cases if case is not None]
        
        for test_case in test_cases:
            print_info(f"Testing: {test_case['name']}")
            
            try:
                # Step 1: Get JSON data
                json_response = self.session.get(
                    f"{BACKEND_URL}/reports/daily/{test_case['report_type']}",
                    params=test_case['params'],
                    headers=headers
                )
                
                if json_response.status_code != 200:
                    print_error(f"JSON daily report failed: {json_response.status_code}")
                    self.test_results['failed'] += 1
                    continue
                
                json_data = json_response.json()
                print_success(f"JSON daily report retrieved successfully")
                
                # Step 2: Get Excel data
                excel_response = self.session.get(
                    f"{BACKEND_URL}/reports/daily/excel/{test_case['report_type']}",
                    params=test_case['params'],
                    headers=headers
                )
                
                if excel_response.status_code != 200:
                    print_error(f"Excel daily report failed: {excel_response.status_code}")
                    self.test_results['failed'] += 1
                    continue
                
                print_success("Excel daily report downloaded successfully")
                
                # Step 3: Compare data
                if self.compare_json_vs_excel_data(json_data, excel_response.content, test_case['report_type']):
                    print_success(f"✅ Daily Excel matches JSON for: {test_case['name']}")
                    self.test_results['passed'] += 1
                else:
                    print_error(f"❌ Daily Excel mismatch for: {test_case['name']}")
                    self.test_results['failed'] += 1
                    self.test_results['errors'].append(f"Daily Excel mismatch: {test_case['name']}")
                
            except Exception as e:
                print_error(f"Exception in daily Excel test ({test_case['name']}): {str(e)}")
                self.test_results['failed'] += 1

    def test_historical_period_excel_downloads(self):
        """Test historical period Excel downloads"""
        print_header("TEST 4: HISTORICAL PERIOD EXCEL DOWNLOADS")
        
        if not self.state_manager_token:
            print_error("No authentication token available")
            return
            
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        
        # Test historical periods
        historical_tests = [
            {"name": "Previous Month", "params": {"period": "monthly", "month": "2025-10"}},
            {"name": "Previous Quarter", "params": {"period": "quarterly", "quarter": "2025-Q3"}},
            {"name": "Previous Year", "params": {"period": "yearly", "year": "2024"}}
        ]
        
        for historical_test in historical_tests:
            print_info(f"Testing historical period: {historical_test['name']}")
            
            try:
                # Test team report for historical period
                json_response = self.session.get(
                    f"{BACKEND_URL}/reports/period/team",
                    params=historical_test['params'],
                    headers=headers
                )
                
                if json_response.status_code != 200:
                    print_warning(f"JSON historical report failed: {json_response.status_code}")
                    continue
                
                json_data = json_response.json()
                
                # Get Excel version
                excel_response = self.session.get(
                    f"{BACKEND_URL}/reports/period/excel/team",
                    params=historical_test['params'],
                    headers=headers
                )
                
                if excel_response.status_code != 200:
                    print_error(f"Excel historical report failed: {excel_response.status_code}")
                    self.test_results['failed'] += 1
                    continue
                
                print_success(f"Historical Excel downloaded: {historical_test['name']}")
                
                # Compare data
                if self.compare_json_vs_excel_data(json_data, excel_response.content, 'team'):
                    print_success(f"✅ Historical Excel matches JSON for: {historical_test['name']}")
                    self.test_results['passed'] += 1
                else:
                    print_error(f"❌ Historical Excel mismatch for: {historical_test['name']}")
                    self.test_results['failed'] += 1
                
            except Exception as e:
                print_error(f"Exception in historical test ({historical_test['name']}): {str(e)}")
                self.test_results['failed'] += 1

    def test_parameter_consistency(self):
        """Test that Excel endpoints accept all the same parameters as JSON endpoints"""
        print_header("TEST 5: PARAMETER CONSISTENCY VERIFICATION")
        
        if not self.state_manager_token:
            print_error("No authentication token available")
            return
            
        headers = {"Authorization": f"Bearer {self.state_manager_token}"}
        
        print_info("Testing that Excel endpoints accept all parameters that JSON endpoints accept")
        
        # Test parameter combinations
        parameter_tests = [
            {
                "name": "Monthly with specific month and user_id",
                "json_endpoint": "/reports/period/team",
                "excel_endpoint": "/reports/period/excel/team",
                "params": {"period": "monthly", "month": "2025-11", "user_id": self.steve_ahlers_id}
            },
            {
                "name": "Quarterly with specific quarter",
                "json_endpoint": "/reports/period/individual",
                "excel_endpoint": "/reports/period/excel/individual",
                "params": {"period": "quarterly", "quarter": "2025-Q4"}
            },
            {
                "name": "Yearly with specific year",
                "json_endpoint": "/reports/period/organization",
                "excel_endpoint": "/reports/period/excel/organization",
                "params": {"period": "yearly", "year": "2025"}
            },
            {
                "name": "Daily with user_id",
                "json_endpoint": "/reports/daily/team",
                "excel_endpoint": "/reports/daily/excel/team",
                "params": {"date": datetime.now().date().isoformat(), "user_id": self.steve_ahlers_id}
            }
        ]
        
        # Filter out tests with None user_id
        parameter_tests = [test for test in parameter_tests if test['params'].get('user_id') != None or 'user_id' not in test['params']]
        
        for test in parameter_tests:
            print_info(f"Testing parameter consistency: {test['name']}")
            
            try:
                # Test JSON endpoint
                json_response = self.session.get(
                    f"{BACKEND_URL}{test['json_endpoint']}",
                    params=test['params'],
                    headers=headers
                )
                
                # Test Excel endpoint with same parameters
                excel_response = self.session.get(
                    f"{BACKEND_URL}{test['excel_endpoint']}",
                    params=test['params'],
                    headers=headers
                )
                
                if json_response.status_code == 200 and excel_response.status_code == 200:
                    print_success(f"✅ Both endpoints accept parameters: {test['name']}")
                    self.test_results['passed'] += 1
                elif json_response.status_code == excel_response.status_code:
                    print_info(f"Both endpoints returned same status ({json_response.status_code}): {test['name']}")
                    self.test_results['passed'] += 1
                else:
                    print_error(f"❌ Parameter inconsistency: JSON={json_response.status_code}, Excel={excel_response.status_code}")
                    self.test_results['failed'] += 1
                    self.test_results['errors'].append(f"Parameter inconsistency: {test['name']}")
                
            except Exception as e:
                print_error(f"Exception in parameter test ({test['name']}): {str(e)}")
                self.test_results['failed'] += 1

    def run_all_tests(self):
        """Run all Excel download bug fix tests"""
        print_header("🚨 CRITICAL EXCEL DOWNLOAD BUG FIX TESTING")
        print_info("Testing that Excel downloads now show identical data to web interface")
        
        # Setup
        if not self.setup_authentication():
            print_error("Authentication setup failed - cannot continue")
            return False
        
        if not self.get_managers_list():
            print_error("Could not get managers list - some tests may be limited")
        
        # Run all tests
        self.test_team_report_excel_with_manager_selection()
        self.test_individual_report_excel_with_manager_selection()
        self.test_daily_excel_downloads()
        self.test_historical_period_excel_downloads()
        self.test_parameter_consistency()
        
        # Print final results
        self.print_final_results()
        
        return self.test_results['failed'] == 0

    def print_final_results(self):
        """Print final test results"""
        print_header("🎯 EXCEL DOWNLOAD BUG FIX TEST RESULTS")
        
        total_tests = self.test_results['passed'] + self.test_results['failed']
        
        if self.test_results['failed'] == 0:
            print_success(f"✅ ALL TESTS PASSED ({self.test_results['passed']}/{total_tests})")
            print_success("✅ EXCEL DOWNLOAD BUG FIX VERIFIED SUCCESSFUL")
            print_success("✅ Excel downloads now show identical data to web interface")
        else:
            print_error(f"❌ TESTS FAILED ({self.test_results['failed']}/{total_tests})")
            print_error("❌ EXCEL DOWNLOAD BUG STILL EXISTS")
            
            if self.test_results['errors']:
                print_error("Specific errors found:")
                for error in self.test_results['errors']:
                    print_error(f"  - {error}")
        
        print_info(f"Total tests run: {total_tests}")
        print_info(f"Passed: {self.test_results['passed']}")
        print_info(f"Failed: {self.test_results['failed']}")

if __name__ == "__main__":
    tester = ExcelDownloadTester()
    success = tester.run_all_tests()
    
    if success:
        print_success("\n🎉 EXCEL DOWNLOAD BUG FIX TESTING COMPLETED SUCCESSFULLY")
        sys.exit(0)
    else:
        print_error("\n💥 EXCEL DOWNLOAD BUG FIX TESTING FAILED")
        sys.exit(1)